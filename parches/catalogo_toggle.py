import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

import os
import re
import json
import difflib
from datetime import datetime
import unicodedata
import subprocess
import sys
import shutil
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    XLImage = None

try:
    from PIL import Image as PILImage
    from PIL import ImageTk
except Exception:
    PILImage = None
    ImageTk = None


PRE_REGEX_ALL = re.compile(
    r'(<pre\b[^>]*\bid\s*=\s*["\']productos-tsv["\'][^>]*>)([\s\S]*?)(</pre>)',
    flags=re.IGNORECASE
)

HTML_COMMENT_REGEX = re.compile(r"<!--[\s\S]*?-->", flags=re.MULTILINE)

EXPECTED_COLS = ["Codigo", "Nombre producto", "Categoria", "Marca", "Valor unitario", "Stock"]

EXCEL_BASE_NAME = "productos"
EXCEL_SHEET_NAME = "Catalogo"


def get_settings_path() -> str:
    appdata = os.getenv("APPDATA")
    if appdata:
        folder = os.path.join(appdata, "CatalogoToggle")
    else:
        folder = os.path.join(os.path.expanduser("~"), ".catalogo_toggle")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "settings.json")


def load_settings() -> dict:
    path = get_settings_path()
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
    except Exception:
        pass
    return {}


def save_settings(data: dict) -> None:
    path = get_settings_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def remember_paths(settings: dict, source_html: str = "", excel_path: str = "", images_dir: str = "") -> dict:
    settings = settings or {}
    if source_html:
        settings["last_source"] = os.path.abspath(source_html)
        settings["last_dir"] = os.path.dirname(os.path.abspath(source_html))
    if excel_path:
        settings["last_excel"] = os.path.abspath(excel_path)
        settings["last_dir"] = os.path.dirname(os.path.abspath(excel_path))
    if images_dir:
        settings["last_images_dir"] = os.path.abspath(images_dir)
        settings["last_dir"] = os.path.abspath(images_dir)
    save_settings(settings)
    return settings


def open_with_default_app(path: str) -> None:
    path = (path or "").strip()
    if not path:
        raise RuntimeError("No hay ruta para abrir.")
    if not os.path.exists(path):
        raise RuntimeError("La ruta no existe.")

    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
            return
        subprocess.run(["xdg-open", path], check=False)
    except Exception as e:
        raise RuntimeError(f"No se pudo abrir. Detalle: {e}")


def open_folder_of(path: str) -> None:
    path = (path or "").strip()
    if not path:
        raise RuntimeError("No hay ruta.")
    folder = path if os.path.isdir(path) else os.path.dirname(path)
    if not folder or not os.path.isdir(folder):
        raise RuntimeError("No se pudo determinar la carpeta.")
    open_with_default_app(folder)


def norm_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s


def split_tsv_line(line: str):
    return line.rstrip("\r\n").split("\t")


def _digits_only(x: str) -> str:
    return re.sub(r"[^\d]", "", str(x or "").strip())


def _int_from_digits(x: str):
    d = _digits_only(x)
    return int(d) if d.isdigit() else None


def looks_like_stock(x: str) -> bool:
    raw = str(x or "").strip()
    if raw == "":
        return False
    d = _digits_only(raw)
    return d.isdigit() and int(d) >= 0


def _clean_number_text(s: str) -> str:
    s = str(s or "").strip()
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[^\d,.\-]", "", s)
    s = s.strip()
    return s


def _decimal_from_text(s: str):
    s = _clean_number_text(s)
    if not s:
        return None

    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:].strip()

    if not s:
        return None

    if "," in s and "." in s:
        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_comma > last_dot:
            s2 = s.replace(".", "")
            s2 = s2.replace(",", ".")
        else:
            s2 = s.replace(",", "")
    elif "," in s and "." not in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            s2 = parts[0].replace(".", "") + "." + parts[1]
        else:
            s2 = s.replace(",", "")
    else:
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            s2 = parts[0].replace(",", "") + "." + parts[1]
        else:
            s2 = s.replace(".", "").replace(",", "")

    try:
        d = Decimal(s2)
        if neg:
            d = -d
        return d
    except (InvalidOperation, ValueError):
        return None


def looks_like_valor_unitario(x: str) -> bool:
    d = _decimal_from_text(x)
    return d is not None


def _rows_equal_for_diff(before6: list, after6: list) -> bool:
    b = (before6 + [""] * 6)[:6]
    a = (after6 + [""] * 6)[:6]

    if str(b[0] or "").strip() != str(a[0] or "").strip():
        return False
    if str(b[1] or "").strip() != str(a[1] or "").strip():
        return False
    if str(b[2] or "").strip() != str(a[2] or "").strip():
        return False
    if str(b[3] or "").strip() != str(a[3] or "").strip():
        return False

    b_val = _decimal_from_text(b[4])
    a_val = _decimal_from_text(a[4])
    if b_val != a_val:
        return False

    if _int_from_digits(b[5]) != _int_from_digits(a[5]):
        return False

    return True


def read_text_preserve_bom(path: str):
    data = open(path, "rb").read()
    had_bom = data.startswith(b"\xef\xbb\xbf")
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(enc)
            return text, had_bom, enc
        except UnicodeDecodeError:
            continue
    text = data.decode("utf-8", errors="replace")
    return text, had_bom, "utf-8"


def write_text_preserve_bom(path: str, text: str, had_bom: bool, _encoding_hint: str):
    payload = text.encode("utf-8")
    if had_bom and not payload.startswith(b"\xef\xbb\xbf"):
        payload = b"\xef\xbb\xbf" + payload
    with open(path, "wb") as f:
        f.write(payload)


def find_single_pre_block(html: str):
    matches = list(PRE_REGEX_ALL.finditer(html))
    if not matches:
        return None, 'No encontré <pre id="productos-tsv">...</pre>.'
    if len(matches) > 1:
        return None, 'Encontré más de un <pre id="productos-tsv">. Para seguridad, no haré cambios.'
    m = matches[0]
    open_tag, inner, close_tag = m.group(1), m.group(2), m.group(3)
    inner_start = m.start(2)
    inner_end = m.end(2)
    return (open_tag, inner, close_tag, inner_start, inner_end), None


def validate_tsv_structure(inner: str):
    issues = []
    cleaned = HTML_COMMENT_REGEX.sub("", inner)

    if "<" in cleaned or ">" in cleaned:
        issues.append("El <pre> contiene '<' o '>' (parece HTML incrustado). Por seguridad no se modifica.")

    lines = cleaned.splitlines(True)
    i = 0
    while i < len(lines) and lines[i].strip() == "":
        i += 1
    if i >= len(lines):
        issues.append("El <pre> está vacío.")
        return {"ok": False, "issues": issues}

    header_line = lines[i]
    header_cols = split_tsv_line(header_line)
    if len(header_cols) < 6:
        issues.append("El encabezado no tiene al menos 6 columnas separadas por TAB.")

    norm_cols = [norm_text(c) for c in header_cols[:6]]
    expected = [norm_text(c) for c in EXPECTED_COLS]
    if norm_cols != expected:
        issues.append(
            "El encabezado TSV no coincide con el esperado: "
            "\"Codigo\\tNombre producto\\tCategoria\\tMarca\\tValor unitario\\tStock\"."
        )

    j = i + 1
    while j < len(lines) and lines[j].strip() == "":
        j += 1
    if j >= len(lines):
        issues.append("No hay ningún producto después del encabezado.")
        return {"ok": False, "issues": issues, "header_idx": i}

    first_product_line = lines[j]
    fp_cols = split_tsv_line(first_product_line)
    if len(fp_cols) < 6:
        issues.append("La primera fila de producto no tiene al menos 6 columnas TSV.")
    else:
        if not str(fp_cols[0] or "").strip():
            issues.append("El 'Codigo' del primer producto está vacío.")
        if not looks_like_valor_unitario(fp_cols[4]):
            issues.append("El 'Valor unitario' del primer producto no parece un número válido.")
        if not looks_like_stock(fp_cols[5]):
            issues.append("El 'Stock' del primer producto no parece un número válido.")

    checked = 0
    for k in range(j, len(lines)):
        if checked >= 50:
            break
        ln = lines[k]
        if not ln.strip():
            continue
        cols = split_tsv_line(ln)
        if len(cols) < 6:
            issues.append(f"Hay una fila con menos de 6 columnas TSV (línea aprox #{k+1}).")
            break
        if not str(cols[0] or "").strip():
            issues.append(f"Hay una fila con 'Codigo' vacío (línea aprox #{k+1}).")
            break
        if not looks_like_valor_unitario(cols[4]):
            issues.append(f"Hay una fila con 'Valor unitario' no válido (línea aprox #{k+1}).")
            break
        if not looks_like_stock(cols[5]):
            issues.append(f"Hay una fila con 'Stock' no válido (línea aprox #{k+1}).")
            break
        checked += 1

    ok = len(issues) == 0
    return {
        "ok": ok,
        "issues": issues,
        "header_idx": i,
        "first_product_idx": j,
        "header_line": header_line.rstrip("\r\n"),
        "first_product_line": first_product_line.rstrip("\r\n"),
        "cleaned_inner": cleaned,
    }


def _detect_newline_style(text: str) -> str:
    return "\r\n" if "\r\n" in text else "\n"


def _tsv_line(cols: list[str], nl: str) -> str:
    return "\t".join([str(x) for x in cols]).rstrip("\r\n") + nl


def _parse_catalog_from_cleaned_inner(cleaned_inner: str):
    lines = [ln for ln in cleaned_inner.splitlines() if ln.strip()]
    if not lines:
        raise RuntimeError("El catálogo está vacío.")
    header_cols = split_tsv_line(lines[0])
    rows = [split_tsv_line(ln) for ln in lines[1:]]
    return header_cols, rows


def program_root_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def default_excel_path() -> str:
    return os.path.join(program_root_dir(), f"{EXCEL_BASE_NAME}.xlsx")


def make_next_excel_path(base_path: str):
    base_path = os.path.abspath(base_path)
    if not os.path.exists(base_path):
        return base_path, False

    folder = os.path.dirname(base_path)
    filename = os.path.basename(base_path)
    stem, ext = os.path.splitext(filename)
    if not ext:
        ext = ".xlsx"

    n = 1
    while True:
        candidate = os.path.join(folder, f"{stem}_{n}{ext}")
        if not os.path.exists(candidate):
            return candidate, True
        n += 1


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _excel_write_header_style(ws, ncols: int):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = center


def _autosize_columns(ws, max_row: int, max_col: int, min_w=10, max_w=60, skip_cols=None):
    skip_cols = set(skip_cols or [])
    for col in range(1, max_col + 1):
        if col in skip_cols:
            continue
        letter = get_column_letter(col)
        width = min_w
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            v = str(v)
            width = max(width, len(v) + 2)
        width = min(max_w, max(min_w, width))
        ws.column_dimensions[letter].width = width


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
CM_TO_INCH = 1.0 / 2.54
EXCEL_DPI = 96
IMG_SIZE_CM = 5.0
IMG_SIZE_PX = int(round(IMG_SIZE_CM * CM_TO_INCH * EXCEL_DPI))
ROW_HEIGHT_PT = IMG_SIZE_CM * CM_TO_INCH * 72
TREE_ROW_HEIGHT_PX = IMG_SIZE_PX


def parse_product_from_image_filename(filename: str):
    base = os.path.basename(filename)
    name_no_ext, _ = os.path.splitext(base)
    parts = name_no_ext.split("_")

    codigo = ""
    nombre = ""
    categoria = ""
    marca = ""
    valor_unitario = ""
    stock = ""

    if len(parts) >= 6:
        codigo = parts[0].strip()
        stock = parts[-1].strip()
        valor_unitario = parts[-2].strip()
        marca = parts[-3].strip()
        categoria = parts[-4].strip()
        nombre = "_".join(parts[1:-4]).strip()
    else:
        parts = parts + [""] * (6 - len(parts))
        codigo = parts[0].strip()
        nombre = parts[1].strip()
        categoria = parts[2].strip()
        marca = parts[3].strip()
        valor_unitario = parts[4].strip()
        stock = parts[5].strip()

    return {
        "imagen": base,
        "codigo": codigo,
        "nombre": nombre,
        "categoria": categoria,
        "marca": marca,
        "valor_unitario": valor_unitario,
        "stock": stock,
        "path": filename,
    }


def load_products_from_folder(images_dir: str):
    products = []
    if not images_dir or not os.path.isdir(images_dir):
        return products

    for fn in sorted(os.listdir(images_dir)):
        p = os.path.join(images_dir, fn)
        if not os.path.isfile(p):
            continue
        ext = os.path.splitext(fn)[1].lower()
        if ext in IMAGE_EXTS:
            products.append(parse_product_from_image_filename(p))
    return products


def _find_image_for_codigo(images_dir: str, codigo: str):
    if not images_dir or not os.path.isdir(images_dir) or not codigo:
        return ""
    try:
        pref = f"{codigo}_"
        for fn in os.listdir(images_dir):
            ext = os.path.splitext(fn)[1].lower()
            if ext not in IMAGE_EXTS:
                continue
            if fn.startswith(pref):
                return os.path.join(images_dir, fn)
    except Exception:
        return ""
    return ""


def export_products_to_excel(products: list[dict], excel_path: str, images_dir: str = ""):
    if XLImage is None:
        raise RuntimeError("openpyxl no pudo cargar soporte de imágenes. Instala Pillow (pip install pillow).")

    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    headers = ["Codigo", "Nombre producto", "Categoria", "Marca", "Valor unitario", "Stock", "Imagen"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    _excel_write_header_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(products) + 1)}"

    ws.column_dimensions["G"].width = 30
    ws.row_dimensions[1].height = 18

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for idx, prod in enumerate(products, start=2):
        codigo = str(prod.get("codigo", "")).strip()
        nombre = str(prod.get("nombre", "")).strip()
        categoria = str(prod.get("categoria", "")).strip()
        marca = str(prod.get("marca", "")).strip()
        valor = str(prod.get("valor_unitario", "")).strip()
        stock = str(prod.get("stock", "")).strip()
        imagen = str(prod.get("imagen", "")).strip()

        ws.row_dimensions[idx].height = ROW_HEIGHT_PT

        ws.cell(row=idx, column=1, value=codigo).alignment = center
        ws.cell(row=idx, column=2, value=nombre).alignment = left
        ws.cell(row=idx, column=3, value=categoria).alignment = left
        ws.cell(row=idx, column=4, value=marca).alignment = left

        dval = _decimal_from_text(valor)
        if dval is not None:
            if dval == dval.to_integral_value():
                ws.cell(row=idx, column=5, value=int(dval)).alignment = right
            else:
                ws.cell(row=idx, column=5, value=float(dval)).alignment = right
        else:
            ws.cell(row=idx, column=5, value=valor).alignment = right

        stock_digits = _digits_only(stock)
        if stock_digits.isdigit():
            ws.cell(row=idx, column=6, value=int(stock_digits)).alignment = center
        else:
            ws.cell(row=idx, column=6, value=stock).alignment = center

        ws.cell(row=idx, column=7, value=imagen).alignment = center

        img_path = prod.get("path", "")
        if not img_path and images_dir and imagen:
            img_path = os.path.join(images_dir, imagen)
        if img_path and os.path.exists(img_path):
            try:
                xl_img = XLImage(img_path)
                xl_img.width = IMG_SIZE_PX
                xl_img.height = IMG_SIZE_PX
                xl_img.anchor = f"G{idx}"
                ws.add_image(xl_img)
            except Exception:
                pass

    _autosize_columns(ws, max_row=max(2, len(products) + 1), max_col=len(headers), skip_cols=[7])
    wb.save(excel_path)
    return excel_path


def read_catalog_excel_flexible(excel_path: str):
    if not os.path.exists(excel_path):
        raise RuntimeError("No existe el Excel seleccionado.")

    wb = load_workbook(excel_path)
    ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb.active

    header_map = {}
    header_dups = set()
    for c in range(1, ws.max_column + 1):
        h = _safe_str(ws.cell(row=1, column=c).value)
        if not h:
            continue
        key = norm_text(h)
        if key in header_map:
            header_dups.add(key)
        else:
            header_map[key] = c

    if header_dups:
        req_dups = [k for k in header_dups if k in [norm_text(x) for x in EXPECTED_COLS]]
        if req_dups:
            raise RuntimeError("El Excel tiene encabezados duplicados en columnas requeridas. Corrige eso antes de actualizar.")

    missing = []
    col_idx = []
    for colname in EXPECTED_COLS:
        k = norm_text(colname)
        if k not in header_map:
            missing.append(colname)
        else:
            col_idx.append(header_map[k])

    if missing:
        raise RuntimeError("Faltan encabezados requeridos en el Excel: " + ", ".join(missing))

    idx_cod, idx_nombre, idx_cat, idx_marca, idx_valor, idx_stock = col_idx

    rows_by_codigo = {}
    order_codigos = []
    duplicates = []

    for r in range(2, ws.max_row + 1):
        codigo = _safe_str(ws.cell(row=r, column=idx_cod).value)
        nombre = _safe_str(ws.cell(row=r, column=idx_nombre).value)
        categoria = _safe_str(ws.cell(row=r, column=idx_cat).value)
        marca = _safe_str(ws.cell(row=r, column=idx_marca).value)
        valor = _safe_str(ws.cell(row=r, column=idx_valor).value)
        stock = _safe_str(ws.cell(row=r, column=idx_stock).value)

        if not any([codigo, nombre, categoria, marca, valor, stock]):
            continue

        if not codigo:
            raise RuntimeError(f"Hay una fila en Excel sin 'Codigo' (fila {r}).")

        if not looks_like_valor_unitario(valor):
            raise RuntimeError(f"En el Excel, 'Valor unitario' no parece válido en la fila {r}: '{valor}'")

        if not looks_like_stock(stock):
            raise RuntimeError(f"En el Excel, 'Stock' no parece válido en la fila {r}: '{stock}'")

        row6 = [codigo, nombre, categoria, marca, valor, stock]

        if codigo in rows_by_codigo:
            duplicates.append(codigo)
        else:
            rows_by_codigo[codigo] = row6
            order_codigos.append(codigo)

    if duplicates:
        dup_list = ", ".join(duplicates[:20]) + (" ..." if len(duplicates) > 20 else "")
        raise RuntimeError(f"El Excel tiene codigos duplicados. Ejemplos: {dup_list}")

    return rows_by_codigo, order_codigos


def export_catalog_html_to_excel(source_html: str, excel_path: str, images_dir: str = ""):
    excel_path, existed = make_next_excel_path(excel_path)

    html, _, _ = read_text_preserve_bom(source_html)
    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, _, _ = block
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se exporta.\n- " + "\n- ".join(val["issues"]))

    header_cols, rows = _parse_catalog_from_cleaned_inner(val["cleaned_inner"])
    if len(header_cols) < 6:
        raise RuntimeError("El encabezado TSV tiene menos de 6 columnas; no se exporta.")
    if [norm_text(c) for c in header_cols[:6]] != [norm_text(c) for c in EXPECTED_COLS]:
        raise RuntimeError("El encabezado TSV no coincide con el esperado; no se exporta.")

    header6 = header_cols[:6]
    rows6 = [r[:6] + ([""] * (6 - len(r))) if len(r) < 6 else r[:6] for r in rows]

    if XLImage is None:
        raise RuntimeError("openpyxl no pudo cargar soporte de imágenes. Instala Pillow (pip install pillow).")

    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    headers = ["Codigo", "Nombre producto", "Categoria", "Marca", "Valor unitario", "Stock", "Imagen"]
    for c, name in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=name)

    _excel_write_header_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows6) + 1)}"

    ws.column_dimensions["G"].width = 30
    ws.row_dimensions[1].height = 18

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows6, start=2):
        codigo = str(row[0] or "").strip()
        nombre = str(row[1] or "").strip()
        categoria = str(row[2] or "").strip()
        marca = str(row[3] or "").strip()
        valor = str(row[4] or "").strip()
        stock = str(row[5] or "").strip()

        ws.row_dimensions[r_idx].height = ROW_HEIGHT_PT

        ws.cell(row=r_idx, column=1, value=codigo).alignment = center
        ws.cell(row=r_idx, column=2, value=nombre).alignment = left
        ws.cell(row=r_idx, column=3, value=categoria).alignment = left
        ws.cell(row=r_idx, column=4, value=marca).alignment = left

        dval = _decimal_from_text(valor)
        if dval is not None:
            if dval == dval.to_integral_value():
                ws.cell(row=r_idx, column=5, value=int(dval)).alignment = right
            else:
                ws.cell(row=r_idx, column=5, value=float(dval)).alignment = right
        else:
            ws.cell(row=r_idx, column=5, value=valor).alignment = right

        stock_digits = _digits_only(stock)
        if stock_digits.isdigit():
            ws.cell(row=r_idx, column=6, value=int(stock_digits)).alignment = center
        else:
            ws.cell(row=r_idx, column=6, value=stock).alignment = center

        img_path = _find_image_for_codigo(images_dir, codigo) if images_dir else ""
        ws.cell(row=r_idx, column=7, value=os.path.basename(img_path) if img_path else "").alignment = center
        if img_path and os.path.exists(img_path):
            try:
                xl_img = XLImage(img_path)
                xl_img.width = IMG_SIZE_PX
                xl_img.height = IMG_SIZE_PX
                xl_img.anchor = f"G{r_idx}"
                ws.add_image(xl_img)
            except Exception:
                pass

    _autosize_columns(ws, max_row=max(2, len(rows6) + 1), max_col=len(headers), skip_cols=[7])
    wb.save(excel_path)
    return excel_path, len(rows6), existed


def _parse_html_catalog(inner: str):
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se puede actualizar.\n- " + "\n- ".join(val["issues"]))
    header_cols, rows = _parse_catalog_from_cleaned_inner(val["cleaned_inner"])
    if len(header_cols) < 6:
        raise RuntimeError("El HTML fuente no tiene al menos 6 columnas en el encabezado.")
    if [norm_text(c) for c in header_cols[:6]] != [norm_text(c) for c in EXPECTED_COLS]:
        raise RuntimeError("El encabezado del HTML fuente no coincide con el esperado.")
    html_map = {}
    dups = []
    for r in rows:
        if len(r) < 6:
            continue
        codigo = str(r[0] or "").strip()
        if not codigo:
            continue
        if codigo in html_map:
            dups.append(codigo)
        else:
            html_map[codigo] = r
    if dups:
        raise RuntimeError("El HTML fuente tiene codigos duplicados. No se actualiza por seguridad.")
    return header_cols, rows, html_map


def compute_diffs_and_build_new_html_exact(source_html: str, rows_by_codigo: dict, order_codigos: list):
    html, had_bom, enc = read_text_preserve_bom(source_html)
    nl = _detect_newline_style(html)

    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, inner_start, inner_end = block
    header_cols, html_rows, html_map = _parse_html_catalog(inner)

    html_codigos = set(html_map.keys())
    excel_codigos = set(rows_by_codigo.keys())

    diffs = []
    modified = 0
    added = 0
    deleted = 0

    def _codigo_sort_key(x: str):
        d = _digits_only(x)
        return (0, int(d)) if d.isdigit() else (1, x)

    for codigo in sorted(html_codigos - excel_codigos, key=_codigo_sort_key):
        before6 = (html_map[codigo] + [""] * 6)[:6]
        diffs.append({
            "type": "ELIMINADO",
            "id": codigo,
            "before": "\t".join([str(x) for x in before6]),
            "after": "(se elimina del HTML)",
        })
        deleted += 1

    for codigo in order_codigos:
        row6 = rows_by_codigo.get(codigo)
        if not row6:
            continue

        after6 = (list(row6) + [""] * 6)[:6]
        after6[0] = codigo

        if codigo not in html_map:
            diffs.append({
                "type": "NUEVO",
                "id": codigo,
                "before": "(no existía en HTML)",
                "after": "\t".join([str(x) for x in after6]),
            })
            added += 1
        else:
            before_row = html_map[codigo]
            before6 = (before_row + [""] * 6)[:6]

            if not _rows_equal_for_diff(before6, after6):
                diffs.append({
                    "type": "MODIFICADO",
                    "id": codigo,
                    "before": "\t".join([str(x) for x in before6]),
                    "after": "\t".join([str(x) for x in after6]),
                })
                modified += 1

    if not diffs:
        return diffs, None, {"modified": 0, "new": 0, "deleted": 0}

    out = []
    out.append(_tsv_line(header_cols[:6], nl))

    for codigo in order_codigos:
        row6 = rows_by_codigo.get(codigo)
        if not row6:
            continue
        row6_out = (list(row6) + [""] * 6)[:6]
        row6_out[0] = codigo
        out.append(_tsv_line(row6_out, nl))

    new_inner = "".join(out)
    new_html = html[:inner_start] + new_inner + html[inner_end:]
    return diffs, (new_html, had_bom, enc), {"modified": modified, "new": added, "deleted": deleted}


def apply_overwrite_html(source_html: str, new_html_tuple, make_backup: bool = False):
    new_html, had_bom, enc = new_html_tuple

    if make_backup and os.path.exists(source_html):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{source_html}.bak_{ts}"
        try:
            shutil.copy2(source_html, backup_path)
        except Exception:
            pass

    tmp_path = source_html + ".tmp_write"
    write_text_preserve_bom(tmp_path, new_html, had_bom, enc)
    os.replace(tmp_path, source_html)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.settings = load_settings()

        self.title("Catálogo (Imágenes ↔ HTML ↔ Excel)")
        self.geometry("1180x740")
        self.minsize(1060, 660)
        self.resizable(True, True)

        self.source_var = tk.StringVar(value="")
        self.excel_var = tk.StringVar(value="")
        self.images_dir_var = tk.StringVar(value="")

        self.status_var = tk.StringVar(value="Selecciona el catalogo.html fuente y/o una carpeta de imágenes.")
        self.state_var = tk.StringVar(value="Estado: (sin archivo)")
        self.backup_var = tk.BooleanVar(value=False)

        self._refresh_job = None

        self.products_all = []
        self.products = []

        self._tree_items = {}
        self._prod_to_iid = {}
        self._tk_imgs = []

        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

        self._sort_col = ""
        self._sort_desc = False

        self.filter_col_var = tk.StringVar(value="Nombre producto")
        self.filter_val_var = tk.StringVar(value="")
        self.filter_info_var = tk.StringVar(value="")

        self._filter_mode = "none"
        self._filter_col = ""
        self._filter_value = ""

        self._setup_style()
        self._build_menu()
        self._build_ui()

        self._load_last_paths()
        self._wire_auto_refresh()
        self.refresh_state()

    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self._bg = "#f6f7fb"
        self._card = "#ffffff"
        self._text = "#111827"
        self._muted = "#4b5563"
        self._accent = "#2563eb"
        self._accent2 = "#1d4ed8"
        self._danger = "#b91c1c"
        self._ok = "#065f46"

        self.configure(bg=self._bg)

        style.configure("App.TFrame", background=self._bg)
        style.configure("Card.TLabelframe", background=self._card, padding=(12, 10))
        style.configure("Card.TLabelframe.Label", background=self._card, foreground=self._text, font=("Segoe UI", 10, "bold"))
        style.configure("Card.TFrame", background=self._card)

        style.configure("Title.TLabel", background=self._bg, foreground=self._text, font=("Segoe UI", 16, "bold"))
        style.configure("Sub.TLabel", background=self._bg, foreground=self._muted, font=("Segoe UI", 10))
        style.configure("Info.TLabel", background=self._card, foreground=self._muted, font=("Segoe UI", 10))
        style.configure("State.TLabel", background=self._card, foreground="#0b5394", font=("Segoe UI", 10, "bold"))

        style.configure("TLabel", background=self._card, foreground=self._text)
        style.configure("TEntry", padding=(6, 4))
        style.configure("TButton", padding=(10, 7))

        style.configure("Accent.TButton", foreground="white", background=self._accent, font=("Segoe UI", 10, "bold"))
        style.map("Accent.TButton",
                  background=[("active", self._accent2), ("pressed", self._accent2)],
                  foreground=[("disabled", "#e5e7eb")])

        style.configure("Danger.TButton", foreground="white", background=self._danger, font=("Segoe UI", 10, "bold"))
        style.map("Danger.TButton",
                  background=[("active", "#991b1b"), ("pressed", "#7f1d1d")],
                  foreground=[("disabled", "#e5e7eb")])

        style.configure("Ok.TButton", foreground="white", background=self._ok, font=("Segoe UI", 10, "bold"))
        style.map("Ok.TButton",
                  background=[("active", "#064e3b"), ("pressed", "#064e3b")],
                  foreground=[("disabled", "#e5e7eb")])

        style.configure("Status.TFrame", background=self._bg)
        style.configure("Status.TLabel", background=self._bg, foreground=self._muted, font=("Segoe UI", 9))

        style.configure("Treeview", rowheight=TREE_ROW_HEIGHT_PX)

    def _build_menu(self):
        menubar = tk.Menu(self)

        m_file = tk.Menu(menubar, tearoff=0)
        m_file.add_command(label="Seleccionar FUENTE (HTML)…", command=self.pick_source)
        m_file.add_command(label="Seleccionar Excel…", command=self.pick_excel)
        m_file.add_command(label="Seleccionar carpeta de imágenes…", command=self.pick_images_folder)
        m_file.add_separator()
        m_file.add_command(label="Abrir carpeta de trabajo", command=self.on_open_folder)
        m_file.add_separator()
        m_file.add_command(label="Salir", command=self.destroy)
        menubar.add_cascade(label="Archivo", menu=m_file)

        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label="Acerca de…", command=self._about)
        menubar.add_cascade(label="Ayuda", menu=m_help)

        self.config(menu=menubar)

    def _about(self):
        messagebox.showinfo(
            "Acerca de",
            "Herramienta para:\n"
            "• Cargar productos desde imágenes (codigo_nombre_categoria_marca_valorunitario_stock.ext)\n"
            "• Editar campos en la interfaz (Codigo, Nombre producto, Categoria, Marca, Valor unitario, Stock)\n"
            "• Exportar a Excel con imagen (5cm x 5cm)\n"
            "• Actualizar el HTML fuente desde el Excel virtual o desde un Excel externo\n"
        )

    def _build_ui(self):
        root = ttk.Frame(self, style="App.TFrame")
        root.pack(fill="both", expand=True, padx=14, pady=12)

        header = ttk.Frame(root, style="App.TFrame")
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Catálogo (Imágenes ↔ HTML ↔ Excel)", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="Excel virtual desde imágenes · Edición en tabla · Exportar con imagen · Actualizar HTML", style="Sub.TLabel").pack(anchor="w", pady=(2, 0))

        nb = ttk.Notebook(root)
        nb.pack(fill="both", expand=True)

        tab_opts = ttk.Frame(nb, style="App.TFrame")
        tab_cat = ttk.Frame(nb, style="App.TFrame")

        nb.add(tab_opts, text="Opciones")
        nb.add(tab_cat, text="Catálogo virtual")

        opt_content = ttk.Frame(tab_opts, style="App.TFrame")
        opt_content.pack(fill="both", expand=True)

        left = ttk.Frame(opt_content, style="App.TFrame")
        right = ttk.Frame(opt_content, style="App.TFrame")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        right.grid(row=0, column=1, sticky="nsew")

        opt_content.columnconfigure(0, weight=3)
        opt_content.columnconfigure(1, weight=2)
        opt_content.rowconfigure(0, weight=1)

        lf_paths = ttk.LabelFrame(left, text="Rutas principales", style="Card.TLabelframe")
        lf_paths.pack(fill="x", pady=(0, 10))

        self._build_path_row(
            lf_paths, 0, "Fuente (HTML con <pre id='productos-tsv'>):", self.source_var,
            browse=self.pick_source, open_cmd=lambda: self.open_path(self.source_var.get()),
            folder_cmd=lambda: self.open_folder(self.source_var.get())
        )
        self._build_path_row(
            lf_paths, 1, "Excel (opcional):", self.excel_var,
            browse=self.pick_excel, open_cmd=lambda: self.open_path(self.excel_var.get()),
            folder_cmd=lambda: self.open_folder(self.excel_var.get())
        )
        self._build_path_row(
            lf_paths, 2, "Carpeta de imágenes:", self.images_dir_var,
            browse=self.pick_images_folder, open_cmd=lambda: self.open_path(self.images_dir_var.get()),
            folder_cmd=lambda: self.open_folder(self.images_dir_var.get())
        )

        lf_ops = ttk.LabelFrame(right, text="Operaciones HTML ↔ Excel", style="Card.TLabelframe")
        lf_ops.pack(fill="x", pady=(0, 10))

        ops_grid = ttk.Frame(lf_ops, style="Card.TFrame")
        ops_grid.pack(fill="x")

        ttk.Button(ops_grid, text="Exportar HTML → Excel", style="Accent.TButton", command=self.on_export_excel_from_html).grid(row=0, column=0, sticky="we", padx=(0, 8), pady=(0, 8))
        ttk.Button(ops_grid, text="Actualizar HTML desde Excel", style="Danger.TButton", command=self.on_update_source_from_excel).grid(row=0, column=1, sticky="we", pady=(0, 8))

        ttk.Separator(ops_grid, orient="horizontal").grid(row=1, column=0, columnspan=2, sticky="we", pady=(2, 10))

        update_row = ttk.Frame(ops_grid, style="Card.TFrame")
        update_row.grid(row=2, column=0, columnspan=2, sticky="we")
        ttk.Checkbutton(update_row, text="Crear copia .bak antes de sobrescribir el HTML fuente", variable=self.backup_var).pack(anchor="w", pady=(0, 6))
        ttk.Button(update_row, text="Abrir carpeta", command=self.on_open_folder).pack(fill="x")

        ops_grid.columnconfigure(0, weight=1)
        ops_grid.columnconfigure(1, weight=1)

        lf_state = ttk.LabelFrame(right, text="Estado y validación", style="Card.TLabelframe")
        lf_state.pack(fill="both", expand=True)

        ttk.Label(lf_state, textvariable=self.state_var, style="State.TLabel").pack(fill="x", pady=(0, 8))

        ttk.Label(lf_state, text="Detalles / advertencias:", style="Info.TLabel").pack(anchor="w")
        self.txt_valid = ScrolledText(lf_state, height=16, wrap="word")
        self.txt_valid.pack(fill="both", expand=True, pady=(6, 8))
        self.txt_valid.configure(state="disabled", font=("Segoe UI", 9))

        btns = ttk.Frame(lf_state, style="Card.TFrame")
        btns.pack(fill="x")
        ttk.Button(btns, text="Refrescar estado", command=self.refresh_state).pack(side="left")
        ttk.Button(btns, text="Copiar estado", command=self.copy_state).pack(side="left", padx=(8, 0))

        lf_virtual = ttk.LabelFrame(tab_cat, text="Excel virtual (desde imágenes)", style="Card.TLabelframe")
        lf_virtual.pack(fill="both", expand=True, padx=0, pady=0)

        topbar = ttk.Frame(lf_virtual, style="Card.TFrame")
        topbar.pack(fill="x", pady=(0, 8))
        ttk.Button(topbar, text="Cargar/Refrescar carpeta", style="Accent.TButton", command=self.on_load_images).pack(side="left")
        ttk.Button(topbar, text="Exportar Excel (virtual)", style="Ok.TButton", command=self.on_export_virtual_excel).pack(side="left", padx=(10, 0))
        ttk.Button(topbar, text="Actualizar HTML desde virtual", style="Danger.TButton", command=self.on_update_source_from_virtual).pack(side="left", padx=(10, 0))

        filterbar = ttk.Frame(lf_virtual, style="Card.TFrame")
        filterbar.pack(fill="x", pady=(0, 8))

        cols = ("Codigo", "Nombre producto", "Categoria", "Marca", "Valor unitario", "Stock")
        filter_cols = ["Imagen"] + list(cols)

        ttk.Label(filterbar, text="Columna:").pack(side="left")
        cmb = ttk.Combobox(filterbar, values=filter_cols, textvariable=self.filter_col_var, state="readonly", width=18)
        cmb.pack(side="left", padx=(8, 12))

        ttk.Label(filterbar, text="Valor (opcional):").pack(side="left")
        ent = ttk.Entry(filterbar, textvariable=self.filter_val_var, width=30)
        ent.pack(side="left", padx=(8, 12))
        ent.bind("<Return>", lambda _e: self.on_apply_value_filter())

        ttk.Button(filterbar, text="Filtrar = valor", command=self.on_apply_value_filter).pack(side="left", padx=(0, 8))
        ttk.Button(filterbar, text="Mostrar duplicados", command=self.on_show_duplicates).pack(side="left", padx=(0, 8))
        ttk.Button(filterbar, text="Limpiar", command=self.on_clear_filter).pack(side="left", padx=(0, 12))

        ttk.Label(filterbar, textvariable=self.filter_info_var, style="Info.TLabel").pack(side="right")

        ttk.Separator(lf_virtual, orient="horizontal").pack(fill="x", pady=(0, 8))

        self.tree = ttk.Treeview(lf_virtual, columns=cols, show="tree headings", height=6)
        self.tree.pack(fill="both", expand=True)

        self.tree.heading("#0", text="Imagen", command=lambda: self._sort_tree_by_column("Imagen"))
        self.tree.column("#0", width=max(210, IMG_SIZE_PX + 16), minwidth=max(190, IMG_SIZE_PX), stretch=False, anchor="center")

        for c in cols:
            self.tree.heading(c, text=c, command=lambda _c=c: self._sort_tree_by_column(_c))
            if c == "Nombre producto":
                self.tree.column(c, width=320, minwidth=220, stretch=True, anchor="w")
            elif c in ("Categoria", "Marca"):
                self.tree.column(c, width=200, minwidth=140, stretch=True, anchor="w")
            elif c == "Codigo":
                self.tree.column(c, width=160, minwidth=120, stretch=False, anchor="center")
            elif c == "Valor unitario":
                self.tree.column(c, width=140, minwidth=110, stretch=False, anchor="e")
            else:
                self.tree.column(c, width=110, minwidth=80, stretch=False, anchor="center")

        self.tree.bind("<Double-1>", self._on_tree_double_click)
        self.tree.bind("<Button-1>", self._on_tree_single_click)

        status = ttk.Frame(root, style="Status.TFrame")
        status.pack(fill="x", pady=(10, 0))
        ttk.Label(status, textvariable=self.status_var, style="Status.TLabel").pack(side="left", fill="x", expand=True)
        ttk.Sizegrip(status).pack(side="right")

    def _build_path_row(self, parent, row, label, var, browse, open_cmd, folder_cmd):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.grid(row=row, column=0, sticky="we", pady=(0, 10))
        parent.columnconfigure(0, weight=1)

        ttk.Label(frame, text=label).grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(frame, textvariable=var).grid(row=1, column=0, columnspan=5, sticky="we")

        ttk.Button(frame, text="Seleccionar…", command=browse).grid(row=1, column=5, padx=(8, 0))
        ttk.Button(frame, text="Abrir", command=open_cmd).grid(row=1, column=6, padx=(8, 0))
        ttk.Button(frame, text="Carpeta", command=folder_cmd).grid(row=1, column=7, padx=(8, 0))
        ttk.Button(frame, text="Copiar", command=lambda v=var: self.copy_to_clipboard(v.get())).grid(row=1, column=8, padx=(8, 0))

        frame.columnconfigure(0, weight=1)

    def _wire_auto_refresh(self):
        def schedule(*_):
            if self._refresh_job is not None:
                try:
                    self.after_cancel(self._refresh_job)
                except Exception:
                    pass
            self._refresh_job = self.after(350, self.refresh_state)

        for v in (self.source_var, self.excel_var, self.images_dir_var):
            v.trace_add("write", schedule)

    def _load_last_paths(self):
        last_source = (self.settings.get("last_source") or "").strip()
        last_excel = (self.settings.get("last_excel") or "").strip()
        last_images = (self.settings.get("last_images_dir") or "").strip()

        if last_source and os.path.exists(last_source):
            self.source_var.set(last_source)

        if last_excel and os.path.exists(last_excel):
            self.excel_var.set(last_excel)
        else:
            self.excel_var.set(default_excel_path())

        if last_images and os.path.isdir(last_images):
            self.images_dir_var.set(last_images)

        if self.images_dir_var.get().strip():
            self.on_load_images()

    def copy_to_clipboard(self, text: str):
        text = (text or "").strip()
        if not text:
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self.status_var.set("Copiado al portapapeles.")

    def copy_state(self):
        self.copy_to_clipboard(self.state_var.get())

    def open_path(self, path: str):
        try:
            open_with_default_app(path.strip())
        except Exception as e:
            messagebox.showerror("No se pudo abrir", str(e))

    def open_folder(self, path: str):
        try:
            open_folder_of(path.strip())
        except Exception as e:
            messagebox.showerror("No se pudo abrir carpeta", str(e))

    def pick_source(self):
        initialdir = None
        last_dir = (self.settings.get("last_dir") or "").strip()
        if last_dir and os.path.isdir(last_dir):
            initialdir = last_dir

        path = filedialog.askopenfilename(
            title="Selecciona el catalogo.html FUENTE (con <pre id='productos-tsv'>)",
            initialdir=initialdir,
            filetypes=[("HTML", "*.html;*.htm"), ("Todos", "*.*")]
        )
        if path:
            self.source_var.set(path)
            self.settings = remember_paths(self.settings, source_html=path)
            self.status_var.set("Fuente seleccionada.")
            self.refresh_state()

    def pick_excel(self):
        initialdir = program_root_dir()
        last_dir = (self.settings.get("last_dir") or "").strip()
        if last_dir and os.path.isdir(last_dir):
            initialdir = last_dir

        path = filedialog.askopenfilename(
            title="Selecciona el Excel",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if path:
            self.excel_var.set(path)
            self.settings = remember_paths(self.settings, excel_path=path)
            self.status_var.set("Excel seleccionado.")
            self.refresh_state()

    def pick_images_folder(self):
        initialdir = None
        last_images = (self.settings.get("last_images_dir") or "").strip()
        last_dir = (self.settings.get("last_dir") or "").strip()
        if last_images and os.path.isdir(last_images):
            initialdir = last_images
        elif last_dir and os.path.isdir(last_dir):
            initialdir = last_dir

        path = filedialog.askdirectory(
            title="Selecciona la carpeta de imágenes",
            initialdir=initialdir
        )
        if path:
            self.images_dir_var.set(path)
            self.settings = remember_paths(self.settings, images_dir=path)
            self.status_var.set("Carpeta de imágenes seleccionada.")
            self.on_load_images()
            self.refresh_state()

    def on_open_folder(self):
        path = self.images_dir_var.get().strip() or self.source_var.get().strip() or self.excel_var.get().strip() or program_root_dir()
        try:
            open_folder_of(path)
        except Exception as e:
            messagebox.showerror("No se pudo abrir carpeta", str(e))

    def _describe_html_state(self, html_path: str):
        if not html_path or not os.path.exists(html_path):
            return "(no existe)", ""

        try:
            html, _, _ = read_text_preserve_bom(html_path)
            block, err = find_single_pre_block(html)
            if err:
                return "NO DISPONIBLE", err

            _, inner, _, _, _ = block
            val = validate_tsv_structure(inner)
            if not val["ok"]:
                return "NO VÁLIDO", "Validación:\n- " + "\n- ".join(val["issues"])

            return "OK (catálogo detectado)", ""
        except Exception as e:
            return "ERROR", str(e)

    def _set_valid_text(self, text: str):
        self.txt_valid.configure(state="normal")
        self.txt_valid.delete("1.0", "end")
        if text.strip():
            self.txt_valid.insert("1.0", text.strip())
        self.txt_valid.configure(state="disabled")

    def refresh_state(self):
        src = self.source_var.get().strip()
        excel = self.excel_var.get().strip()
        imgdir = self.images_dir_var.get().strip()

        src_state, src_msg = self._describe_html_state(src)
        excel_ok = "OK" if (excel and os.path.exists(excel)) else "NO"
        img_ok = "OK" if (imgdir and os.path.isdir(imgdir)) else "NO"

        total = len(self.products_all)
        visible = len(self.products)

        lines = [
            f"Fuente: {src_state}",
            f"Excel: {excel_ok}",
            f"Imágenes: {img_ok}",
            f"Virtual: {visible}/{total} productos",
        ]
        self.state_var.set("Estado: " + "   |   ".join(lines))

        warn = []
        if src_msg:
            warn.append("FUENTE:\n" + src_msg)

        if img_ok == "OK" and total == 0:
            warn.append("VIRTUAL:\nNo se encontraron imágenes soportadas en la carpeta.")
        if total > 0:
            dups = self._find_duplicate_codigos()
            if dups:
                warn.append("VIRTUAL:\nHay códigos duplicados. Ejemplos: " + ", ".join(dups[:20]) + (" ..." if len(dups) > 20 else ""))

            bad_stock = []
            for p in self.products_all:
                if not looks_like_stock(p.get("stock", "")):
                    bad_stock.append(p.get("imagen", ""))
                    if len(bad_stock) >= 10:
                        break
            if bad_stock:
                warn.append("VIRTUAL:\nHay productos con Stock no válido (debe ser número). Ejemplos: " + ", ".join(bad_stock))

            bad_valor = []
            for p in self.products_all:
                if not looks_like_valor_unitario(p.get("valor_unitario", "")):
                    bad_valor.append(p.get("imagen", ""))
                    if len(bad_valor) >= 10:
                        break
            if bad_valor:
                warn.append("VIRTUAL:\nHay productos con Valor unitario no válido. Ejemplos: " + ", ".join(bad_valor))

            empty_code = []
            for p in self.products_all:
                if not str(p.get("codigo", "")).strip():
                    empty_code.append(p.get("imagen", ""))
                    if len(empty_code) >= 10:
                        break
            if empty_code:
                warn.append("VIRTUAL:\nHay productos sin Codigo. Ejemplos: " + ", ".join(empty_code))

        self._set_valid_text("\n\n".join(warn))

    def _find_duplicate_codigos(self):
        seen = set()
        dups = []
        for p in self.products_all:
            c = str(p.get("codigo", "")).strip()
            if not c:
                continue
            if c in seen:
                dups.append(c)
            else:
                seen.add(c)
        return dups

    def _make_thumbnail(self, img_path: str, size_px: int = IMG_SIZE_PX):
        if not img_path or not os.path.exists(img_path):
            return None

        if PILImage is not None and ImageTk is not None:
            try:
                im = PILImage.open(img_path)
                im = im.convert("RGBA")
                im.thumbnail((size_px, size_px))
                return ImageTk.PhotoImage(im)
            except Exception:
                return None

        ext = os.path.splitext(img_path)[1].lower()
        if ext in (".png", ".gif"):
            try:
                return tk.PhotoImage(file=img_path)
            except Exception:
                return None
        return None

    def on_load_images(self):
        imgdir = self.images_dir_var.get().strip()
        if not imgdir or not os.path.isdir(imgdir):
            messagebox.showwarning("Falta carpeta", "Selecciona una carpeta de imágenes válida.")
            return

        self.products_all = load_products_from_folder(imgdir)
        self._apply_view()
        self.settings = remember_paths(self.settings, images_dir=imgdir)
        self.status_var.set(f"Cargados {len(self.products_all)} productos desde imágenes.")
        self.refresh_state()

    def _value_by_column(self, prod: dict, col_name: str) -> str:
        if col_name == "Imagen":
            return str(prod.get("imagen", "")).strip()
        if col_name == "Codigo":
            return str(prod.get("codigo", "")).strip()
        if col_name == "Nombre producto":
            return str(prod.get("nombre", "")).strip()
        if col_name == "Categoria":
            return str(prod.get("categoria", "")).strip()
        if col_name == "Marca":
            return str(prod.get("marca", "")).strip()
        if col_name == "Valor unitario":
            return str(prod.get("valor_unitario", "")).strip()
        if col_name == "Stock":
            return str(prod.get("stock", "")).strip()
        return ""

    def _norm_key_for_filter(self, prod: dict, col_name: str) -> str:
        v = self._value_by_column(prod, col_name)
        if col_name == "Valor unitario":
            d = _decimal_from_text(v)
            return str(d) if d is not None else ""
        if col_name == "Stock":
            n = _int_from_digits(v)
            return str(n) if n is not None else ""
        if col_name == "Codigo":
            return str(v).strip()
        return norm_text(v)

    def _sort_key_for_product(self, prod: dict, col_name: str):
        if col_name == "Imagen":
            v = self._value_by_column(prod, "Imagen")
            nv = norm_text(v)
            return (1 if nv == "" else 0, nv, v)
        if col_name == "Codigo":
            v = self._value_by_column(prod, "Codigo")
            d = _digits_only(v)
            if d.isdigit():
                return (0, int(d), v)
            return (1, norm_text(v), v)
        if col_name == "Stock":
            v = self._value_by_column(prod, "Stock")
            n = _int_from_digits(v)
            if n is None:
                return (1, 0)
            return (0, n)
        if col_name == "Valor unitario":
            v = self._value_by_column(prod, "Valor unitario")
            d = _decimal_from_text(v)
            if d is None:
                return (1, Decimal("0"))
            return (0, d)
        v = self._value_by_column(prod, col_name)
        nv = norm_text(v)
        return (1 if nv == "" else 0, nv, v)

    def _update_sort_markers(self):
        cols = ("Codigo", "Nombre producto", "Categoria", "Marca", "Valor unitario", "Stock")
        marker = ""
        if self._sort_col == "Imagen":
            marker = " ▼" if self._sort_desc else " ▲"
        self.tree.heading("#0", text="Imagen" + marker)

        for c in cols:
            marker = ""
            if self._sort_col == c:
                marker = " ▼" if self._sort_desc else " ▲"
            self.tree.heading(c, text=c + marker)

    def _apply_view(self, keep_prod: dict = None):
        if not self.products_all:
            self.products = []
            self._render_tree()
            self.filter_info_var.set("0 productos")
            self._update_sort_markers()
            return

        if keep_prod is None:
            sel = self.tree.selection()
            if sel:
                keep_prod = self._tree_items.get(sel[0])

        y = None
        try:
            y = self.tree.yview()[0]
        except Exception:
            y = None

        view = list(self.products_all)

        mode = (self._filter_mode or "none").lower()
        col = (self._filter_col or "").strip()
        val = (self._filter_value or "").strip()

        if mode == "value" and col:
            if val == "":
                view = list(self.products_all)
            else:
                if col in ("Valor unitario", "Stock"):
                    if col == "Valor unitario":
                        target = _decimal_from_text(val)
                        if target is None:
                            view = []
                        else:
                            out = []
                            for p in self.products_all:
                                d = _decimal_from_text(self._value_by_column(p, col))
                                if d is not None and d == target:
                                    out.append(p)
                            view = out
                    else:
                        target = _int_from_digits(val)
                        if target is None:
                            view = []
                        else:
                            out = []
                            for p in self.products_all:
                                n = _int_from_digits(self._value_by_column(p, col))
                                if n is not None and n == target:
                                    out.append(p)
                            view = out
                elif col == "Codigo":
                    t = val.strip()
                    view = [p for p in self.products_all if self._value_by_column(p, col).strip() == t]
                else:
                    t = norm_text(val)
                    view = [p for p in self.products_all if self._norm_key_for_filter(p, col) == t]

        elif mode == "dups" and col:
            counts = {}
            keys = []
            for p in self.products_all:
                k = self._norm_key_for_filter(p, col)
                keys.append(k)
                if k == "":
                    continue
                counts[k] = counts.get(k, 0) + 1
            out = []
            for p in self.products_all:
                k = self._norm_key_for_filter(p, col)
                if k != "" and counts.get(k, 0) > 1:
                    out.append(p)
            view = out

        if self._sort_col:
            view = sorted(view, key=lambda p: self._sort_key_for_product(p, self._sort_col), reverse=self._sort_desc)

        self.products = view
        self._render_tree()

        if keep_prod is not None:
            iid = self._prod_to_iid.get(id(keep_prod))
            if iid:
                try:
                    self.tree.selection_set(iid)
                    self.tree.see(iid)
                except Exception:
                    pass

        if y is not None:
            try:
                self.tree.yview_moveto(y)
            except Exception:
                pass

        total = len(self.products_all)
        shown = len(self.products)
        if self._filter_mode != "none":
            self.filter_info_var.set(f"Mostrando {shown} de {total}")
        else:
            self.filter_info_var.set(f"{total} productos")

        self._update_sort_markers()

    def _render_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._tree_items.clear()
        self._prod_to_iid.clear()
        self._tk_imgs.clear()

        for p in self.products:
            img = self._make_thumbnail(p.get("path", ""), size_px=IMG_SIZE_PX)
            if img is not None:
                self._tk_imgs.append(img)

            iid = self.tree.insert(
                "",
                "end",
                image=img if img is not None else "",
                text="",
                values=(
                    str(p.get("codigo", "")),
                    str(p.get("nombre", "")),
                    str(p.get("categoria", "")),
                    str(p.get("marca", "")),
                    str(p.get("valor_unitario", "")),
                    str(p.get("stock", "")),
                )
            )
            self._tree_items[iid] = p
            self._prod_to_iid[id(p)] = iid

    def on_apply_value_filter(self):
        col = (self.filter_col_var.get() or "").strip()
        self._filter_mode = "value"
        self._filter_col = col
        self._filter_value = (self.filter_val_var.get() or "").strip()
        self._apply_view()
        self.status_var.set("Filtro aplicado.")
        self.refresh_state()

    def on_show_duplicates(self):
        col = (self.filter_col_var.get() or "").strip()
        self._filter_mode = "dups"
        self._filter_col = col
        self._filter_value = ""
        self._apply_view()
        self.status_var.set("Mostrando duplicados.")
        self.refresh_state()

    def on_clear_filter(self):
        self._filter_mode = "none"
        self._filter_col = ""
        self._filter_value = ""
        self.filter_val_var.set("")
        self._apply_view()
        self.status_var.set("Filtro limpiado.")
        self.refresh_state()

    def _sort_tree_by_column(self, col_name: str):
        col_name = (col_name or "").strip()
        if not col_name:
            return
        if self._sort_col == col_name:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col = col_name
            self._sort_desc = False
        self._apply_view()
        self.status_var.set(f"Ordenado por: {col_name}")

    def _on_tree_single_click(self, _event):
        self._end_cell_edit(commit=False)

    def _on_tree_double_click(self, event):
        self._end_cell_edit(commit=True)

        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item or not col:
            return
        if col == "#0":
            return

        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox

        col_idx = int(col.replace("#", "")) - 1
        current = self.tree.item(item, "values")
        if col_idx < 0 or col_idx >= len(current):
            return

        self._edit_item = item
        self._edit_col = col

        entry = ttk.Entry(self.tree)
        entry.insert(0, current[col_idx])
        entry.select_range(0, "end")
        entry.focus_set()
        entry.place(x=x, y=y, width=w, height=h)

        entry.bind("<Return>", lambda e: self._end_cell_edit(commit=True))
        entry.bind("<Escape>", lambda e: self._end_cell_edit(commit=False))
        entry.bind("<FocusOut>", lambda e: self._end_cell_edit(commit=True))

        self._edit_entry = entry

    def _end_cell_edit(self, commit: bool):
        if self._edit_entry is None:
            return
        entry = self._edit_entry
        item = self._edit_item
        col = self._edit_col

        try:
            val = entry.get()
        except Exception:
            val = ""

        try:
            entry.place_forget()
        except Exception:
            pass
        try:
            entry.destroy()
        except Exception:
            pass

        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

        if not commit or not item or not col:
            return

        prod = self._tree_items.get(item)
        if prod is None:
            return

        values = list(self.tree.item(item, "values"))
        col_idx = int(col.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(values):
            return

        values[col_idx] = val
        try:
            self.tree.item(item, values=tuple(values))
        except Exception:
            pass

        field_map = {
            0: "codigo",
            1: "nombre",
            2: "categoria",
            3: "marca",
            4: "valor_unitario",
            5: "stock",
        }
        field = field_map.get(col_idx)
        if field:
            prod[field] = val

        self._apply_view(keep_prod=prod)
        self.refresh_state()

    def _diff_changed_word_indices(self, before: str, after: str) -> set[int]:
        b_words = re.findall(r"\S+", before or "")
        a_words = re.findall(r"\S+", after or "")

        sm = difflib.SequenceMatcher(a=b_words, b=a_words)
        changed_a = set()

        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag in ("replace", "insert"):
                for k in range(j1, j2):
                    changed_a.add(k)

        return changed_a

    def _insert_with_red_changed(self, txt: tk.Text, before: str, after: str, red_tag: str):
        changed = self._diff_changed_word_indices(before, after)
        parts = re.split(r"(\s+)", after or "")
        word_idx = 0

        for part in parts:
            if part == "":
                continue
            if part.isspace():
                txt.insert("end", part)
            else:
                if word_idx in changed:
                    txt.insert("end", part, red_tag)
                else:
                    txt.insert("end", part)
                word_idx += 1

    def _show_preview_and_confirm(self, diffs: list, summary: dict):
        top = tk.Toplevel(self)
        top.title("Vista previa de cambios (→ HTML fuente)")
        top.geometry("980x560")
        top.minsize(900, 520)
        top.transient(self)
        top.grab_set()

        frm = ttk.Frame(top, padding=12)
        frm.pack(fill="both", expand=True)

        header = (
            "Sincronización EXACTA (copia fiel) hacia el HTML usando 6 columnas:\n"
            "Codigo\tNombre producto\tCategoria\tMarca\tValor unitario\tStock\n\n"
            f"- MODIFICADOS: {summary.get('modified', 0)}\n"
            f"- NUEVOS:      {summary.get('new', 0)}\n"
            f"- ELIMINADOS:  {summary.get('deleted', 0)}\n\n"
        )
        if self.backup_var.get():
            header += "• Se creará una copia .bak antes de sobrescribir el HTML.\n"
        else:
            header += "• No se creará copia .bak (opción desactivada).\n"
        header += "-" * 92 + "\n\n"

        txt = ScrolledText(frm, wrap="none", height=24)
        txt.pack(fill="both", expand=True)

        txt.tag_configure("redchg", foreground="#b91c1c")
        txt.tag_configure("muted", foreground="#374151")

        txt.insert("1.0", header, ("muted",))

        max_show = 300
        for i, d in enumerate(diffs[:max_show], start=1):
            tipo = d.get("type", "")
            rid = d.get("id", "")
            before_line = d.get("before", "")
            after_line = d.get("after", "")

            txt.insert("end", f"[{i}] {tipo} | codigo={rid}\n")
            txt.insert("end", "    ANTES:   ")
            txt.insert("end", f"{before_line}\n")
            txt.insert("end", "    DESPUÉS: ")

            if tipo == "MODIFICADO":
                self._insert_with_red_changed(txt, before_line, after_line, "redchg")
                txt.insert("end", "\n\n")
            elif tipo == "NUEVO":
                txt.insert("end", after_line, "redchg")
                txt.insert("end", "\n\n")
            else:
                txt.insert("end", f"{after_line}\n\n")

        if len(diffs) > max_show:
            txt.insert("end", f"(Hay {len(diffs)} filas afectadas; se muestran solo las primeras {max_show}.)\n", ("muted",))

        txt.configure(state="disabled", font=("Consolas", 9))

        decision = tk.StringVar(value="")
        btn_frame = ttk.Frame(frm)
        btn_frame.pack(pady=(10, 0), fill="x")

        def accept():
            decision.set("yes")
            top.destroy()

        def cancel():
            decision.set("no")
            top.destroy()

        ttk.Button(btn_frame, text="Aplicar cambios", style="Danger.TButton", command=accept).pack(side="left")
        ttk.Button(btn_frame, text="Cancelar", command=cancel).pack(side="left", padx=(10, 0))

        self.wait_variable(decision)
        return decision.get() == "yes"

    def on_export_virtual_excel(self):
        if not self.products_all:
            messagebox.showwarning("Sin productos", "Carga primero una carpeta de imágenes para crear el Excel virtual.")
            return

        dups = self._find_duplicate_codigos()
        if dups:
            messagebox.showerror("Códigos duplicados", "Corrige códigos duplicados antes de exportar.")
            return

        for p in self.products_all:
            if not str(p.get("codigo", "")).strip():
                messagebox.showerror("Codigo vacío", "Hay productos con Codigo vacío. Corrige antes de exportar.")
                return
            if not looks_like_valor_unitario(p.get("valor_unitario", "")):
                messagebox.showerror("Valor unitario inválido", "Hay productos con Valor unitario no numérico. Corrige antes de exportar.")
                return
            if not looks_like_stock(p.get("stock", "")):
                messagebox.showerror("Stock inválido", "Hay productos con Stock no numérico. Corrige antes de exportar.")
                return

        initialdir = program_root_dir()
        default_name = "productos_virtual.xlsx"
        out = filedialog.asksaveasfilename(
            title="Guardar Excel (virtual)",
            initialdir=initialdir,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not out:
            return

        try:
            export_products_to_excel(self.products_all, out, images_dir=self.images_dir_var.get().strip())
            self.excel_var.set(out)
            self.settings = remember_paths(self.settings, excel_path=out)
            msg = f"Excel virtual exportado: {os.path.basename(out)}"
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)
        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Error", str(e))

    def on_export_excel_from_html(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return

        try:
            base_excel = default_excel_path()
            imgdir = self.images_dir_var.get().strip()
            excel_path, rows_count, existed = export_catalog_html_to_excel(src, base_excel, images_dir=imgdir)

            self.excel_var.set(excel_path)
            self.settings = remember_paths(self.settings, excel_path=excel_path)

            if existed:
                msg = (
                    f"Ya existía '{os.path.basename(base_excel)}'.\n"
                    f"Para no sobrescribirlo, se creó: '{os.path.basename(excel_path)}'.\n\n"
                    f"Filas exportadas: {rows_count}"
                )
            else:
                msg = f"Excel creado: '{os.path.basename(excel_path)}'\n\nFilas exportadas: {rows_count}"

            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Exportación completada", msg)

        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))

    def on_update_source_from_virtual(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return
        if not self.products_all:
            messagebox.showwarning("Sin productos", "Carga primero una carpeta de imágenes.")
            return

        dups = self._find_duplicate_codigos()
        if dups:
            messagebox.showerror("Códigos duplicados", "Corrige códigos duplicados antes de actualizar el HTML.")
            return

        rows_by_codigo = {}
        order_codigos = []
        for p in self.products_all:
            codigo = str(p.get("codigo", "")).strip()
            nombre = str(p.get("nombre", "")).strip()
            categoria = str(p.get("categoria", "")).strip()
            marca = str(p.get("marca", "")).strip()
            valor_unitario = str(p.get("valor_unitario", "")).strip()
            stock = str(p.get("stock", "")).strip()

            if not codigo:
                messagebox.showerror("Codigo vacío", "Hay productos con Codigo vacío. Corrige antes de actualizar el HTML.")
                return
            if not looks_like_valor_unitario(valor_unitario):
                messagebox.showerror("Valor unitario inválido", "Hay productos con Valor unitario no numérico. Corrige antes de actualizar el HTML.")
                return
            if not looks_like_stock(stock):
                messagebox.showerror("Stock inválido", "Hay productos con Stock no numérico. Corrige antes de actualizar el HTML.")
                return

            row6 = [codigo, nombre, categoria, marca, valor_unitario, stock]
            rows_by_codigo[codigo] = row6
            order_codigos.append(codigo)

        try:
            diffs, new_html_tuple, summary = compute_diffs_and_build_new_html_exact(
                src, rows_by_codigo, order_codigos
            )

            if not diffs:
                msg = "No hay cambios: el HTML ya coincide con el Excel virtual."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Sin cambios", msg)
                return

            ok = self._show_preview_and_confirm(diffs, summary)
            if not ok:
                msg = "Operación cancelada. No se modificó el HTML fuente."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Cancelado", msg)
                return

            apply_overwrite_html(src, new_html_tuple, make_backup=self.backup_var.get())

            msg = (
                "Actualización aplicada al HTML fuente desde el Excel virtual.\n\n"
                f"- Modificados: {summary.get('modified', 0)}\n"
                f"- Nuevos: {summary.get('new', 0)}\n"
                f"- Eliminados: {summary.get('deleted', 0)}\n"
            )
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)

        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))

    def on_update_source_from_excel(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return

        excel_path = self.excel_var.get().strip()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Falta Excel", "Selecciona un Excel válido primero.")
            return

        try:
            rows_by_codigo, order_codigos = read_catalog_excel_flexible(excel_path)

            diffs, new_html_tuple, summary = compute_diffs_and_build_new_html_exact(
                src, rows_by_codigo, order_codigos
            )

            if not diffs:
                msg = "No hay cambios: el HTML ya coincide con el Excel."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Sin cambios", msg)
                return

            ok = self._show_preview_and_confirm(diffs, summary)
            if not ok:
                msg = "Operación cancelada. No se modificó el HTML fuente."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Cancelado", msg)
                return

            apply_overwrite_html(src, new_html_tuple, make_backup=self.backup_var.get())

            msg = (
                "Actualización aplicada al HTML fuente desde el Excel.\n\n"
                f"- Modificados: {summary.get('modified', 0)}\n"
                f"- Nuevos: {summary.get('new', 0)}\n"
                f"- Eliminados: {summary.get('deleted', 0)}\n"
            )
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)

        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))


if __name__ == "__main__":
    App().mainloop()
