import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

import os
import re
import json
import difflib
from datetime import datetime
import unicodedata
import subprocess
import sys
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# =========================
#  Reglas/regex del HTML
# =========================

PRE_REGEX_ALL = re.compile(
    r'(<pre\b[^>]*\bid\s*=\s*["\']productos-tsv["\'][^>]*>)([\s\S]*?)(</pre>)',
    flags=re.IGNORECASE
)

HTML_COMMENT_REGEX = re.compile(r"<!--[\s\S]*?-->", flags=re.MULTILINE)
EXPECTED_COLS = ["nombre producto", "categoria", "marca", "valor unitario", "id"]

CATALOG_TSV_SUFFIX = ".CATALOGO.tsv"
CATALOG_META_SUFFIX = ".CATALOGO.json"

# Excel en raíz del programa
EXCEL_BASE_NAME = "productos"   # => productos.xlsx, productos_1.xlsx, productos_2.xlsx...
EXCEL_SHEET_NAME = "Catalogo"


# =========================
#  Persistencia (último archivo/carpeta)
# =========================

def get_settings_path() -> str:
    appdata = os.getenv("APPDATA")
    if appdata:
        folder = os.path.join(appdata, "CatalogoToggle")
    else:
        folder = os.path.join(os.path.expanduser("~"), ".catalogo_toggle")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "settings.json")


def load_settings() -> dict:
    path = get_settings_path()
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
    except Exception:
        pass
    return {}


def save_settings(data: dict) -> None:
    path = get_settings_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def remember_paths(settings: dict, source_html: str = "", stripped_html: str = "", tsv_path: str = "", meta_path: str = "",
                   excel_path: str = "") -> dict:
    settings = settings or {}
    if source_html:
        settings["last_source"] = os.path.abspath(source_html)
        settings["last_dir"] = os.path.dirname(os.path.abspath(source_html))
    if stripped_html:
        settings["last_stripped"] = os.path.abspath(stripped_html)
        settings["last_dir"] = os.path.dirname(os.path.abspath(stripped_html))
    if tsv_path:
        settings["last_tsv"] = os.path.abspath(tsv_path)
    if meta_path:
        settings["last_meta"] = os.path.abspath(meta_path)
    if excel_path:
        settings["last_excel"] = os.path.abspath(excel_path)
    save_settings(settings)
    return settings


# =========================
#  Abrir archivo/carpeta con app predeterminada
# =========================

def open_with_default_app(path: str) -> None:
    path = (path or "").strip()
    if not path:
        raise RuntimeError("No hay ruta para abrir.")
    if not os.path.exists(path):
        raise RuntimeError("La ruta no existe.")

    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
            return
        subprocess.run(["xdg-open", path], check=False)
    except Exception as e:
        raise RuntimeError(f"No se pudo abrir. Detalle: {e}")


def open_folder_of(path: str) -> None:
    path = (path or "").strip()
    if not path:
        raise RuntimeError("No hay ruta.")
    folder = path if os.path.isdir(path) else os.path.dirname(path)
    if not folder or not os.path.isdir(folder):
        raise RuntimeError("No se pudo determinar la carpeta.")
    open_with_default_app(folder)


# =========================
#  Utilidades TSV / HTML
# =========================

def norm_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s


def split_tsv_line(line: str):
    return line.rstrip("\r\n").split("\t")


def looks_like_id(x: str) -> bool:
    return bool(re.fullmatch(r"\s*\d+\s*", str(x or "")))


def looks_like_price(x: str) -> bool:
    raw = str(x or "").strip()
    if not raw:
        return False
    raw = re.sub(r"[^\d]", "", raw)
    return raw.isdigit() and int(raw) >= 0


# --- comparaciones/formatos tolerantes (40.000 == 40000) ---

def _digits_only(x: str) -> str:
    return re.sub(r"[^\d]", "", str(x or "").strip())


def _int_from_digits(x: str):
    d = _digits_only(x)
    return int(d) if d.isdigit() else None


def _detect_price_style_from_html_rows(rows: list) -> str:
    """
    Detecta el estilo del 'Valor unitario' mirando el HTML actual:
      - 'dot'   : 40.000
      - 'comma' : 40,000
      - 'plain' : 40000
    """
    dot_pat = re.compile(r"^\d{1,3}(\.\d{3})+$")
    comma_pat = re.compile(r"^\d{1,3}(,\d{3})+$")

    for r in rows or []:
        if not r or len(r) < 4:
            continue
        v = str(r[3] or "").strip()
        if not v:
            continue
        if dot_pat.fullmatch(v):
            return "dot"
        if comma_pat.fullmatch(v):
            return "comma"
        if "." in v:
            return "dot"
        if "," in v:
            return "comma"
    return "plain"


def _format_price_for_html(value, style: str) -> str:
    n = _int_from_digits(value)
    if n is None:
        return str(value or "").strip()
    if style == "dot":
        return f"{n:,}".replace(",", ".")
    if style == "comma":
        return f"{n:,}"
    return str(n)


def _rows_equal_for_diff(before5: list, after5: list) -> bool:
    """
    Compara:
      - columnas 1-3 exactas (texto)
      - valor unitario por número (ignora separadores)
      - id por número
    """
    b = (before5 + [""] * 5)[:5]
    a = (after5 + [""] * 5)[:5]

    if b[:3] != a[:3]:
        return False
    if _int_from_digits(b[3]) != _int_from_digits(a[3]):
        return False
    if _int_from_digits(b[4]) != _int_from_digits(a[4]):
        return False
    return True


def read_text_preserve_bom(path: str):
    data = open(path, "rb").read()
    had_bom = data.startswith(b"\xef\xbb\xbf")
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(enc)
            return text, had_bom, enc
        except UnicodeDecodeError:
            continue
    text = data.decode("utf-8", errors="replace")
    return text, had_bom, "utf-8"


def write_text_preserve_bom(path: str, text: str, had_bom: bool, _encoding_hint: str):
    payload = text.encode("utf-8")
    if had_bom and not payload.startswith(b"\xef\xbb\xbf"):
        payload = b"\xef\xbb\xbf" + payload
    with open(path, "wb") as f:
        f.write(payload)


def find_single_pre_block(html: str):
    matches = list(PRE_REGEX_ALL.finditer(html))
    if not matches:
        return None, 'No encontré <pre id="productos-tsv">...</pre>.'
    if len(matches) > 1:
        return None, 'Encontré más de un <pre id="productos-tsv">. Para seguridad, no haré cambios.'
    m = matches[0]
    open_tag, inner, close_tag = m.group(1), m.group(2), m.group(3)
    inner_start = m.start(2)
    inner_end = m.end(2)
    return (open_tag, inner, close_tag, inner_start, inner_end), None


def validate_tsv_structure(inner: str):
    issues = []
    cleaned = HTML_COMMENT_REGEX.sub("", inner)

    if "<" in cleaned or ">" in cleaned:
        issues.append("El <pre> contiene '<' o '>' (parece HTML incrustado). Por seguridad no se modifica.")

    lines = cleaned.splitlines(True)
    i = 0
    while i < len(lines) and lines[i].strip() == "":
        i += 1
    if i >= len(lines):
        issues.append("El <pre> está vacío.")
        return {"ok": False, "issues": issues}

    header_line = lines[i]
    header_cols = split_tsv_line(header_line)
    if len(header_cols) < 5:
        issues.append("El encabezado no tiene al menos 5 columnas separadas por TAB.")

    norm_cols = [norm_text(c) for c in header_cols[:5]]
    expected = [norm_text(c) for c in EXPECTED_COLS]
    if norm_cols != expected:
        issues.append(
            "El encabezado TSV no coincide con el esperado: "
            "\"Nombre producto\\tCategoria\\tMarca\\tValor unitario\\tid\"."
        )

    j = i + 1
    while j < len(lines) and lines[j].strip() == "":
        j += 1
    if j >= len(lines):
        issues.append("No hay ningún producto después del encabezado.")
        return {"ok": False, "issues": issues, "header_idx": i}

    first_product_line = lines[j]
    fp_cols = split_tsv_line(first_product_line)
    if len(fp_cols) < 5:
        issues.append("La primera fila de producto no tiene al menos 5 columnas TSV.")
    else:
        if not looks_like_price(fp_cols[3]):
            issues.append("El 'Valor unitario' del primer producto no parece un precio válido.")
        if not looks_like_id(fp_cols[4]):
            issues.append("El 'id' del primer producto no es numérico.")

    checked = 0
    for k in range(j, len(lines)):
        if checked >= 50:
            break
        ln = lines[k]
        if not ln.strip():
            continue
        cols = split_tsv_line(ln)
        if len(cols) < 5:
            issues.append(f"Hay una fila con menos de 5 columnas TSV (línea aprox #{k+1}).")
            break
        if not looks_like_id(cols[4]):
            issues.append(f"Hay una fila con id no numérico (línea aprox #{k+1}).")
            break
        checked += 1

    ok = len(issues) == 0
    return {
        "ok": ok,
        "issues": issues,
        "header_idx": i,
        "first_product_idx": j,
        "header_line": header_line.rstrip("\r\n"),
        "first_product_line": first_product_line.rstrip("\r\n"),
        "cleaned_inner": cleaned,
    }


def compute_kept_removed_from_cleaned(cleaned_inner: str):
    lines = cleaned_inner.splitlines(True)

    header_idx = None
    for i, ln in enumerate(lines):
        if ln.strip():
            header_idx = i
            break
    if header_idx is None:
        return cleaned_inner, ""

    first_idx = None
    for j in range(header_idx + 1, len(lines)):
        if lines[j].strip():
            first_idx = j
            break
    if first_idx is None:
        return "".join(lines), ""

    kept = "".join(lines[:first_idx + 1])
    removed = "".join(lines[first_idx + 1:])

    if kept and not kept.endswith("\n") and not kept.endswith("\r\n"):
        kept += "\n"
    return kept, removed


def inner_is_trimmed(inner: str):
    val = validate_tsv_structure(inner)
    if not val.get("ok"):
        return None
    kept, removed = compute_kept_removed_from_cleaned(val["cleaned_inner"])
    return removed.strip() == ""


def make_unique_path(path: str) -> str:
    path = os.path.abspath(path)
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def default_stripped_path(source_html: str) -> str:
    source_html = os.path.abspath(source_html)
    folder = os.path.dirname(source_html)
    name = os.path.basename(source_html)
    stem, ext = os.path.splitext(name)
    if ext.lower() not in (".html", ".htm"):
        ext = ".html"
    candidate = os.path.join(folder, f"{stem}.SIN_CATALOGO{ext}")
    return make_unique_path(candidate)


def sidecar_paths_for_stripped(stripped_html: str):
    base = os.path.splitext(os.path.abspath(stripped_html))[0]
    return base + CATALOG_TSV_SUFFIX, base + CATALOG_META_SUFFIX


def _detect_newline_style(text: str) -> str:
    return "\r\n" if "\r\n" in text else "\n"


# =========================
#  Excel: raíz del programa
# =========================

def program_root_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def default_excel_path() -> str:
    return os.path.join(program_root_dir(), f"{EXCEL_BASE_NAME}.xlsx")


def make_next_excel_path(base_path: str):
    """
    Si base_path existe, devuelve base_1.xlsx, base_2.xlsx, ... (sufijo ascendente)
    Retorna (path, existed_base)
    """
    base_path = os.path.abspath(base_path)
    if not os.path.exists(base_path):
        return base_path, False

    folder = os.path.dirname(base_path)
    filename = os.path.basename(base_path)
    stem, ext = os.path.splitext(filename)
    if not ext:
        ext = ".xlsx"

    n = 1
    while True:
        candidate = os.path.join(folder, f"{stem}_{n}{ext}")
        if not os.path.exists(candidate):
            return candidate, True
        n += 1


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _excel_write_header_style(ws, ncols: int):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = center


def _autosize_columns(ws, max_row: int, max_col: int, min_w=10, max_w=60):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width = min_w
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            v = str(v)
            width = max(width, len(v) + 2)
        width = min(max_w, max(min_w, width))
        ws.column_dimensions[letter].width = width


def _parse_catalog_from_cleaned_inner(cleaned_inner: str):
    lines = [ln for ln in cleaned_inner.splitlines() if ln.strip()]
    if not lines:
        raise RuntimeError("El catálogo está vacío.")
    header_cols = split_tsv_line(lines[0])
    rows = [split_tsv_line(ln) for ln in lines[1:]]
    return header_cols, rows


def export_catalog_html_to_excel(source_html: str, excel_path: str):
    """
    Exporta SOLO las 5 columnas estándar al Excel.
    Si el archivo ya existe, NO se sobrescribe: se crea uno con sufijo _N.
    """
    excel_path, existed = make_next_excel_path(excel_path)

    html, _, _ = read_text_preserve_bom(source_html)
    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, _, _ = block
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se exporta.\n- " + "\n- ".join(val["issues"]))

    header_cols, rows = _parse_catalog_from_cleaned_inner(val["cleaned_inner"])
    if len(header_cols) < 5:
        raise RuntimeError("El encabezado TSV tiene menos de 5 columnas; no se exporta.")
    if [norm_text(c) for c in header_cols[:5]] != [norm_text(c) for c in EXPECTED_COLS]:
        raise RuntimeError("El encabezado TSV no coincide con el esperado; no se exporta.")

    header5 = header_cols[:5]
    rows5 = [r[:5] + ([""] * (5 - len(r))) if len(r) < 5 else r[:5] for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    for c, name in enumerate(header5, start=1):
        ws.cell(row=1, column=c, value=name)

    for r_idx, row in enumerate(rows5, start=2):
        for c_idx, v in enumerate(row, start=1):
            v = "" if v is None else v
            if c_idx == 4:
                raw = re.sub(r"[^\d]", "", str(v))
                ws.cell(row=r_idx, column=c_idx, value=int(raw) if raw.isdigit() else str(v))
            elif c_idx == 5:
                raw = re.sub(r"[^\d]", "", str(v))
                ws.cell(row=r_idx, column=c_idx, value=int(raw) if raw.isdigit() else str(v))
            else:
                ws.cell(row=r_idx, column=c_idx, value=str(v))

    _excel_write_header_style(ws, 5)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(5)}{max(1, len(rows5) + 1)}"
    _autosize_columns(ws, max_row=max(2, len(rows5) + 1), max_col=5)

    wb.save(excel_path)
    return excel_path, len(rows5), existed


def read_catalog_excel_flexible(excel_path: str):
    """
    Lee el Excel aunque:
      - tenga columnas nuevas,
      - el usuario haya movido el orden de las columnas,
      - existan encabezados extra.

    SOLO toma las 5 columnas requeridas por nombre de encabezado (ignorando las demás).
    Devuelve:
      rows_by_id: dict[id] -> [nombre, categoria, marca, valor, id]
      order_ids: ids en el orden de filas del Excel
    """
    if not os.path.exists(excel_path):
        raise RuntimeError("No existe el Excel seleccionado.")

    wb = load_workbook(excel_path)
    ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb.active

    header_map = {}
    header_dups = set()
    for c in range(1, ws.max_column + 1):
        h = _safe_str(ws.cell(row=1, column=c).value)
        if not h:
            continue
        key = norm_text(h)
        if key in header_map:
            header_dups.add(key)
        else:
            header_map[key] = c

    if header_dups:
        req_dups = [k for k in header_dups if k in [norm_text(x) for x in EXPECTED_COLS]]
        if req_dups:
            raise RuntimeError("El Excel tiene encabezados duplicados en columnas requeridas. Corrige eso antes de actualizar.")

    missing = []
    col_idx = []
    for colname in EXPECTED_COLS:
        k = norm_text(colname)
        if k not in header_map:
            missing.append(colname)
        else:
            col_idx.append(header_map[k])

    if missing:
        raise RuntimeError("Faltan encabezados requeridos en el Excel: " + ", ".join(missing))

    idx_nombre, idx_categoria, idx_marca, idx_valor, idx_id = col_idx

    rows_by_id = {}
    order_ids = []
    duplicates = []

    for r in range(2, ws.max_row + 1):
        nombre = _safe_str(ws.cell(row=r, column=idx_nombre).value)
        categoria = _safe_str(ws.cell(row=r, column=idx_categoria).value)
        marca = _safe_str(ws.cell(row=r, column=idx_marca).value)
        valor = _safe_str(ws.cell(row=r, column=idx_valor).value)
        rid = _safe_str(ws.cell(row=r, column=idx_id).value)

        if not any([nombre, categoria, marca, valor, rid]):
            continue

        if not rid:
            raise RuntimeError(f"Hay una fila en Excel sin 'id' (fila {r}).")

        if not looks_like_id(rid):
            raise RuntimeError(f"En el Excel, el id no es numérico en la fila {r}: '{rid}'")

        if re.fullmatch(r"\d+\.\d+", valor):
            raise RuntimeError(f"En el Excel, 'Valor unitario' tiene decimales en la fila {r}: '{valor}'")

        if not looks_like_price(valor):
            raise RuntimeError(f"En el Excel, 'Valor unitario' no parece válido en la fila {r}: '{valor}'")

        row5 = [nombre, categoria, marca, valor, rid]

        if rid in rows_by_id:
            duplicates.append(rid)
        else:
            rows_by_id[rid] = row5
            order_ids.append(rid)

    if duplicates:
        dup_list = ", ".join(duplicates[:20]) + (" ..." if len(duplicates) > 20 else "")
        raise RuntimeError(f"El Excel tiene ids duplicados. Ejemplos: {dup_list}")

    return rows_by_id, order_ids


# =========================
#  Operaciones SIN_CATALOGO
# =========================

def create_new_without_catalog(source_html: str):
    html, had_bom, enc = read_text_preserve_bom(source_html)
    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, inner_start, inner_end = block
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se modifica.\n- " + "\n- ".join(val["issues"]))

    kept, removed = compute_kept_removed_from_cleaned(val["cleaned_inner"])
    if removed.strip() == "":
        raise RuntimeError("Este archivo ya parece estar recortado (no hay catálogo extra para guardar).")

    stripped_path = default_stripped_path(source_html)
    tsv_path, meta_path = sidecar_paths_for_stripped(stripped_path)

    if os.path.exists(tsv_path) or os.path.exists(meta_path):
        stripped_path = make_unique_path(stripped_path)
        tsv_path, meta_path = sidecar_paths_for_stripped(stripped_path)

    new_html = html[:inner_start] + kept + html[inner_end:]
    write_text_preserve_bom(stripped_path, new_html, had_bom, enc)

    with open(tsv_path, "w", encoding="utf-8", newline="") as f:
        f.write(removed)

    meta = {
        "version": 1,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source_html_path": os.path.abspath(source_html),
        "stripped_html_path": os.path.abspath(stripped_path),
        "tsv_path": os.path.abspath(tsv_path),
        "header_line": val["header_line"],
        "first_product_line": val["first_product_line"],
        "removed_lines_count": len([ln for ln in removed.splitlines() if ln.strip()]),
    }
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    return stripped_path, tsv_path, meta_path, meta["removed_lines_count"]


def restore_catalog_to_stripped(stripped_html: str):
    stripped_html = os.path.abspath(stripped_html)
    tsv_path, meta_path = sidecar_paths_for_stripped(stripped_html)

    if not os.path.exists(meta_path):
        raise RuntimeError("No encuentro el archivo de metadatos .CATALOGO.json junto al SIN_CATALOGO.")
    if not os.path.exists(tsv_path):
        raise RuntimeError("No encuentro el archivo .CATALOGO.tsv junto al SIN_CATALOGO.")

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    if meta.get("version") != 1:
        raise RuntimeError("Versión de metadatos desconocida. Por seguridad no se restaura.")

    html, had_bom, enc = read_text_preserve_bom(stripped_html)
    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, inner_start, inner_end = block
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se modifica.\n- " + "\n- ".join(val["issues"]))

    state = inner_is_trimmed(inner)
    if state is None:
        raise RuntimeError("No se puede determinar estado (estructura inválida).")
    if state is False:
        return "Este archivo ya está COMPLETO (parece que el catálogo ya está puesto)."

    b_header = norm_text(meta.get("header_line", ""))
    b_first = norm_text(meta.get("first_product_line", ""))
    c_header = norm_text(val.get("header_line", ""))
    c_first = norm_text(val.get("first_product_line", ""))

    if b_header != c_header or b_first != c_first:
        raise RuntimeError(
            "El encabezado o el primer producto NO coinciden con los metadatos guardados.\n"
            "Esto evita pegar catálogo en un archivo equivocado."
        )

    removed_text = open(tsv_path, "r", encoding="utf-8", errors="replace").read()
    if not removed_text.strip():
        raise RuntimeError("El .CATALOGO.tsv está vacío. No hay nada para restaurar.")

    sample_lines = [ln for ln in removed_text.splitlines() if ln.strip()][:10]
    for ln in sample_lines[:5]:
        cols = ln.split("\t")
        if len(cols) < 5 or not looks_like_id(cols[4]):
            raise RuntimeError("El .CATALOGO.tsv no parece TSV válido. Por seguridad no se restaura.")

    kept, _ = compute_kept_removed_from_cleaned(val["cleaned_inner"])
    new_inner = kept + removed_text
    new_html = html[:inner_start] + new_inner + html[inner_end:]

    tmp_path = stripped_html + ".tmp_write"
    write_text_preserve_bom(tmp_path, new_html, had_bom, enc)
    os.replace(tmp_path, stripped_html)

    return "Catálogo restaurado correctamente en el archivo SIN_CATALOGO."


# =========================
#  Actualizar HTML fuente desde Excel
# =========================

def _parse_html_catalog(inner: str):
    val = validate_tsv_structure(inner)
    if not val["ok"]:
        raise RuntimeError("Estructura inválida, no se puede actualizar.\n- " + "\n- ".join(val["issues"]))
    header_cols, rows = _parse_catalog_from_cleaned_inner(val["cleaned_inner"])
    if len(header_cols) < 5:
        raise RuntimeError("El HTML fuente no tiene al menos 5 columnas en el encabezado.")
    if [norm_text(c) for c in header_cols[:5]] != [norm_text(c) for c in EXPECTED_COLS]:
        raise RuntimeError("El encabezado del HTML fuente no coincide con el esperado.")
    html_map = {}
    dups = []
    for r in rows:
        if len(r) < 5:
            continue
        rid = str(r[4]).strip()
        if not rid:
            continue
        if rid in html_map:
            dups.append(rid)
        else:
            html_map[rid] = r
    if dups:
        raise RuntimeError("El HTML fuente tiene ids duplicados. No se actualiza por seguridad.")
    return header_cols, rows, html_map


def _tsv_line(cols: list[str], nl: str) -> str:
    return "\t".join([str(x) for x in cols]).rstrip("\r\n") + nl


def compute_diffs_and_build_new_html_exact(source_html: str, excel_rows_by_id: dict, excel_order_ids: list):
    html, had_bom, enc = read_text_preserve_bom(source_html)
    nl = _detect_newline_style(html)

    block, err = find_single_pre_block(html)
    if err:
        raise RuntimeError(err)

    _, inner, _, inner_start, inner_end = block
    header_cols, html_rows, html_map = _parse_html_catalog(inner)

    price_style = _detect_price_style_from_html_rows(html_rows)

    html_ids = set(html_map.keys())
    excel_ids = set(excel_rows_by_id.keys())

    diffs = []
    modified = 0
    added = 0
    deleted = 0

    for rid in sorted(html_ids - excel_ids, key=lambda x: int(re.sub(r"[^\d]", "", x) or "0")):
        before5 = (html_map[rid] + [""] * 5)[:5]
        diffs.append({
            "type": "ELIMINADO",
            "id": rid,
            "before": "\t".join([str(x) for x in before5]),
            "after": "(se elimina del HTML)",
        })
        deleted += 1

    for rid in excel_order_ids:
        row5 = excel_rows_by_id.get(rid)
        if not row5:
            continue

        after5 = (list(row5) + [""] * 5)[:5]
        after5[4] = rid

        after5_display = list(after5)
        after5_display[3] = _format_price_for_html(after5_display[3], price_style)

        if rid not in html_map:
            diffs.append({
                "type": "NUEVO",
                "id": rid,
                "before": "(no existía en HTML)",
                "after": "\t".join([str(x) for x in after5_display]),
            })
            added += 1
        else:
            before_row = html_map[rid]
            before5 = (before_row + [""] * 5)[:5]

            if not _rows_equal_for_diff(before5, after5):
                diffs.append({
                    "type": "MODIFICADO",
                    "id": rid,
                    "before": "\t".join([str(x) for x in before5]),
                    "after": "\t".join([str(x) for x in after5_display]),
                })
                modified += 1

    if not diffs:
        return diffs, None, {"modified": 0, "new": 0, "deleted": 0}

    out = []
    out.append(_tsv_line(header_cols[:5], nl))

    for rid in excel_order_ids:
        row5 = excel_rows_by_id.get(rid)
        if not row5:
            continue

        row5_out = (list(row5) + [""] * 5)[:5]
        row5_out[4] = rid
        row5_out[3] = _format_price_for_html(row5_out[3], price_style)
        out.append(_tsv_line(row5_out, nl))

    new_inner = "".join(out)
    new_html = html[:inner_start] + new_inner + html[inner_end:]
    return diffs, (new_html, had_bom, enc), {"modified": modified, "new": added, "deleted": deleted}


def apply_overwrite_html(source_html: str, new_html_tuple, make_backup: bool = False):
    new_html, had_bom, enc = new_html_tuple

    if make_backup and os.path.exists(source_html):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{source_html}.bak_{ts}"
        try:
            shutil.copy2(source_html, backup_path)
        except Exception:
            pass

    tmp_path = source_html + ".tmp_write"
    write_text_preserve_bom(tmp_path, new_html, had_bom, enc)
    os.replace(tmp_path, source_html)


# =========================
#  UI Tkinter (mejorada)
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.settings = load_settings()

        self.title("Catálogo TSV (HTML ↔ Excel) — SIN_CATALOGO")
        self.geometry("1040x720")
        self.minsize(980, 660)
        self.resizable(True, True)

        self.source_var = tk.StringVar(value="")
        self.stripped_var = tk.StringVar(value="")
        self.tsv_var = tk.StringVar(value="")
        self.meta_var = tk.StringVar(value="")
        self.excel_var = tk.StringVar(value="")

        self.status_var = tk.StringVar(value="Selecciona el catalogo.html fuente (con catálogo) o usa el último recordado.")
        self.state_var = tk.StringVar(value="Estado: (sin archivo)")
        self.backup_var = tk.BooleanVar(value=False)

        self._refresh_job = None

        self._setup_style()
        self._build_menu()
        self._build_ui()

        self._load_last_paths()
        self._wire_auto_refresh()
        self.refresh_state()

    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self._bg = "#f6f7fb"
        self._card = "#ffffff"
        self._text = "#111827"
        self._muted = "#4b5563"
        self._accent = "#2563eb"
        self._accent2 = "#1d4ed8"
        self._danger = "#b91c1c"

        self.configure(bg=self._bg)

        style.configure("App.TFrame", background=self._bg)
        style.configure("Card.TLabelframe", background=self._card, padding=(12, 10))
        style.configure("Card.TLabelframe.Label", background=self._card, foreground=self._text, font=("Segoe UI", 10, "bold"))
        style.configure("Card.TFrame", background=self._card)

        style.configure("Title.TLabel", background=self._bg, foreground=self._text, font=("Segoe UI", 16, "bold"))
        style.configure("Sub.TLabel", background=self._bg, foreground=self._muted, font=("Segoe UI", 10))
        style.configure("Info.TLabel", background=self._card, foreground=self._muted, font=("Segoe UI", 10))
        style.configure("State.TLabel", background=self._card, foreground="#0b5394", font=("Segoe UI", 10, "bold"))

        style.configure("TLabel", background=self._card, foreground=self._text)
        style.configure("TEntry", padding=(6, 4))
        style.configure("TButton", padding=(10, 7))

        style.configure("Accent.TButton", foreground="white", background=self._accent, font=("Segoe UI", 10, "bold"))
        style.map("Accent.TButton",
                  background=[("active", self._accent2), ("pressed", self._accent2)],
                  foreground=[("disabled", "#e5e7eb")])

        style.configure("Danger.TButton", foreground="white", background=self._danger, font=("Segoe UI", 10, "bold"))
        style.map("Danger.TButton",
                  background=[("active", "#991b1b"), ("pressed", "#7f1d1d")],
                  foreground=[("disabled", "#e5e7eb")])

        style.configure("Status.TFrame", background=self._bg)
        style.configure("Status.TLabel", background=self._bg, foreground=self._muted, font=("Segoe UI", 9))

    def _build_menu(self):
        menubar = tk.Menu(self)

        m_file = tk.Menu(menubar, tearoff=0)
        m_file.add_command(label="Seleccionar FUENTE…", command=self.pick_source)
        m_file.add_command(label="Seleccionar SIN_CATALOGO…", command=self.pick_stripped)
        m_file.add_command(label="Seleccionar Excel…", command=self.pick_excel)
        m_file.add_separator()
        m_file.add_command(label="Abrir carpeta de trabajo", command=self.on_open_folder)
        m_file.add_separator()
        m_file.add_command(label="Salir", command=self.destroy)
        menubar.add_cascade(label="Archivo", menu=m_file)

        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label="Acerca de…", command=self._about)
        menubar.add_cascade(label="Ayuda", menu=m_help)

        self.config(menu=menubar)

    def _about(self):
        messagebox.showinfo(
            "Acerca de",
            "Herramienta para:\n"
            "• Crear SIN_CATALOGO (recorta el catálogo y lo guarda como sidecar TSV/JSON)\n"
            "• Restaurar catálogo al SIN_CATALOGO\n"
            "• Exportar catálogo del HTML fuente a Excel\n"
            "• Actualizar el HTML fuente desde Excel (usando solo 5 columnas estándar)"
        )

    def _build_ui(self):
        root = ttk.Frame(self, style="App.TFrame")
        root.pack(fill="both", expand=True, padx=14, pady=12)

        header = ttk.Frame(root, style="App.TFrame")
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Catálogo TSV (HTML ↔ Excel)", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="Interfaz mejorada · Validación visible · Vista previa con cambios en rojo", style="Sub.TLabel").pack(anchor="w", pady=(2, 0))

        content = ttk.Frame(root, style="App.TFrame")
        content.pack(fill="both", expand=True)

        left = ttk.Frame(content, style="App.TFrame")
        right = ttk.Frame(content, style="App.TFrame")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        right.grid(row=0, column=1, sticky="nsew")

        content.columnconfigure(0, weight=3)
        content.columnconfigure(1, weight=2)
        content.rowconfigure(0, weight=1)

        lf_paths = ttk.LabelFrame(left, text="Rutas principales", style="Card.TLabelframe")
        lf_paths.pack(fill="x", pady=(0, 10))

        self._build_path_row(
            lf_paths, 0, "Fuente (con catálogo):", self.source_var,
            browse=self.pick_source, open_cmd=lambda: self.open_path(self.source_var.get()),
            folder_cmd=lambda: self.open_folder(self.source_var.get())
        )
        self._build_path_row(
            lf_paths, 1, "Archivo SIN_CATALOGO:", self.stripped_var,
            browse=self.pick_stripped, open_cmd=lambda: self.open_path(self.stripped_var.get()),
            folder_cmd=lambda: self.open_folder(self.stripped_var.get())
        )
        self._build_path_row(
            lf_paths, 2, "Excel del catálogo:", self.excel_var,
            browse=self.pick_excel, open_cmd=lambda: self.open_path(self.excel_var.get()),
            folder_cmd=lambda: self.open_folder(self.excel_var.get())
        )

        lf_side = ttk.LabelFrame(left, text="Archivos asociados (automáticos)", style="Card.TLabelframe")
        lf_side.pack(fill="x", pady=(0, 10))

        self._build_readonly_row(
            lf_side, 0, "Catálogo guardado (.CATALOGO.tsv):", self.tsv_var,
            open_cmd=lambda: self.open_path(self.tsv_var.get()), folder_cmd=lambda: self.open_folder(self.tsv_var.get())
        )
        self._build_readonly_row(
            lf_side, 1, "Metadatos (.CATALOGO.json):", self.meta_var,
            open_cmd=lambda: self.open_path(self.meta_var.get()), folder_cmd=lambda: self.open_folder(self.meta_var.get())
        )

        lf_ops = ttk.LabelFrame(left, text="Operaciones", style="Card.TLabelframe")
        lf_ops.pack(fill="x")

        ops_grid = ttk.Frame(lf_ops, style="Card.TFrame")
        ops_grid.pack(fill="x")

        ttk.Button(ops_grid, text="Crear SIN_CATALOGO", style="Accent.TButton", command=self.on_create).grid(row=0, column=0, sticky="we", padx=(0, 8), pady=(0, 8))
        ttk.Button(ops_grid, text="Restaurar catálogo", command=self.on_restore).grid(row=0, column=1, sticky="we", pady=(0, 8))

        ttk.Button(ops_grid, text="Exportar a Excel (desde FUENTE)", command=self.on_export_excel).grid(row=1, column=0, sticky="we", padx=(0, 8), pady=(0, 8))
        ttk.Button(ops_grid, text="Abrir carpeta", command=self.on_open_folder).grid(row=1, column=1, sticky="we", pady=(0, 8))

        ttk.Separator(ops_grid, orient="horizontal").grid(row=2, column=0, columnspan=2, sticky="we", pady=(2, 10))

        update_row = ttk.Frame(ops_grid, style="Card.TFrame")
        update_row.grid(row=3, column=0, columnspan=2, sticky="we")
        ttk.Checkbutton(update_row, text="Crear copia .bak antes de sobrescribir el HTML fuente", variable=self.backup_var).pack(anchor="w", pady=(0, 8))
        ttk.Button(update_row, text="Actualizar FUENTE desde Excel (copia fiel, ignora columnas extra)", style="Danger.TButton",
                   command=self.on_update_source_from_excel).pack(fill="x")

        ops_grid.columnconfigure(0, weight=1)
        ops_grid.columnconfigure(1, weight=1)

        lf_state = ttk.LabelFrame(right, text="Estado y validación", style="Card.TLabelframe")
        lf_state.pack(fill="both", expand=True)

        ttk.Label(lf_state, textvariable=self.state_var, style="State.TLabel").pack(fill="x", pady=(0, 8))

        ttk.Label(lf_state, text="Detalles / advertencias:", style="Info.TLabel").pack(anchor="w")
        self.txt_valid = ScrolledText(lf_state, height=14, wrap="word")
        self.txt_valid.pack(fill="both", expand=True, pady=(6, 8))
        self.txt_valid.configure(state="disabled", font=("Segoe UI", 9))

        btns = ttk.Frame(lf_state, style="Card.TFrame")
        btns.pack(fill="x")
        ttk.Button(btns, text="Refrescar estado", command=self.refresh_state).pack(side="left")
        ttk.Button(btns, text="Copiar estado", command=self.copy_state).pack(side="left", padx=(8, 0))

        status = ttk.Frame(root, style="Status.TFrame")
        status.pack(fill="x", pady=(10, 0))
        ttk.Label(status, textvariable=self.status_var, style="Status.TLabel").pack(side="left", fill="x", expand=True)
        ttk.Sizegrip(status).pack(side="right")

    def _build_path_row(self, parent, row, label, var, browse, open_cmd, folder_cmd):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.grid(row=row, column=0, sticky="we", pady=(0, 10))
        parent.columnconfigure(0, weight=1)

        ttk.Label(frame, text=label).grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(frame, textvariable=var).grid(row=1, column=0, columnspan=5, sticky="we")

        ttk.Button(frame, text="Seleccionar…", command=browse).grid(row=1, column=5, padx=(8, 0))
        ttk.Button(frame, text="Abrir", command=open_cmd).grid(row=1, column=6, padx=(8, 0))
        ttk.Button(frame, text="Carpeta", command=folder_cmd).grid(row=1, column=7, padx=(8, 0))
        ttk.Button(frame, text="Copiar", command=lambda v=var: self.copy_to_clipboard(v.get())).grid(row=1, column=8, padx=(8, 0))

        frame.columnconfigure(0, weight=1)

    def _build_readonly_row(self, parent, row, label, var, open_cmd, folder_cmd):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.grid(row=row, column=0, sticky="we", pady=(0, 10))
        parent.columnconfigure(0, weight=1)

        ttk.Label(frame, text=label).grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(frame, textvariable=var, state="readonly").grid(row=1, column=0, columnspan=4, sticky="we")
        ttk.Button(frame, text="Abrir", command=open_cmd).grid(row=1, column=4, padx=(8, 0))
        ttk.Button(frame, text="Carpeta", command=folder_cmd).grid(row=1, column=5, padx=(8, 0))
        ttk.Button(frame, text="Copiar", command=lambda v=var: self.copy_to_clipboard(v.get())).grid(row=1, column=6, padx=(8, 0))

        frame.columnconfigure(0, weight=1)

    def _wire_auto_refresh(self):
        def schedule(*_):
            if self._refresh_job is not None:
                try:
                    self.after_cancel(self._refresh_job)
                except Exception:
                    pass
            self._refresh_job = self.after(350, self.refresh_state)

        for v in (self.source_var, self.stripped_var, self.excel_var):
            v.trace_add("write", schedule)

    def _load_last_paths(self):
        last_source = (self.settings.get("last_source") or "").strip()
        last_stripped = (self.settings.get("last_stripped") or "").strip()
        last_excel = (self.settings.get("last_excel") or "").strip()

        if last_source and os.path.exists(last_source):
            self.source_var.set(last_source)

        if last_stripped and os.path.exists(last_stripped):
            self.stripped_var.set(last_stripped)
            self._sync_sidecar_from_stripped()

        if last_excel and os.path.exists(last_excel):
            self.excel_var.set(last_excel)
        else:
            self.excel_var.set(default_excel_path())

    def copy_to_clipboard(self, text: str):
        text = (text or "").strip()
        if not text:
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self.status_var.set("Copiado al portapapeles.")

    def copy_state(self):
        self.copy_to_clipboard(self.state_var.get())

    def open_path(self, path: str):
        try:
            open_with_default_app(path.strip())
        except Exception as e:
            messagebox.showerror("No se pudo abrir", str(e))

    def open_folder(self, path: str):
        try:
            open_folder_of(path.strip())
        except Exception as e:
            messagebox.showerror("No se pudo abrir carpeta", str(e))

    def _sync_sidecar_from_stripped(self):
        stripped = self.stripped_var.get().strip()
        if stripped:
            tsv_path, meta_path = sidecar_paths_for_stripped(stripped)
            self.tsv_var.set(tsv_path)
            self.meta_var.set(meta_path)
        else:
            self.tsv_var.set("")
            self.meta_var.set("")

    def pick_source(self):
        initialdir = None
        last_dir = (self.settings.get("last_dir") or "").strip()
        if last_dir and os.path.isdir(last_dir):
            initialdir = last_dir

        path = filedialog.askopenfilename(
            title="Selecciona el catalogo.html FUENTE (con catálogo)",
            initialdir=initialdir,
            filetypes=[("HTML", "*.html;*.htm"), ("Todos", "*.*")]
        )
        if path:
            self.source_var.set(path)
            self.settings = remember_paths(self.settings, source_html=path)
            self.status_var.set("Fuente seleccionada. Puedes crear el SIN_CATALOGO o exportar a Excel.")
            self.refresh_state()

    def pick_stripped(self):
        initialdir = None
        last_dir = (self.settings.get("last_dir") or "").strip()
        if last_dir and os.path.isdir(last_dir):
            initialdir = last_dir

        path = filedialog.askopenfilename(
            title="Selecciona el archivo SIN_CATALOGO.html",
            initialdir=initialdir,
            filetypes=[("HTML", "*.html;*.htm"), ("Todos", "*.*")]
        )
        if path:
            self.stripped_var.set(path)
            self._sync_sidecar_from_stripped()
            self.settings = remember_paths(self.settings, stripped_html=path, tsv_path=self.tsv_var.get(), meta_path=self.meta_var.get())
            self.status_var.set("Archivo SIN_CATALOGO seleccionado. Puedes restaurar si existen los sidecar.")
            self.refresh_state()

    def pick_excel(self):
        initialdir = program_root_dir()
        path = filedialog.askopenfilename(
            title="Selecciona el Excel del catálogo",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if path:
            self.excel_var.set(path)
            self.settings = remember_paths(self.settings, excel_path=path)
            self.status_var.set("Excel seleccionado.")
            self.refresh_state()

    def on_open_folder(self):
        path = self.stripped_var.get().strip() or self.source_var.get().strip() or self.excel_var.get().strip() or program_root_dir()
        try:
            open_folder_of(path)
        except Exception as e:
            messagebox.showerror("No se pudo abrir carpeta", str(e))

    def _describe_html_state(self, html_path: str):
        if not html_path or not os.path.exists(html_path):
            return "(no existe)", ""

        try:
            html, _, _ = read_text_preserve_bom(html_path)
            block, err = find_single_pre_block(html)
            if err:
                return "NO DISPONIBLE", err

            _, inner, _, _, _ = block
            val = validate_tsv_structure(inner)
            if not val["ok"]:
                return "NO VÁLIDO", "Validación:\n- " + "\n- ".join(val["issues"])

            trimmed = inner_is_trimmed(inner)
            if trimmed is True:
                return "RECORTADO (solo 1 producto)", ""
            return "COMPLETO (muchos productos)", ""
        except Exception as e:
            return "ERROR", str(e)

    def _set_valid_text(self, text: str):
        self.txt_valid.configure(state="normal")
        self.txt_valid.delete("1.0", "end")
        if text.strip():
            self.txt_valid.insert("1.0", text.strip())
        self.txt_valid.configure(state="disabled")

    def refresh_state(self):
        self._sync_sidecar_from_stripped()

        src = self.source_var.get().strip()
        dst = self.stripped_var.get().strip()

        src_state, src_msg = self._describe_html_state(src)
        dst_state, dst_msg = self._describe_html_state(dst)

        tsv = self.tsv_var.get().strip()
        meta = self.meta_var.get().strip()
        excel = self.excel_var.get().strip()

        tsv_ok = "OK" if (tsv and os.path.exists(tsv)) else "NO"
        meta_ok = "OK" if (meta and os.path.exists(meta)) else "NO"
        excel_ok = "OK" if (excel and os.path.exists(excel)) else "NO"

        lines = [
            f"Fuente: {src_state}",
            f"SIN_CATALOGO: {dst_state}",
            f"Sidecar TSV: {tsv_ok} | Sidecar JSON: {meta_ok}",
            f"Excel: {excel_ok}",
        ]
        self.state_var.set("Estado: " + "   |   ".join(lines))

        warn = []
        if src_msg:
            warn.append("FUENTE:\n" + src_msg)
        if dst_msg:
            warn.append("SIN_CATALOGO:\n" + dst_msg)
        self._set_valid_text("\n\n".join(warn))

    # =========================
    #  Preview con cambios en ROJO en "DESPUÉS"
    # =========================

    def _diff_changed_word_indices(self, before: str, after: str) -> set[int]:
        """
        Devuelve el conjunto de índices (por palabra) que en AFTER son distintos,
        comparando BEFORE vs AFTER a nivel de tokens (separados por whitespace).
        """
        b_words = re.findall(r"\S+", before or "")
        a_words = re.findall(r"\S+", after or "")

        sm = difflib.SequenceMatcher(a=b_words, b=a_words)
        changed_a = set()

        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag in ("replace", "insert"):
                for k in range(j1, j2):
                    changed_a.add(k)

        return changed_a

    def _insert_with_red_changed(self, txt: tk.Text, before: str, after: str, red_tag: str):
        """
        Inserta AFTER preservando espacios/tabs, pintando en rojo las palabras
        que difieren respecto a BEFORE.
        """
        changed = self._diff_changed_word_indices(before, after)
        parts = re.split(r"(\s+)", after or "")
        word_idx = 0

        for part in parts:
            if part == "":
                continue
            if part.isspace():
                txt.insert("end", part)
            else:
                if word_idx in changed:
                    txt.insert("end", part, red_tag)
                else:
                    txt.insert("end", part)
                word_idx += 1

    def _show_preview_and_confirm(self, diffs: list, summary: dict):
        top = tk.Toplevel(self)
        top.title("Vista previa de cambios (Excel → HTML fuente)")
        top.geometry("980x560")
        top.minsize(900, 520)
        top.transient(self)
        top.grab_set()

        frm = ttk.Frame(top, padding=12)
        frm.pack(fill="both", expand=True)

        header = (
            "Sincronización EXACTA (copia fiel) Excel → HTML usando SOLO 5 columnas estándar.\n"
            "Columnas nuevas del Excel se IGNORAN (no se pasan al HTML).\n\n"
            f"- MODIFICADOS: {summary.get('modified', 0)}\n"
            f"- NUEVOS:      {summary.get('new', 0)}\n"
            f"- ELIMINADOS:  {summary.get('deleted', 0)}\n\n"
        )
        if self.backup_var.get():
            header += "• Se creará una copia .bak antes de sobrescribir el HTML.\n"
        else:
            header += "• No se creará copia .bak (opción desactivada).\n"
        header += "-" * 92 + "\n\n"

        txt = ScrolledText(frm, wrap="none", height=24)
        txt.pack(fill="both", expand=True)

        txt.tag_configure("redchg", foreground="#b91c1c")  # rojo
        txt.tag_configure("muted", foreground="#374151")   # gris

        txt.insert("1.0", header, ("muted",))

        max_show = 300
        for i, d in enumerate(diffs[:max_show], start=1):
            tipo = d.get("type", "")
            rid = d.get("id", "")
            before_line = d.get("before", "")
            after_line = d.get("after", "")

            txt.insert("end", f"[{i}] {tipo} | id={rid}\n")

            txt.insert("end", "    ANTES:   ")
            txt.insert("end", f"{before_line}\n")

            txt.insert("end", "    DESPUÉS: ")

            if tipo == "MODIFICADO":
                self._insert_with_red_changed(txt, before_line, after_line, "redchg")
                txt.insert("end", "\n\n")
            elif tipo == "NUEVO":
                txt.insert("end", after_line, "redchg")
                txt.insert("end", "\n\n")
            else:
                txt.insert("end", f"{after_line}\n\n")

        if len(diffs) > max_show:
            txt.insert("end", f"(Hay {len(diffs)} filas afectadas; se muestran solo las primeras {max_show}.)\n", ("muted",))

        txt.configure(state="disabled", font=("Consolas", 9))

        decision = tk.StringVar(value="")
        btn_frame = ttk.Frame(frm)
        btn_frame.pack(pady=(10, 0), fill="x")

        def accept():
            decision.set("yes")
            top.destroy()

        def cancel():
            decision.set("no")
            top.destroy()

        ttk.Button(btn_frame, text="Aplicar cambios", style="Danger.TButton", command=accept).pack(side="left")
        ttk.Button(btn_frame, text="Cancelar", command=cancel).pack(side="left", padx=(10, 0))

        self.wait_variable(decision)
        return decision.get() == "yes"

    # =========================
    # Acciones principales
    # =========================

    def on_create(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return

        try:
            stripped_path, tsv_path, meta_path, removed_count = create_new_without_catalog(src)

            self.stripped_var.set(stripped_path)
            self.tsv_var.set(tsv_path)
            self.meta_var.set(meta_path)

            self.settings = remember_paths(
                self.settings,
                source_html=src,
                stripped_html=stripped_path,
                tsv_path=tsv_path,
                meta_path=meta_path
            )

            msg = (
                "Creado SIN_CATALOGO sin modificar la fuente.\n\n"
                f"- SIN_CATALOGO: {os.path.basename(stripped_path)}\n"
                f"- Catálogo guardado: {os.path.basename(tsv_path)}\n"
                f"- Filas guardadas (aprox): {removed_count}\n"
            )
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)
        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))

    def on_restore(self):
        dst = self.stripped_var.get().strip()
        if not dst:
            messagebox.showwarning("Falta SIN_CATALOGO", "Selecciona (o crea) primero el archivo SIN_CATALOGO.")
            return
        if not os.path.exists(dst):
            messagebox.showerror("No existe", "El archivo SIN_CATALOGO no existe.")
            return

        self._sync_sidecar_from_stripped()
        self.settings = remember_paths(self.settings, stripped_html=dst, tsv_path=self.tsv_var.get(), meta_path=self.meta_var.get())

        try:
            msg = restore_catalog_to_stripped(dst)
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)
        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))

    def on_export_excel(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return

        try:
            base_excel = default_excel_path()
            excel_path, rows_count, existed = export_catalog_html_to_excel(src, base_excel)

            self.excel_var.set(excel_path)
            self.settings = remember_paths(self.settings, excel_path=excel_path)

            if existed:
                msg = (
                    f"Ya existía '{os.path.basename(base_excel)}'.\n"
                    f"Para no sobrescribirlo, se creó: '{os.path.basename(excel_path)}'.\n\n"
                    f"Filas exportadas: {rows_count}"
                )
            else:
                msg = f"Excel creado: '{os.path.basename(excel_path)}'\n\nFilas exportadas: {rows_count}"

            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Exportación completada", msg)

        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))

    def on_update_source_from_excel(self):
        src = self.source_var.get().strip()
        if not src:
            messagebox.showwarning("Falta fuente", "Selecciona primero el archivo FUENTE (catalogo.html).")
            return
        if not os.path.exists(src):
            messagebox.showerror("No existe", "El archivo FUENTE no existe.")
            return

        excel_path = self.excel_var.get().strip()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Falta Excel", "Selecciona un Excel válido primero.")
            return

        try:
            excel_rows_by_id, excel_order_ids = read_catalog_excel_flexible(excel_path)

            diffs, new_html_tuple, summary = compute_diffs_and_build_new_html_exact(
                src, excel_rows_by_id, excel_order_ids
            )

            if not diffs:
                msg = "No hay cambios: el HTML ya coincide con el Excel (considerando SOLO las 5 columnas estándar)."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Sin cambios", msg)
                return

            ok = self._show_preview_and_confirm(diffs, summary)
            if not ok:
                msg = "Operación cancelada. No se modificó el HTML fuente."
                self.status_var.set(msg)
                self.refresh_state()
                messagebox.showinfo("Cancelado", msg)
                return

            apply_overwrite_html(src, new_html_tuple, make_backup=self.backup_var.get())

            msg = (
                "Actualización aplicada al HTML fuente desde el Excel.\n\n"
                f"- Modificados: {summary.get('modified', 0)}\n"
                f"- Nuevos: {summary.get('new', 0)}\n"
                f"- Eliminados: {summary.get('deleted', 0)}\n\n"
                "Columnas nuevas del Excel fueron ignoradas (no se copiaron al HTML)."
            )
            self.status_var.set(msg)
            self.refresh_state()
            messagebox.showinfo("Hecho", msg)

        except Exception as e:
            self.status_var.set(str(e))
            self.refresh_state()
            messagebox.showerror("Bloqueado por seguridad", str(e))


if __name__ == "__main__":
    App().mainloop()
