# image_renamer_single.py
import json
import os
import re
import sys
import csv
import unicodedata
import subprocess
from bisect import bisect_left
from pathlib import Path
from typing import List

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont

# =========================
# CONFIG (antes config.py)
# =========================

BASE_DIR = Path(__file__).resolve().parent
ARCHIVOS_DIR = BASE_DIR / "archivos"
ARCHIVOS_DIR.mkdir(parents=True, exist_ok=True)

CONFIG_FILE = ARCHIVOS_DIR / "image_renamer_config.json"

FILENAME_PART_SEPARATOR = "_"

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp",
    ".tif", ".tiff", ".webp",
}


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = str(text).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text


def sanitize_filename(text: str) -> str:
    text = str(text)
    return re.sub(r"[^A-Za-z0-9_-]", "_", text)

def to_underscore_text(text: str) -> str:
    """
    Convierte texto a formato consistente con guion bajo:
    - Reemplaza caracteres no permitidos por '_'
    - Colapsa múltiples '_' seguidos a uno solo
    Nota: NO hace .strip("_") para conservar '_' si vienen del saneado.
    """
    s = "" if text is None else str(text)
    s = sanitize_filename(s)
    s = re.sub(r"_+", "_", s)
    return s



# =========================
# UI WIDGETS (antes ui_widgets.py)
# =========================

class ToolTip:
    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tipwindow = None
        self.id = None
        widget.bind("<Enter>", self._enter)
        widget.bind("<Leave>", self._leave)
        widget.bind("<ButtonPress>", self._leave)

    def _enter(self, event=None):
        self._schedule()

    def _leave(self, event=None):
        self._unschedule()
        self._hidetip()

    def _schedule(self):
        self._unschedule()
        self.id = self.widget.after(self.delay, self._showtip)

    def _unschedule(self):
        if self.id is not None:
            self.widget.after_cancel(self.id)
            self.id = None

    def _showtip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 1
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("tahoma", 8, "normal"),
        )
        label.pack(ipadx=1, ipady=1)

    def _hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw is not None:
            tw.destroy()


def create_tooltip(widget, text: str):
    ToolTip(widget, text)


# =========================
# ARCHIVOS (antes archivos.py)
# =========================

def gather_all_files(root: Path) -> List[Path]:
    if not root or not root.is_dir():
        return []
    try:
        files = [p for p in root.rglob("*") if p.is_file()]
    except Exception as e:
        print("Error leyendo carpeta:", e)
        files = []
    return sorted(files, key=lambda x: x.name.lower())


def looks_like_product_code(prefix: str) -> bool:
    if prefix is None:
        return False
    s = str(prefix).strip()
    if not s:
        return False
    if len(s) > 40:
        return False
    if not any(ch.isdigit() for ch in s):
        return False
    return True


def extract_code_key_from_stem(stem: str) -> str:
    stem_sanitized = sanitize_filename(stem)
    if not stem_sanitized:
        return ""
    sep = FILENAME_PART_SEPARATOR or "_"

    if sep and sep in stem_sanitized:
        prefix = stem_sanitized.split(sep, 1)[0]
        if sep == "-":
            return prefix if looks_like_product_code(prefix) else ""
        return prefix

    if sep != "_" and "_" in stem_sanitized:
        return stem_sanitized.split("_", 1)[0]

    if sep != "-" and "-" in stem_sanitized:
        prefix = stem_sanitized.split("-", 1)[0]
        if looks_like_product_code(prefix):
            return prefix

    return stem_sanitized


# =========================
# PRODUCTOS (antes productos.py)
# =========================

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    load_workbook = None
    Alignment = None


def get_excel_sheet_names(path: str):
    path_obj = Path(path)
    ext = path_obj.suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        return []
    if not HAS_OPENPYXL:
        return []
    try:
        wb = load_workbook(filename=str(path_obj), read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def _column_label_default(i: int) -> str:
    return f"Columna {i + 1}"


def _sniff_csv_dialect(path: Path):
    try:
        sample = path.read_text(encoding="utf-8-sig", errors="ignore")[:4096]
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample, delimiters=";,\t|")
        return dialect
    except Exception:
        return csv.excel


def get_table_preview(
    path: str,
    sheet_name: str = None,
    has_headers: bool = True,
    header_row_number: int = 1,
    max_rows: int = 20
):
    if header_row_number < 1:
        header_row_number = 1

    path_obj = Path(path)
    ext = path_obj.suffix.lower()
    if ext == ".csv":
        return _get_csv_preview(path_obj, has_headers, header_row_number, max_rows)
    if ext in {".xlsx", ".xlsm"}:
        return _get_xlsx_preview(path_obj, sheet_name, has_headers, header_row_number, max_rows)
    return ([], [])


def _get_csv_preview(path: Path, has_headers: bool, header_row_number: int, max_rows: int):
    sample = []
    dialect = _sniff_csv_dialect(path)
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, dialect)
            for i, row in enumerate(reader, start=1):
                sample.append(list(row))
                if has_headers:
                    if i >= header_row_number + max_rows:
                        break
                else:
                    if i >= max_rows:
                        break
    except Exception:
        return ([], [])

    if not sample:
        return ([], [])

    header_row = []
    data_rows_raw = []
    if has_headers:
        if header_row_number <= len(sample):
            header_row = sample[header_row_number - 1]
            data_rows_raw = sample[header_row_number:]
        else:
            header_row = []
            data_rows_raw = []
    else:
        data_rows_raw = sample

    col_count = 0
    if header_row:
        col_count = len(header_row)
    if data_rows_raw:
        col_count = max(col_count, max(len(r) for r in data_rows_raw))
    if col_count <= 0:
        return ([], [])

    if has_headers and header_row:
        headers = []
        for i in range(col_count):
            val = header_row[i] if i < len(header_row) else ""
            val = str(val).strip() if val is not None else ""
            headers.append(val or _column_label_default(i))
    else:
        headers = [_column_label_default(i) for i in range(col_count)]

    norm_rows = []
    for r in data_rows_raw[:max_rows]:
        rr = []
        for i in range(col_count):
            v = r[i] if i < len(r) else ""
            rr.append("" if v is None else str(v))
        norm_rows.append(rr)

    return (headers, norm_rows)


def _get_xlsx_preview(path: Path, sheet_name: str, has_headers: bool, header_row_number: int, max_rows: int):
    if not HAS_OPENPYXL:
        return ([], [])

    sample = []
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        limit = header_row_number + max_rows if has_headers else max_rows
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            sample.append([] if row is None else list(row))
            if i >= limit:
                break
        wb.close()
    except Exception:
        return ([], [])

    if not sample:
        return ([], [])

    header_row = []
    data_rows_raw = []
    if has_headers:
        if header_row_number <= len(sample):
            header_row = sample[header_row_number - 1]
            data_rows_raw = sample[header_row_number:]
        else:
            header_row = []
            data_rows_raw = []
    else:
        data_rows_raw = sample

    col_count = 0
    if header_row:
        col_count = len(header_row)
    if data_rows_raw:
        col_count = max(col_count, max(len(r) for r in data_rows_raw))
    if col_count <= 0:
        return ([], [])

    if has_headers and header_row:
        headers = []
        for i in range(col_count):
            val = header_row[i] if i < len(header_row) else ""
            val = str(val).strip() if val is not None else ""
            headers.append(val or _column_label_default(i))
    else:
        headers = [_column_label_default(i) for i in range(col_count)]

    norm_rows = []
    for r in data_rows_raw[:max_rows]:
        rr = []
        for i in range(col_count):
            v = r[i] if i < len(r) else ""
            rr.append("" if v is None else str(v))
        norm_rows.append(rr)

    return (headers, norm_rows)


def _safe_get(row, idx):
    try:
        if idx is None or idx < 0:
            return ""
        if idx < len(row):
            v = row[idx]
            return "" if v is None else str(v).strip()
    except Exception:
        pass
    return ""


def _safe_get_code(row, idx):
    if idx is None:
        return ""
    try:
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx]
    except Exception:
        return ""
    if v is None:
        return ""
    try:
        if isinstance(v, bool):
            return str(v).strip()
        if isinstance(v, int):
            return str(v).strip()
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            s = str(v).strip()
            if re.fullmatch(r"\d+\.0+", s):
                return s.split(".", 1)[0]
            return s
    except Exception:
        pass
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def _safe_get_value(row, idx):
    if idx is None or idx < 0:
        return ""
    try:
        if idx >= len(row):
            return ""
        v = row[idx]
    except Exception:
        return ""
    if v is None:
        return ""
    try:
        if isinstance(v, bool):
            return str(v).strip()
        if isinstance(v, int):
            return str(v).strip()
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            return str(v).strip()
    except Exception:
        pass
    return str(v).strip()


def load_products(
    path: str,
    code_col_index: int = 0,
    name_col_index: int = 1,
    category_col_index: int = -1,
    brand_col_index: int = -1,
    unit_price_col_index: int = -1,
    stock_col_index: int = -1,
    has_headers: bool = False,
    header_row_number: int = 1,
    sheet_name: str = None
):
    if header_row_number < 1:
        header_row_number = 1

    path_obj = Path(path)
    ext = path_obj.suffix.lower()
    try:
        if ext == ".csv":
            products = _load_from_csv(
                path_obj,
                code_col_index, name_col_index,
                category_col_index, brand_col_index,
                unit_price_col_index, stock_col_index,
                has_headers, header_row_number
            )
        elif ext in {".xlsx", ".xlsm"}:
            if not HAS_OPENPYXL:
                messagebox.showerror(
                    "Excel",
                    "Para leer archivos .xlsx/.xlsm necesitas instalar openpyxl:\n\npip install openpyxl"
                )
                return []
            products = _load_from_xlsx(
                path_obj,
                code_col_index, name_col_index,
                category_col_index, brand_col_index,
                unit_price_col_index, stock_col_index,
                has_headers, header_row_number,
                sheet_name
            )
        else:
            messagebox.showerror("Excel/CSV", "Extensión no soportada. Usa .csv, .xlsx o .xlsm.")
            return []
    except Exception as e:
        messagebox.showerror("Excel/CSV", f"Ocurrió un error al leer el archivo:\n{e}")
        return []
    return products


def _load_from_csv(
    path: Path,
    code_col_index: int,
    name_col_index: int,
    category_col_index: int,
    brand_col_index: int,
    unit_price_col_index: int,
    stock_col_index: int,
    has_headers: bool,
    header_row_number: int
):
    products = []
    dialect = _sniff_csv_dialect(path)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, dialect)
        for i, row in enumerate(reader, start=1):
            if row is None:
                continue
            if has_headers:
                if i < header_row_number:
                    continue
                if i == header_row_number:
                    continue

            code = _safe_get_code(row, code_col_index)
            name = _safe_get(row, name_col_index)
            category = _safe_get(row, category_col_index)
            brand = _safe_get(row, brand_col_index)
            unit_price = _safe_get_value(row, unit_price_col_index)
            stock = _safe_get_value(row, stock_col_index)

            if not code and not name and not category and not brand and not unit_price and not stock:
                continue

            search_text = normalize_text(f"{code} {name} {category} {brand} {unit_price} {stock}")
            products.append({
                "code": code,
                "name": name,
                "category": category,
                "brand": brand,
                "unit_price": unit_price,
                "stock": stock,
                "search_text": search_text,
                "row_number": int(i),
            })
    return products


def _load_from_xlsx(
    path: Path,
    code_col_index: int,
    name_col_index: int,
    category_col_index: int,
    brand_col_index: int,
    unit_price_col_index: int,
    stock_col_index: int,
    has_headers: bool,
    header_row_number: int,
    sheet_name: str = None
):
    products = []
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row is None:
            row = ()
        if has_headers:
            if i < header_row_number:
                continue
            if i == header_row_number:
                continue

        row_list = list(row)
        code = _safe_get_code(row_list, code_col_index)
        name = _safe_get(row_list, name_col_index)
        category = _safe_get(row_list, category_col_index)
        brand = _safe_get(row_list, brand_col_index)
        unit_price = _safe_get_value(row_list, unit_price_col_index)
        stock = _safe_get_value(row_list, stock_col_index)

        if not code and not name and not category and not brand and not unit_price and not stock:
            continue

        search_text = normalize_text(f"{code} {name} {category} {brand} {unit_price} {stock}")
        products.append({
            "code": code,
            "name": name,
            "category": category,
            "brand": brand,
            "unit_price": unit_price,
            "stock": stock,
            "search_text": search_text,
            "row_number": int(i),
        })
    wb.close()
    return products


def build_code_to_name_index(products):
    code_to_name = {}
    for p in products:
        code = p.get("code", "")
        name = p.get("name", "")
        if code:
            key = sanitize_filename(code)
            code_to_name[key] = name
    return code_to_name


def _parse_number_for_excel(s: str):
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        return s

    txt = str(s).strip()
    if not txt:
        return ""

    txt = txt.replace("$", "").replace(" ", "").replace("\u00a0", "")
    cleaned = [ch for ch in txt if ch.isdigit() or ch in ",.-"]
    txt = "".join(cleaned).strip()
    if not txt:
        return ""

    try:
        if "," in txt and "." in txt:
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "")
                txt = txt.replace(",", ".")
            else:
                txt = txt.replace(",", "")
        else:
            if "," in txt and "." not in txt:
                txt = txt.replace(",", ".")
        val = float(txt)
        return int(val) if val.is_integer() else val
    except Exception:
        return str(s)


def update_product_in_file(
    path: str,
    row_number: int,
    code_col_index: int,
    name_col_index: int,
    category_col_index: int = -1,
    brand_col_index: int = -1,
    unit_price_col_index: int = -1,
    stock_col_index: int = -1,
    sheet_name: str = None,
    has_headers: bool = True,
    header_row_number: int = 1,
    new_code: str = "",
    new_name: str = "",
    new_category: str = "",
    new_brand: str = "",
    new_unit_price: str = "",
    new_stock: str = ""
):
    try:
        # ✅ NUEVO: asegurar guion bajo en lo que se guarda en Excel/CSV
        new_code = to_underscore_text(new_code)
        new_name = to_underscore_text(new_name)

        p = Path(path)
        if not p.exists() or p.is_dir():
            return (False, "La ruta del archivo no existe o es una carpeta.")

        ext = p.suffix.lower()
        if row_number <= 0:
            return (False, "Fila inválida (row_number).")

        if header_row_number < 1:
            header_row_number = 1

        if has_headers and row_number == header_row_number:
            return (False, "La fila a editar corresponde al encabezado.")

        if ext == ".csv":
            return _update_csv_row(
                p,
                row_number=int(row_number),
                code_col_index=int(code_col_index),
                name_col_index=int(name_col_index),
                category_col_index=int(category_col_index),
                brand_col_index=int(brand_col_index),
                unit_price_col_index=int(unit_price_col_index),
                stock_col_index=int(stock_col_index),
                new_code=new_code,
                new_name=new_name,
                new_category=new_category,
                new_brand=new_brand,
                new_unit_price=new_unit_price,
                new_stock=new_stock,
            )

        if ext in {".xlsx", ".xlsm"}:
            if not HAS_OPENPYXL:
                return (False, "No está instalado openpyxl. Instala con: pip install openpyxl")
            return _update_xlsx_row(
                p,
                row_number=int(row_number),
                code_col_index=int(code_col_index),
                name_col_index=int(name_col_index),
                category_col_index=int(category_col_index),
                brand_col_index=int(brand_col_index),
                unit_price_col_index=int(unit_price_col_index),
                stock_col_index=int(stock_col_index),
                sheet_name=sheet_name,
                new_code=new_code,
                new_name=new_name,
                new_category=new_category,
                new_brand=new_brand,
                new_unit_price=new_unit_price,
                new_stock=new_stock,
            )

        return (False, "Extensión no soportada (solo .csv, .xlsx, .xlsm).")

    except PermissionError:
        return (False, "Permiso denegado. ¿El archivo está abierto en Excel?")
    except Exception as e:
        return (False, str(e))

def append_products_to_file(
    path: str,
    products_to_add: List[dict],
    code_col_index: int,
    name_col_index: int,
    sheet_name: str = None,
    has_headers: bool = True,
    header_row_number: int = 1,
):
    """
    Agrega nuevas filas al final del archivo (CSV/XLSX/XLSM) con Id y Nombre producto.
    products_to_add: lista de dicts: {"code": "...", "name": "..."}.
    Retorna (ok: bool, err: str, added_count: int)
    """
    try:
        p = Path(path)
        if not p.exists() or p.is_dir():
            return (False, "La ruta del archivo no existe o es una carpeta.", 0)

        ext = p.suffix.lower()
        if not products_to_add:
            return (True, "", 0)

        if header_row_number < 1:
            header_row_number = 1

        if ext == ".csv":
            return _append_products_to_csv(
                p,
                products_to_add=products_to_add,
                code_col_index=int(code_col_index),
                name_col_index=int(name_col_index),
            )

        if ext in {".xlsx", ".xlsm"}:
            if not HAS_OPENPYXL:
                return (False, "No está instalado openpyxl. Instala con: pip install openpyxl", 0)
            return _append_products_to_xlsx(
                p,
                products_to_add=products_to_add,
                code_col_index=int(code_col_index),
                name_col_index=int(name_col_index),
                sheet_name=sheet_name,
                has_headers=bool(has_headers),
                header_row_number=int(header_row_number),
            )

        return (False, "Extensión no soportada (solo .csv, .xlsx, .xlsm).", 0)

    except PermissionError:
        return (False, "Permiso denegado. ¿El archivo está abierto en Excel?", 0)
    except Exception as e:
        return (False, str(e), 0)


def _append_products_to_csv(
    path: Path,
    products_to_add: List[dict],
    code_col_index: int,
    name_col_index: int,
):
    dialect = _sniff_csv_dialect(path)
    rows = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, dialect)
            for r in reader:
                rows.append(list(r))
    except Exception as e:
        return (False, f"No se pudo leer el CSV: {e}", 0)

    max_idx = max(code_col_index, name_col_index)
    col_count = 0
    if rows:
        col_count = max(len(r) for r in rows)
    col_count = max(col_count, max_idx + 1)

    added = 0
    for item in products_to_add:
        code = str(item.get("code", "") or "").strip()
        name = str(item.get("name", "") or "").strip()
        new_row = [""] * col_count
        if code_col_index >= 0:
            new_row[code_col_index] = code
        if name_col_index >= 0:
            new_row[name_col_index] = name
        rows.append(new_row)
        added += 1

    try:
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, dialect)
            for r in rows:
                writer.writerow(r)
    except Exception as e:
        return (False, f"No se pudo escribir el CSV: {e}", 0)

    return (True, "", added)


def _append_products_to_xlsx(
    path: Path,
    products_to_add: List[dict],
    code_col_index: int,
    name_col_index: int,
    sheet_name: str,
    has_headers: bool,
    header_row_number: int,
):
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=False)
    except Exception as e:
        return (False, f"No se pudo abrir el Excel: {e}", 0)

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if header_row_number < 1:
            header_row_number = 1

        code_col_1b = int(code_col_index) + 1
        name_col_1b = int(name_col_index) + 1

        # Encuentra la última fila de datos real debajo del header
        data_start = header_row_number + 1 if has_headers else 1
        last = int(ws.max_row or 1)

        if last < data_start:
            last = data_start - 1

        def _cell_has_value(r):
            v1 = ws.cell(row=r, column=code_col_1b).value
            v2 = ws.cell(row=r, column=name_col_1b).value
            return (v1 is not None and str(v1).strip() != "") or (v2 is not None and str(v2).strip() != "")

        while last >= data_start:
            if _cell_has_value(last):
                break
            last -= 1

        write_row = last + 1
        if write_row < data_start:
            write_row = data_start

        added = 0
        for item in products_to_add:
            code = str(item.get("code", "") or "").strip()
            name = str(item.get("name", "") or "").strip()
            r = write_row + added
            ws.cell(row=r, column=code_col_1b).value = code
            ws.cell(row=r, column=name_col_1b).value = name
            added += 1

        wb.save(str(path))
        wb.close()
        return (True, "", added)

    except PermissionError:
        try:
            wb.close()
        except Exception:
            pass
        return (False, "Permiso denegado. ¿El archivo está abierto en Excel?", 0)
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return (False, f"Error guardando Excel: {e}", 0)


def _ensure_row_len(row, need_len: int):
    if len(row) < need_len:
        row.extend([""] * (need_len - len(row)))
    return row


def _update_csv_row(
    path: Path,
    row_number: int,
    code_col_index: int,
    name_col_index: int,
    category_col_index: int,
    brand_col_index: int,
    unit_price_col_index: int,
    stock_col_index: int,
    new_code: str,
    new_name: str,
    new_category: str,
    new_brand: str,
    new_unit_price: str,
    new_stock: str
):
    dialect = _sniff_csv_dialect(path)
    rows = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, dialect)
            for r in reader:
                rows.append(list(r))
    except Exception as e:
        return (False, f"No se pudo leer el CSV: {e}")

    if row_number < 1 or row_number > len(rows):
        return (False, "La fila a editar está fuera de rango.")

    idx = row_number - 1
    row = rows[idx]
    max_idx = max(code_col_index, name_col_index, category_col_index, brand_col_index, unit_price_col_index, stock_col_index)
    need_len = (max_idx + 1) if max_idx >= 0 else len(row)
    row = _ensure_row_len(row, need_len)

    if code_col_index >= 0:
        row[code_col_index] = str(new_code or "")
    if name_col_index >= 0:
        row[name_col_index] = str(new_name or "")
    if category_col_index >= 0:
        row[category_col_index] = str(new_category or "")
    if brand_col_index >= 0:
        row[brand_col_index] = str(new_brand or "")
    if unit_price_col_index >= 0:
        row[unit_price_col_index] = str(new_unit_price or "")
    if stock_col_index >= 0:
        row[stock_col_index] = str(new_stock or "")

    rows[idx] = row

    try:
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, dialect)
            for r in rows:
                writer.writerow(r)
    except Exception as e:
        return (False, f"No se pudo escribir el CSV: {e}")

    return (True, "")


def _update_xlsx_row(
    path: Path,
    row_number: int,
    code_col_index: int,
    name_col_index: int,
    category_col_index: int,
    brand_col_index: int,
    unit_price_col_index: int,
    stock_col_index: int,
    sheet_name: str,
    new_code: str,
    new_name: str,
    new_category: str,
    new_brand: str,
    new_unit_price: str,
    new_stock: str
):
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=False)
    except Exception as e:
        return (False, f"No se pudo abrir el Excel: {e}")

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if code_col_index >= 0:
            ws.cell(row=row_number, column=code_col_index + 1).value = str(new_code or "")
        if name_col_index >= 0:
            ws.cell(row=row_number, column=name_col_index + 1).value = str(new_name or "")
        if category_col_index >= 0:
            ws.cell(row=row_number, column=category_col_index + 1).value = str(new_category or "")
        if brand_col_index >= 0:
            ws.cell(row=row_number, column=brand_col_index + 1).value = str(new_brand or "")
        if unit_price_col_index >= 0:
            ws.cell(row=row_number, column=unit_price_col_index + 1).value = _parse_number_for_excel(new_unit_price)
        if stock_col_index >= 0:
            ws.cell(row=row_number, column=stock_col_index + 1).value = _parse_number_for_excel(new_stock)

        wb.save(str(path))
        wb.close()
        return (True, "")
    except PermissionError:
        try:
            wb.close()
        except Exception:
            pass
        return (False, "Permiso denegado. ¿El archivo está abierto en Excel?")
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return (False, f"Error guardando Excel: {e}")


def _norm_hdr(s: str) -> str:
    return normalize_text("" if s is None else str(s)).strip()


def _coerce_code_cell(v):
    if v is None:
        return ""
    try:
        if isinstance(v, bool):
            return str(v).strip()
        if isinstance(v, int):
            return str(v).strip()
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            s = str(v).strip()
            if re.fullmatch(r"\d+\.0+", s):
                return s.split(".", 1)[0]
            return s
    except Exception:
        pass
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def _path_to_excel_hyperlink_target(path_str: str) -> str:
    s = (path_str or "").strip()
    if not s:
        return ""
    if "://" in s:
        return s
    if s.startswith("\\\\"):
        unc = s.lstrip("\\").replace("\\", "/")
        return "file://" + unc
    if len(s) >= 2 and s[1] == ":":
        return "file:///" + s.replace("\\", "/")
    return s


def update_ruta_column_in_xlsx(
    path: str,
    code_col_index: int,
    codekey_to_path: dict,
    sheet_name: str = "catalogo",
    header_row_number: int = 2,
    ruta_header: str = "Ruta",
):
    if not HAS_OPENPYXL:
        return (False, "No está instalado openpyxl.", 0)

    try:
        p = Path(path)
        if not p.exists():
            return (False, "El archivo no existe.", 0)
        ext = p.suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            return (False, "Extensión no soportada para esta acción (solo .xlsx/.xlsm).", 0)

        try:
            wb = load_workbook(filename=str(p), read_only=False, data_only=False)
        except Exception as e:
            return (False, f"No se pudo abrir el Excel: {e}", 0)

        filas = 0
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if header_row_number < 1:
                header_row_number = 1

            max_col = int(ws.max_column or 1)
            ruta_col_1based = None
            for c in range(1, max_col + 1):
                v = ws.cell(row=header_row_number, column=c).value
                if _norm_hdr(v) == _norm_hdr(ruta_header):
                    ruta_col_1based = c
                    break

            if ruta_col_1based is None:
                wb.close()
                return (False, f"No se encontró el encabezado '{ruta_header}' en la fila {header_row_number}.", 0)

            code_col_1based = int(code_col_index) + 1
            start_row = header_row_number + 1
            last_row = int(ws.max_row or start_row)

            left_alignment = None
            if Alignment is not None:
                try:
                    left_alignment = Alignment(horizontal="left")
                except Exception:
                    left_alignment = None

            for r in range(start_row, last_row + 1):
                code_val = ws.cell(row=r, column=code_col_1based).value
                code = _coerce_code_cell(code_val)
                key = sanitize_filename(code) if code else ""
                new_path = ""
                if key:
                    new_path = str(codekey_to_path.get(key, "") or "")

                cell = ws.cell(row=r, column=ruta_col_1based)

                if left_alignment is not None:
                    try:
                        cell.alignment = left_alignment
                    except Exception:
                        pass

                if new_path:
                    cell.value = new_path
                    cell.hyperlink = _path_to_excel_hyperlink_target(new_path)
                    try:
                        cell.style = "Hyperlink"
                    except Exception:
                        pass
                else:
                    cell.value = ""
                    try:
                        cell.hyperlink = None
                    except Exception:
                        try:
                            cell._hyperlink = None
                        except Exception:
                            pass
                    try:
                        cell.style = "Normal"
                    except Exception:
                        pass

                filas += 1

            wb.save(str(p))
            wb.close()
            return (True, "", filas)

        except PermissionError:
            try:
                wb.close()
            except Exception:
                pass
            return (False, "Permiso denegado. ¿El archivo está abierto en Excel?", 0)
        except Exception as e:
            try:
                wb.close()
            except Exception:
                pass
            return (False, f"Error actualizando Excel: {e}", 0)

    except PermissionError:
        return (False, "Permiso denegado. ¿El archivo está abierto en Excel?", 0)
    except Exception as e:
        return (False, str(e), 0)


# =========================
# PIL opcional (para imágenes)
# =========================

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    Image = ImageTk = None


# =========================
# MIXINS UI (antes app_window.py / app_products_ui.py / app_files_ui.py / app_rename_ui.py)
# =========================

class WindowMixin:
    def _create_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        default_font = tkfont.nametofont("TkDefaultFont")
        status_font = default_font.copy()
        status_font.configure(size=max(10, default_font.cget("size") + 2))
        self.status_font = status_font

        style.configure("TFrame", padding=5)
        style.configure("Status.TLabel", padding=(8, 4), font=self.status_font)
        style.configure("Toolbar.TButton", padding=(6, 2))

    def _create_menu(self):
        try:
            self.config(menu=tk.Menu(self))
        except Exception:
            pass

    def _create_widgets(self):
        main_frame = ttk.Frame(self)
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        main_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        paned = ttk.Panedwindow(main_frame, orient=tk.HORIZONTAL)
        paned.grid(row=0, column=0, sticky="nsew")
        self.paned = paned

        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)
        self._create_left_panel(left_frame)

        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)
        self._create_right_panel(right_frame)

        status_frame = ttk.Frame(self, relief=tk.SUNKEN)
        status_frame.grid(row=1, column=0, sticky="ew")
        status_frame.columnconfigure(1, weight=1)

        lbl = ttk.Label(
            status_frame,
            textvariable=self.status_filename_var,
            style="Status.TLabel",
            anchor="w"
        )
        lbl.grid(row=0, column=0, sticky="w")

        entry = ttk.Entry(status_frame, textvariable=self.status_rename_var)
        entry.grid(row=0, column=1, sticky="ew", padx=(8, 6), pady=2)
        entry.bind("<Return>", self.on_statusbar_rename_apply)
        entry.bind("<Escape>", self.on_statusbar_rename_reset)
        try:
            entry.bind("<FocusIn>", lambda e: entry.select_range(0, tk.END), add="+")
        except Exception:
            pass

        btn_apply = ttk.Button(
            status_frame,
            text="Renombrar",
            style="Toolbar.TButton",
            command=self.on_statusbar_rename_apply
        )
        btn_apply.grid(row=0, column=2, sticky="e", padx=(0, 6), pady=2)

        self.status_rename_entry = entry
        self.status_rename_btn = btn_apply

        create_tooltip(
            entry,
            "Escribe el nuevo nombre (sin extensión) y presiona Enter.\nTip: Esc restaura el nombre actual."
        )
        create_tooltip(
            btn_apply,
            "Renombra la imagen actual usando el texto escrito en la barra de estado."
        )

    def _create_left_panel(self, parent: ttk.Frame):
        self.image_frame = ttk.Frame(parent, relief=tk.SUNKEN)
        self.image_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(5, 0))

        self.image_canvas = tk.Canvas(self.image_frame, bg="black", highlightthickness=0)
        self.image_canvas.pack(fill=tk.BOTH, expand=True)

        try:
            if hasattr(self, "_init_image_canvas_interactions"):
                self._init_image_canvas_interactions()
        except Exception:
            pass

        self.image_frame.bind("<Configure>", self.on_image_frame_resize)

    def _create_context_menu(self):
        m = tk.Menu(self, tearoff=False)
        m.add_command(label="Abrir archivo Excel/CSV", command=self.open_excel_inventory_file)
        m.add_command(label="Abrir carpeta del Excel del inventario", command=self.open_excel_inventory_folder)
        m.add_command(label="Abrir carpeta de imágenes", command=self.open_images_folder)
        self._context_menu = m

    def _show_context_menu(self, event):
        try:
            if hasattr(self, "tree"):
                iid = self.tree.identify_row(event.y)
                if iid:
                    try:
                        self.tree.selection_set(iid)
                        self.tree.focus(iid)
                    except Exception:
                        pass
            self._context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self._context_menu.grab_release()
            except Exception:
                pass

    def _create_right_panel(self, parent: ttk.Frame):
        top_controls = ttk.Frame(parent)
        top_controls.pack(side=tk.TOP, fill=tk.X, padx=6, pady=6)
        top_controls.columnconfigure(1, weight=1)

        def _make_blue_title_frame(master, title: str, padding=(8, 4)):
            lf = ttk.LabelFrame(master, padding=padding)
            try:
                fnt = tkfont.nametofont("TkDefaultFont").copy()
                fnt.configure(weight="bold")
            except Exception:
                fnt = ("TkDefaultFont", 10, "bold")
            lbl = tk.Label(lf, text=title, fg="#0B57D0", font=fnt)
            try:
                lf.configure(labelwidget=lbl)
            except Exception:
                lf.configure(text=title)
            return lf

        r = 0

        ttk.Label(top_controls, text="Excel/CSV:").grid(row=r, column=0, sticky="w")
        entry_excel = ttk.Entry(top_controls, textvariable=self.excel_path_var)
        entry_excel.grid(row=r, column=1, sticky="ew", padx=(6, 6))
        entry_excel.bind("<Return>", self.on_load_excel_from_entry)
        btn_load_excel = ttk.Button(top_controls, text="Cargar", command=self.on_load_excel_clicked)
        btn_load_excel.grid(row=r, column=2, sticky="e")
        r += 1

        # ========= NUEVO: fila de encabezados desde interfaz =========
        ttk.Label(top_controls, text="Encabezados (fila):").grid(row=r, column=0, sticky="w", pady=(4, 0))
        spn_hdr = ttk.Spinbox(
            top_controls,
            from_=1,
            to=500,
            width=8,
            textvariable=self.excel_header_row_var,
            command=self.on_excel_header_row_spin_change
        )
        spn_hdr.grid(row=r, column=1, sticky="w", padx=(6, 6), pady=(4, 0))

        btn_apply_hdr = ttk.Button(top_controls, text="Aplicar", command=self.on_excel_header_row_apply)
        btn_apply_hdr.grid(row=r, column=2, sticky="e", pady=(4, 0))

        spn_hdr.bind("<Return>", self.on_excel_header_row_apply)
        spn_hdr.bind("<FocusOut>", self.on_excel_header_row_apply)
        create_tooltip(
            spn_hdr,
            "Indica la fila donde están los encabezados (1 = primera fila).\n"
            "Luego pulsa 'Aplicar' para recargar el Excel usando esa fila."
        )
        create_tooltip(
            btn_apply_hdr,
            "Guarda la fila de encabezados y recarga el Excel/CSV actual (si existe)."
        )
        self._hdr_row_spinbox = spn_hdr
        r += 1
        # ===========================================================

        ttk.Label(top_controls, text="Imágenes:").grid(row=r, column=0, sticky="w", pady=(4, 0))
        entry_img_folder = ttk.Entry(top_controls, textvariable=self.image_folder_path_var)
        entry_img_folder.grid(row=r, column=1, sticky="ew", padx=(6, 6), pady=(4, 0))
        entry_img_folder.bind("<Return>", self.on_load_image_folder_from_entry)
        btn_load_folder = ttk.Button(top_controls, text="Cargar", command=self.on_load_image_folder_clicked)
        btn_load_folder.grid(row=r, column=2, sticky="e", pady=(4, 0))
        r += 1

        ttk.Label(top_controls, text="Buscar:").grid(row=r, column=0, sticky="w", pady=(6, 0))
        entry_search = ttk.Entry(top_controls, textvariable=self.search_var)
        entry_search.grid(row=r, column=1, columnspan=2, sticky="ew", pady=(6, 0))
        entry_search.focus_set()
        self.search_entry = entry_search
        r += 1

        search_opts = _make_blue_title_frame(top_controls, "Opciones de búsqueda", padding=(8, 4))
        search_opts.grid(row=r, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        r += 1

        filters_row = ttk.Frame(search_opts)
        filters_row.pack(side=tk.TOP, fill=tk.X, anchor="w")

        chk_matches = ttk.Checkbutton(
            filters_row,
            text="Solo imagenes coincidentes con la busqueda(Ant/sig)",
            variable=getattr(self, "show_matches_var"),
            command=self.on_toggle_show_matches
        )
        chk_matches.pack(side=tk.LEFT, padx=(0, 10))
        create_tooltip(
            chk_matches,
            "Cuando está activo, Anterior/Siguiente solo recorre imágenes\n"
            "cuyos códigos estén en el Excel y dentro del filtro de búsqueda actual."
        )

        chk_no_code_imgs = ttk.Checkbutton(
            filters_row,
            text="Solo imagnes sin codigo en Excel (Ant/sig)",
            variable=self.show_no_code_images_only_var,
            command=self.on_toggle_no_code_images_filter
        )
        chk_no_code_imgs.pack(side=tk.LEFT, padx=(0, 10))
        create_tooltip(
            chk_no_code_imgs,
            "Cuando está activo, Anterior/Siguiente solo recorre imágenes\n"
            "consideradas 'sin código'."
        )

        nav_row = ttk.Frame(search_opts)
        nav_row.pack(side=tk.TOP, fill=tk.X, anchor="w", pady=(6, 0))
        ttk.Button(nav_row, text="Anterior", style="Toolbar.TButton", command=self.show_prev_image)\
            .pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(nav_row, text="Siguiente", style="Toolbar.TButton", command=self.show_next_image)\
            .pack(side=tk.LEFT, padx=(0, 6))

        rename_opts = _make_blue_title_frame(top_controls, "Opciones de renombrado", padding=(8, 4))
        rename_opts.grid(row=r, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        r += 1

        row1 = ttk.Frame(rename_opts)
        row1.pack(side=tk.TOP, fill=tk.X, anchor="w")
        ttk.Label(row1, text="Renombrado:").pack(side=tk.LEFT)
        ttk.Radiobutton(row1, text="Código", value="code", variable=self.rename_mode_var)\
            .pack(side=tk.LEFT, padx=(6, 2))
        ttk.Radiobutton(row1, text="Nombre", value="name", variable=self.rename_mode_var)\
            .pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(row1, text="Ambos", value="both", variable=self.rename_mode_var)\
            .pack(side=tk.LEFT, padx=2)

        ttk.Label(row1, text="   |   ").pack(side=tk.LEFT)

        scope_chk = ttk.Checkbutton(
            row1,
            text="Ampliar alcance (subcarpetas)",
            variable=self.expand_scope_var
        )
        scope_chk.pack(side=tk.LEFT, padx=(0, 6))
        create_tooltip(
            scope_chk,
            "Si está marcado, 'Renombrar todas' trabajará también en todas las subcarpetas.\n"
            "Si no está marcado, solo renombra imágenes en la carpeta principal."
        )

        row2 = ttk.Frame(rename_opts)
        row2.pack(side=tk.TOP, fill=tk.X, anchor="w", pady=(6, 0))

        self.btn_undo = ttk.Button(row2, text="Deshacer", style="Toolbar.TButton", command=self.on_undo)
        self.btn_undo.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_undo.state(["disabled"])
        create_tooltip(
            self.btn_undo,
            "Deshace el último renombrado realizado.\nImportante: solo revierte el cambio más reciente (una imagen)."
        )

        self.btn_rename_current = ttk.Button(
            row2, text="Renombrar actual (Excel)", style="Toolbar.TButton",
            command=self.on_rename_current_from_product
        )
        self.btn_rename_current.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_rename_all = ttk.Button(
            row2, text="Renombrar todas (Excel)", style="Toolbar.TButton",
            command=self.on_rename_all_from_product
        )
        self.btn_rename_all.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_update_routes = ttk.Button(
            row2, text="Actualizar ruta en Excel", style="Toolbar.TButton",
            command=self.on_update_routes_in_excel
        )
        self.btn_update_routes.pack(side=tk.LEFT, padx=(0, 6))
        create_tooltip(
            self.btn_update_routes,
            "Borra y vuelve a escribir la columna 'Ruta' del Excel\n"
            "con la ruta actual real de cada imagen según el código."
        )


        # ===== NUEVO: fila 3 de opciones (asignar códigos / agregar productos) =====
        row3 = ttk.Frame(rename_opts)
        row3.pack(side=tk.TOP, fill=tk.X, anchor="w", pady=(6, 0))

        self.btn_assign_codes_missing = ttk.Button(
            row3,
            text="Asignar códigos a imágenes sin código",
            style="Toolbar.TButton",
            command=self.on_assign_codes_to_images_without_code
        )
        self.btn_assign_codes_missing.pack(side=tk.LEFT, padx=(0, 6))
        create_tooltip(
            self.btn_assign_codes_missing,
            "Renombra imágenes cuyo nombre NO inicia con un código válido.\n"
            "Los códigos asignados se toman de los 'huecos' entre el menor y mayor Id del Excel."
        )

        self.btn_add_missing_products_excel = ttk.Button(
            row3,
            text="Agregar productos faltantes al Excel",
            style="Toolbar.TButton",
            command=self.on_add_missing_products_to_excel
        )
        self.btn_add_missing_products_excel.pack(side=tk.LEFT, padx=(0, 6))
        create_tooltip(
            self.btn_add_missing_products_excel,
            "Escanea las imágenes con código y, si el código no existe en el Excel,\n"
            "agrega una fila nueva (Id / Nombre producto)."
        )
        # ========================================================================

        info_frame = ttk.Frame(top_controls)
        info_frame.grid(row=r, column=0, columnspan=3, sticky="w", pady=(10, 0))
        r += 1

        self.lbl_files_count = ttk.Label(info_frame, text="Productos (Excel): 0")
        self.lbl_files_count.pack(side=tk.LEFT)
        self.lbl_view_count = ttk.Label(info_frame, text=" | Mostrando: 0")
        self.lbl_view_count.pack(side=tk.LEFT)
        self.lbl_images_count = ttk.Label(info_frame, text=" | Imágenes: 0")
        self.lbl_images_count.pack(side=tk.LEFT)

        list_frame = ttk.Frame(parent)
        list_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=6, pady=(6, 6))

        initial_columns = ("code", "name", "folder")
        self.tree = ttk.Treeview(list_frame, columns=initial_columns, show="headings", selectmode="browse")
        self.tree.heading("code", text="Código")
        self.tree.heading("name", text="Nombre")
        self.tree.heading("folder", text="Carpeta")
        self.tree.column("code", width=90, anchor="w")
        self.tree.column("name", width=280, anchor="w")
        self.tree.column("folder", width=180, anchor="w")

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.bind("<F2>", self.on_tree_edit_request)
        self.tree.bind("<Return>", self.on_tree_edit_request)
        self.tree.bind("<Button-1>", self.on_tree_click_store_column, add="+")
        self.tree.bind("<Escape>", self.on_tree_edit_cancel_request, add="+")

        self._create_context_menu()
        self.tree.bind("<Button-3>", self._show_context_menu, add="+")
        self.tree.bind("<Button-2>", self._show_context_menu, add="+")

        try:
            if hasattr(self, "configure_product_tree_columns"):
                self.configure_product_tree_columns()
        except Exception:
            pass


class FilesUIMixin:
    def _folder_label_for_file(self, file_path: Path) -> str:
        if not file_path or not isinstance(file_path, Path):
            return ""
        root = getattr(self, "image_folder", None)
        if not root or not isinstance(root, Path):
            try:
                return file_path.parent.name
            except Exception:
                return ""
        try:
            if not root.is_dir():
                return file_path.parent.name
        except Exception:
            return file_path.parent.name
        try:
            root_abs = os.path.abspath(str(root))
            parent_abs = os.path.abspath(str(file_path.parent))
            root_cmp = os.path.normcase(os.path.normpath(root_abs))
            parent_cmp = os.path.normcase(os.path.normpath(parent_abs))
            try:
                common = os.path.commonpath([root_cmp, parent_cmp])
            except Exception:
                return file_path.parent.name
            if common != root_cmp:
                return file_path.parent.name
            rel = os.path.relpath(parent_abs, start=root_abs)
            rel = rel.replace("\\", "/").strip()
            if rel in ("", "."):
                return "(principal)"
            while rel.startswith("./"):
                rel = rel[2:]
            rel = rel.strip("/")
            return rel or "(principal)"
        except Exception:
            try:
                return file_path.parent.name
            except Exception:
                return ""

    def _set_statusbar_rename_enabled(self, enabled: bool):
        for attr in ("status_rename_entry", "status_rename_btn"):
            w = getattr(self, attr, None)
            if w is None:
                continue
            try:
                w.state(["!disabled"] if enabled else ["disabled"])
            except Exception:
                try:
                    w.configure(state=("normal" if enabled else "disabled"))
                except Exception:
                    pass

    def _is_image_without_code(self, path: Path) -> bool:
        if not path or not isinstance(path, Path):
            return False
        try:
            ck = extract_code_key_from_stem(path.stem)
        except Exception:
            ck = ""
        if not looks_like_product_code(ck):
            return True
        excel_codes = set((getattr(self, "code_to_name", {}) or {}).keys())
        if excel_codes and ck not in excel_codes:
            return True
        return False

    def _get_nav_indices(self):
        files = list(getattr(self, "image_files", []) or [])
        if not files:
            return []

        use_matches = False
        try:
            v = getattr(self, "show_matches_var", None)
            if v is not None:
                use_matches = bool(v.get())
        except Exception:
            use_matches = False

        if use_matches:
            allowed_codes = set()
            try:
                filtered = list(getattr(self, "filtered_products", []) or [])
                for p in filtered:
                    code = str(p.get("code", "") or "").strip()
                    if code:
                        allowed_codes.add(sanitize_filename(code))
            except Exception:
                allowed_codes = set()

            if not allowed_codes:
                allowed_codes = set((getattr(self, "code_to_name", {}) or {}).keys())

            if not allowed_codes:
                return []

            idxs = []
            for i, p in enumerate(files):
                try:
                    ck = extract_code_key_from_stem(p.stem)
                except Exception:
                    ck = ""
                if not looks_like_product_code(ck):
                    continue
                if ck in allowed_codes:
                    idxs.append(i)
            return idxs

        use_no_code = False
        try:
            v = getattr(self, "show_no_code_images_only_var", None)
            if v is not None:
                use_no_code = bool(v.get())
        except Exception:
            use_no_code = False

        if not use_no_code:
            return list(range(len(files)))

        idxs = []
        for i, p in enumerate(files):
            try:
                if self._is_image_without_code(p):
                    idxs.append(i)
            except Exception:
                pass
        return idxs

    def on_toggle_no_code_images_filter(self):
        files = list(getattr(self, "image_files", []) or [])
        if not files:
            return

        use_no_code = False
        try:
            use_no_code = bool(getattr(self, "show_no_code_images_only_var").get())
        except Exception:
            use_no_code = False

        if not use_no_code:
            try:
                self.show_current_image()
            except Exception:
                pass
            return

        try:
            if hasattr(self, "show_matches_var"):
                self.show_matches_var.set(False)
        except Exception:
            pass

        idxs = self._get_nav_indices()
        if not idxs:
            messagebox.showinfo("Imágenes", "No hay imágenes sin código para navegar con el filtro activado.")
            try:
                getattr(self, "show_no_code_images_only_var").set(False)
            except Exception:
                pass
            return

        cur = getattr(self, "current_image_index", -1)
        if cur is None or cur < 0:
            cur = 0

        pos = bisect_left(idxs, cur)
        if pos >= len(idxs):
            pos = len(idxs) - 1
        if idxs[pos] != cur:
            self.current_image_index = idxs[pos]
        self.show_current_image()

    def on_load_image_folder_clicked(self):
        raw_path = self.image_folder_path_var.get().strip().strip('"')
        initialdir = None
        if raw_path:
            p = Path(raw_path)
            try:
                initialdir = str(p) if p.is_dir() else str(p.parent)
            except Exception:
                initialdir = None

        selected = filedialog.askdirectory(
            title="Seleccionar carpeta principal de imágenes",
            initialdir=initialdir if initialdir else None,
        )
        if not selected:
            return

        folder = Path(selected)
        self.image_folder_path_var.set(str(folder))
        self.set_image_folder(folder)
        self.save_config()

    def on_load_image_folder_from_entry(self, event=None):
        raw_path = self.image_folder_path_var.get().strip().strip('"')
        if not raw_path:
            messagebox.showerror("Imágenes", "Por favor pega la ruta de la carpeta de imágenes.")
            return
        folder = Path(raw_path)
        if not folder.is_dir():
            messagebox.showerror("Imágenes", f"La carpeta no existe o no es válida:\n{raw_path}")
            return
        self.set_image_folder(folder)
        self.save_config()

    def set_image_folder(self, folder_path: Path):
        if not folder_path or not folder_path.is_dir():
            return
        self.image_folder = folder_path
        self.image_folder_path_var.set(str(folder_path))

        self.reload_image_files(keep_current=False)

        if getattr(self, "_config_last_image_path", None) is not None:
            try:
                if self._config_last_image_path in self.image_files:
                    self.current_image_index = self.image_files.index(self._config_last_image_path)
            except Exception:
                pass
            self._config_last_image_path = None

        if self.image_files and self.current_image_index < 0:
            self.current_image_index = 0

        self.show_current_image()
        try:
            if hasattr(self, "refresh_product_list"):
                self.refresh_product_list()
        except Exception:
            pass

        self.save_config()

    def reload_image_files(self, keep_current: bool = True):
        current_path = None
        if keep_current and self.image_files and 0 <= self.current_image_index < len(self.image_files):
            current_path = self.image_files[self.current_image_index]

        files = []
        if self.image_folder and self.image_folder.is_dir():
            files = gather_all_files(self.image_folder)
            files = [p for p in files if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS]

        self.all_image_files = files
        self.image_files = list(files)

        if hasattr(self, "lbl_images_count"):
            try:
                self.lbl_images_count.config(text=f" | Imágenes: {len(self.image_files)}")
            except Exception:
                pass

        code_to_folder = {}
        code_to_image_path = {}
        for p in self.all_image_files:
            try:
                ck = extract_code_key_from_stem(p.stem)
            except Exception:
                ck = ""
            if not looks_like_product_code(ck):
                continue
            if ck not in code_to_image_path:
                code_to_image_path[ck] = p
                code_to_folder[ck] = self._folder_label_for_file(p)
            else:
                if not code_to_folder.get(ck):
                    code_to_folder[ck] = self._folder_label_for_file(p)

        self.code_to_folder = code_to_folder
        self.code_to_image_path = code_to_image_path

        if not self.image_files:
            self.current_image_index = -1
            return

        if current_path and current_path in self.image_files:
            self.current_image_index = self.image_files.index(current_path)
        else:
            self.current_image_index = min(max(self.current_image_index, 0), len(self.image_files) - 1)

    def _get_view_mode(self) -> str:
        return str(getattr(self, "_view_mode", "single"))

    def _set_view_mode(self, mode: str):
        if mode not in ("single",):
            mode = "single"
        self._view_mode = mode

    def update_status_bar(self):
        if not self.image_files or not (0 <= self.current_image_index < len(self.image_files)):
            self.status_filename_var.set("(sin archivo)")
            try:
                self.status_rename_var.set("")
            except Exception:
                pass
            self._set_statusbar_rename_enabled(False)
            return

        path = self.image_files[self.current_image_index]
        self.status_filename_var.set(f"{path.name}")
        try:
            self.status_rename_var.set(path.stem)
        except Exception:
            pass
        self._set_statusbar_rename_enabled(True)

    # ---- zoom/pan ----
    def _ensure_zoom_pan_state(self):
        if not hasattr(self, "_zoom_factor"):
            self._zoom_factor = 1.0
        if not hasattr(self, "_pan_x"):
            self._pan_x = 0.0
        if not hasattr(self, "_pan_y"):
            self._pan_y = 0.0
        if not hasattr(self, "_fit_scale"):
            self._fit_scale = 1.0
        if not hasattr(self, "_canvas_img_id"):
            self._canvas_img_id = None
        if not hasattr(self, "_canvas_interactions_ready"):
            self._canvas_interactions_ready = False
        if not hasattr(self, "_last_canvas_w"):
            self._last_canvas_w = 0
        if not hasattr(self, "_last_canvas_h"):
            self._last_canvas_h = 0
        if not hasattr(self, "_last_display_w"):
            self._last_display_w = 0
        if not hasattr(self, "_last_display_h"):
            self._last_display_h = 0
        if not hasattr(self, "_last_base_x"):
            self._last_base_x = 0
        if not hasattr(self, "_last_base_y"):
            self._last_base_y = 0
        if not hasattr(self, "_dragging"):
            self._dragging = False
        if not hasattr(self, "_drag_start_x"):
            self._drag_start_x = 0
        if not hasattr(self, "_drag_start_y"):
            self._drag_start_y = 0
        if not hasattr(self, "_drag_start_pan_x"):
            self._drag_start_pan_x = 0.0
        if not hasattr(self, "_drag_start_pan_y"):
            self._drag_start_pan_y = 0.0
        if not hasattr(self, "_zoom_step"):
            self._zoom_step = 1.12
        if not hasattr(self, "_zoom_min"):
            self._zoom_min = 0.20
        if not hasattr(self, "_zoom_max"):
            self._zoom_max = 8.0
        if not hasattr(self, "_max_render_side"):
            self._max_render_side = 8000

    def _reset_zoom_and_pan(self):
        self._ensure_zoom_pan_state()
        self._zoom_factor = 1.0
        self._pan_x = 0.0
        self._pan_y = 0.0

    def _clamp_pan(self, canvas_w, canvas_h, display_w, display_h, base_x, base_y):
        self._ensure_zoom_pan_state()
        pan_x = float(getattr(self, "_pan_x", 0.0))
        pan_y = float(getattr(self, "_pan_y", 0.0))

        if display_w <= canvas_w:
            pan_x = 0.0
        else:
            max_pan_x = -base_x
            min_pan_x = (canvas_w - display_w) - base_x
            pan_x = min(max(pan_x, min_pan_x), max_pan_x)

        if display_h <= canvas_h:
            pan_y = 0.0
        else:
            max_pan_y = -base_y
            min_pan_y = (canvas_h - display_h) - base_y
            pan_y = min(max(pan_y, min_pan_y), max_pan_y)

        self._pan_x = pan_x
        self._pan_y = pan_y

    def _init_image_canvas_interactions(self):
        self._ensure_zoom_pan_state()
        c = getattr(self, "image_canvas", None)
        if c is None:
            return
        if bool(getattr(self, "_canvas_interactions_ready", False)):
            return
        self._canvas_interactions_ready = True

        try:
            c.bind("<Enter>", lambda e: c.focus_set(), add="+")
        except Exception:
            pass
        try:
            c.bind("<MouseWheel>", self._on_canvas_mousewheel, add="+")
        except Exception:
            pass
        try:
            c.bind("<Button-4>", self._on_canvas_mousewheel_linux, add="+")
            c.bind("<Button-5>", self._on_canvas_mousewheel_linux, add="+")
        except Exception:
            pass
        try:
            c.bind("<ButtonPress-1>", self._on_canvas_drag_start, add="+")
            c.bind("<B1-Motion>", self._on_canvas_drag_move, add="+")
            c.bind("<ButtonRelease-1>", self._on_canvas_drag_end, add="+")
        except Exception:
            pass
        try:
            c.bind("<Double-Button-1>", self._on_canvas_double_click_reset, add="+")
        except Exception:
            pass

    def _on_canvas_double_click_reset(self, event=None):
        if self.current_image is None:
            return
        self._reset_zoom_and_pan()
        self.update_image_display()

    def _on_canvas_mousewheel(self, event):
        if self.current_image is None:
            return
        self._ensure_zoom_pan_state()
        try:
            delta = int(event.delta)
        except Exception:
            delta = 0
        if delta == 0:
            return
        step = float(getattr(self, "_zoom_step", 1.12))
        factor = step if delta > 0 else (1.0 / step)
        self._zoom_at(event.x, event.y, factor)

    def _on_canvas_mousewheel_linux(self, event):
        if self.current_image is None:
            return
        self._ensure_zoom_pan_state()
        step = float(getattr(self, "_zoom_step", 1.12))
        try:
            factor = step if int(getattr(event, "num", 0)) == 4 else (1.0 / step)
        except Exception:
            factor = step
        self._zoom_at(event.x, event.y, factor)

    def _zoom_at(self, cx, cy, factor: float):
        if self.current_image is None:
            return
        self._ensure_zoom_pan_state()
        canvas = getattr(self, "image_canvas", None)
        if canvas is None:
            return

        canvas_w = int(canvas.winfo_width() or 0)
        canvas_h = int(canvas.winfo_height() or 0)
        if canvas_w < 10 or canvas_h < 10:
            return

        img_w, img_h = self.current_image.size
        if img_w <= 0 or img_h <= 0:
            return

        fit_scale = min(canvas_w / img_w, canvas_h / img_h)
        if fit_scale <= 0:
            return

        old_zoom = float(getattr(self, "_zoom_factor", 1.0))
        old_zoom = max(float(getattr(self, "_zoom_min", 0.2)), min(float(getattr(self, "_zoom_max", 8.0)), old_zoom))
        old_scale = fit_scale * old_zoom
        if old_scale <= 0:
            return

        old_display_w = max(1, int(round(img_w * old_scale)))
        old_display_h = max(1, int(round(img_h * old_scale)))
        old_base_x = (canvas_w - old_display_w) // 2
        old_base_y = (canvas_h - old_display_h) // 2
        old_left = old_base_x + float(getattr(self, "_pan_x", 0.0))
        old_top = old_base_y + float(getattr(self, "_pan_y", 0.0))

        try:
            ix = (cx - old_left) / old_scale
            iy = (cy - old_top) / old_scale
        except Exception:
            ix, iy = (img_w / 2.0), (img_h / 2.0)

        zmin = float(getattr(self, "_zoom_min", 0.2))
        zmax = float(getattr(self, "_zoom_max", 8.0))
        new_zoom = old_zoom * float(factor)
        new_zoom = max(zmin, min(zmax, new_zoom))
        if abs(new_zoom - old_zoom) < 1e-9:
            return

        self._zoom_factor = new_zoom
        new_scale = fit_scale * new_zoom
        new_display_w = max(1, int(round(img_w * new_scale)))
        new_display_h = max(1, int(round(img_h * new_scale)))
        new_base_x = (canvas_w - new_display_w) // 2
        new_base_y = (canvas_h - new_display_h) // 2

        new_left = cx - (ix * new_scale)
        new_top = cy - (iy * new_scale)
        self._pan_x = float(new_left - new_base_x)
        self._pan_y = float(new_top - new_base_y)

        self._clamp_pan(canvas_w, canvas_h, new_display_w, new_display_h, new_base_x, new_base_y)
        self.update_image_display()

    def _on_canvas_drag_start(self, event):
        if self.current_image is None:
            return
        self._ensure_zoom_pan_state()
        self._dragging = True
        self._drag_start_x = int(getattr(event, "x", 0))
        self._drag_start_y = int(getattr(event, "y", 0))
        self._drag_start_pan_x = float(getattr(self, "_pan_x", 0.0))
        self._drag_start_pan_y = float(getattr(self, "_pan_y", 0.0))

    def _on_canvas_drag_move(self, event):
        if not getattr(self, "_dragging", False):
            return
        if self.current_image is None:
            return

        self._ensure_zoom_pan_state()
        dx = float(int(getattr(event, "x", 0)) - int(getattr(self, "_drag_start_x", 0)))
        dy = float(int(getattr(event, "y", 0)) - int(getattr(self, "_drag_start_y", 0)))
        self._pan_x = float(getattr(self, "_drag_start_pan_x", 0.0)) + dx
        self._pan_y = float(getattr(self, "_drag_start_pan_y", 0.0)) + dy

        canvas_w = int(getattr(self, "_last_canvas_w", 0) or 0)
        canvas_h = int(getattr(self, "_last_canvas_h", 0) or 0)
        display_w = int(getattr(self, "_last_display_w", 0) or 0)
        display_h = int(getattr(self, "_last_display_h", 0) or 0)
        base_x = int(getattr(self, "_last_base_x", 0) or 0)
        base_y = int(getattr(self, "_last_base_y", 0) or 0)

        if canvas_w < 10 or canvas_h < 10 or display_w <= 0 or display_h <= 0:
            self.update_image_display()
            return

        self._clamp_pan(canvas_w, canvas_h, display_w, display_h, base_x, base_y)
        self._update_canvas_image_position()

    def _on_canvas_drag_end(self, event=None):
        self._dragging = False

    def _update_canvas_image_position(self):
        c = getattr(self, "image_canvas", None)
        iid = getattr(self, "_canvas_img_id", None)
        if c is None or iid is None:
            return

        base_x = int(getattr(self, "_last_base_x", 0) or 0)
        base_y = int(getattr(self, "_last_base_y", 0) or 0)
        pan_x = float(getattr(self, "_pan_x", 0.0))
        pan_y = float(getattr(self, "_pan_y", 0.0))
        x = base_x + pan_x
        y = base_y + pan_y
        try:
            c.coords(iid, x, y)
        except Exception:
            try:
                self.update_image_display()
            except Exception:
                pass

    # ---- visualización ----
    def show_current_image(self):
        self._set_view_mode("single")
        if not self.image_files or not (0 <= self.current_image_index < len(self.image_files)):
            self.current_image = None
            self.current_photo = None
            self._ensure_zoom_pan_state()
            self._canvas_img_id = None
            self.image_canvas.delete("all")
            self.update_status_bar()
            return

        path = self.image_files[self.current_image_index]
        ext = path.suffix.lower()
        if ext not in IMAGE_EXTENSIONS:
            self.current_image = None
            self.current_photo = None
            self._ensure_zoom_pan_state()
            self._canvas_img_id = None
            self.image_canvas.delete("all")
            self.update_status_bar()
            return

        if not HAS_PIL:
            messagebox.showerror(
                "PIL/Pillow",
                "Para visualizar imágenes necesitas instalar Pillow:\n\npip install pillow"
            )
            self.current_image = None
            self.current_photo = None
            self._ensure_zoom_pan_state()
            self._canvas_img_id = None
            self.image_canvas.delete("all")
            self.update_status_bar()
            return

        try:
            with Image.open(path) as img:
                self.current_image = img.copy()
        except Exception as e:
            messagebox.showerror("Imágenes", f"No se pudo abrir la imagen:\n{path.name}\n\n{e}")
            self.current_image = None
            self.current_photo = None
            self._ensure_zoom_pan_state()
            self._canvas_img_id = None
            self.image_canvas.delete("all")
            self.update_status_bar()
            return

        self._reset_zoom_and_pan()
        self.update_image_display()
        self.update_status_bar()

    def update_image_display(self):
        if not HAS_PIL or self.current_image is None:
            return
        self._ensure_zoom_pan_state()

        canvas = getattr(self, "image_canvas", None)
        if canvas is None:
            return
        canvas_width = int(canvas.winfo_width() or 0)
        canvas_height = int(canvas.winfo_height() or 0)
        if canvas_width < 10 or canvas_height < 10:
            return

        img_w, img_h = self.current_image.size
        if img_w <= 0 or img_h <= 0:
            return

        fit_scale = min(canvas_width / img_w, canvas_height / img_h)
        if fit_scale <= 0:
            fit_scale = 1.0
        self._fit_scale = fit_scale

        zoom = float(getattr(self, "_zoom_factor", 1.0))
        zmin = float(getattr(self, "_zoom_min", 0.2))
        zmax = float(getattr(self, "_zoom_max", 8.0))
        zoom = max(zmin, min(zmax, zoom))
        self._zoom_factor = zoom

        scale = fit_scale * zoom
        if scale <= 0:
            scale = fit_scale

        display_w = max(1, int(round(img_w * scale)))
        display_h = max(1, int(round(img_h * scale)))

        max_side = int(getattr(self, "_max_render_side", 8000))
        if max(display_w, display_h) > max_side:
            shrink = max_side / float(max(display_w, display_h))
            display_w = max(1, int(round(display_w * shrink)))
            display_h = max(1, int(round(display_h * shrink)))

        base_x = (canvas_width - display_w) // 2
        base_y = (canvas_height - display_h) // 2

        self._clamp_pan(canvas_width, canvas_height, display_w, display_h, base_x, base_y)

        offset_x = base_x + float(getattr(self, "_pan_x", 0.0))
        offset_y = base_y + float(getattr(self, "_pan_y", 0.0))

        try:
            img = self.current_image.resize((display_w, display_h), Image.LANCZOS)
        except Exception:
            img = self.current_image.resize((display_w, display_h))

        self.current_photo = ImageTk.PhotoImage(img)
        try:
            self.image_canvas.delete("all")
        except Exception:
            pass

        try:
            self._canvas_img_id = self.image_canvas.create_image(offset_x, offset_y, image=self.current_photo, anchor="nw")
        except Exception:
            self._canvas_img_id = None

        self._last_canvas_w = canvas_width
        self._last_canvas_h = canvas_height
        self._last_display_w = display_w
        self._last_display_h = display_h
        self._last_base_x = base_x
        self._last_base_y = base_y

    def show_prev_image(self):
        if not self.image_files:
            return
        idxs = self._get_nav_indices()
        if not idxs:
            return
        cur = self.current_image_index
        if cur is None or cur < 0:
            cur = 0
        pos = bisect_left(idxs, cur)
        new_pos = pos - 1
        if new_pos < 0:
            return
        self.current_image_index = idxs[new_pos]
        self.show_current_image()

    def show_next_image(self):
        if not self.image_files:
            return
        idxs = self._get_nav_indices()
        if not idxs:
            return
        cur = self.current_image_index
        if cur is None or cur < 0:
            cur = 0
        pos = bisect_left(idxs, cur)
        new_pos = pos + 1 if (pos < len(idxs) and idxs[pos] == cur) else pos
        if new_pos >= len(idxs):
            return
        self.current_image_index = idxs[new_pos]
        self.show_current_image()

    def on_image_frame_resize(self, event):
        self.update_image_display()


class ProductsUIMixin:
    # =========================
    # NUEVO: helpers para fila encabezados
    # =========================
    def _get_excel_header_row_from_ui(self) -> int:
        n = None
        try:
            v = getattr(self, "excel_header_row_var", None)
            if v is not None:
                n = int(v.get())
        except Exception:
            n = None
        if n is None:
            try:
                n = int(getattr(self, "excel_header_row_number", 2))
            except Exception:
                n = 2
        if n < 1:
            n = 1
        return int(n)

    def _sync_excel_header_row(self):
        n = self._get_excel_header_row_from_ui()
        self.excel_header_row_number = int(n)
        try:
            if hasattr(self, "excel_header_row_var"):
                self.excel_header_row_var.set(int(n))
        except Exception:
            pass

    def on_excel_header_row_spin_change(self):
        # cambio por flechas del spinbox: solo sincroniza y guarda config (sin popups)
        try:
            self._sync_excel_header_row()
        except Exception:
            pass
        try:
            self.save_config()
        except Exception:
            pass

    def on_excel_header_row_apply(self, event=None):
        # Aplica la fila y recarga el Excel actual (si existe)
        try:
            self._sync_excel_header_row()
        except Exception:
            pass

        try:
            self.save_config()
        except Exception:
            pass

        raw = (self.excel_path_var.get() or "").strip().strip('"')
        if not raw:
            return

        p = Path(raw)
        if not p.exists() or p.is_dir():
            return

        # Recargar sin el mensaje largo de "Reglas usadas" (evita ruido)
        # Si falla, load_excel mostrará el error correspondiente.
        self.load_excel(str(p), interactive=False)

    # =========================
    # columnas TreeView
    # =========================
    def configure_product_tree_columns(self):
        if not hasattr(self, "tree"):
            return
        cols = ["code", "name", "folder"]
        self.tree["columns"] = tuple(cols)
        self.tree["show"] = "headings"
        headers = {"code": "Código", "name": "Nombre", "folder": "Carpeta"}
        default_widths = {"code": 90, "name": 280, "folder": 180}
        for col in cols:
            self.tree.heading(col, text=headers.get(col, col))
            self.tree.column(col, width=default_widths.get(col, 120), anchor="w")

        saved = getattr(self, "_config_tree_columns", None)
        if isinstance(saved, dict):
            for col in cols:
                if col in saved:
                    try:
                        self.tree.column(col, width=int(saved[col]))
                    except Exception:
                        pass

    def _update_products_labels(self):
        total = len(getattr(self, "products", []) or [])
        showing = len(getattr(self, "filtered_products", []) or [])
        if hasattr(self, "lbl_files_count"):
            try:
                self.lbl_files_count.config(text=f"Productos (Excel): {total}")
            except Exception:
                pass
        if hasattr(self, "lbl_view_count"):
            try:
                self.lbl_view_count.config(text=f" | Mostrando: {showing}")
            except Exception:
                pass

    def _open_path(self, path: Path):
        try:
            if hasattr(os, "startfile"):
                os.startfile(str(path))
                return
        except Exception:
            pass
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception:
            pass

    def open_excel_inventory_file(self):
        raw = (self.excel_path_var.get() or "").strip().strip('"')
        if not raw:
            messagebox.showwarning("Excel", "No hay ruta de Excel/CSV cargada.")
            return
        p = Path(raw)
        if not p.exists():
            messagebox.showerror("Excel", f"No se encontró el archivo:\n{p}")
            return
        if p.is_dir():
            messagebox.showwarning("Excel", "La ruta actual es una carpeta. Carga un archivo Excel/CSV.")
            return
        self._open_path(p)

    def open_excel_inventory_folder(self):
        raw = (self.excel_path_var.get() or "").strip().strip('"')
        if not raw:
            messagebox.showwarning("Excel", "No hay ruta de Excel/CSV cargada.")
            return
        p = Path(raw)
        folder = p if p.is_dir() else p.parent
        if not folder.exists():
            messagebox.showerror("Excel", f"No se encontró la carpeta:\n{folder}")
            return
        self._open_path(folder)

    def open_images_folder(self):
        root = getattr(self, "image_folder", None)
        if root and hasattr(root, "is_dir") and root.is_dir():
            self._open_path(root)
            return
        raw = (self.image_folder_path_var.get() or "").strip().strip('"')
        if not raw:
            messagebox.showwarning("Imágenes", "No hay carpeta de imágenes cargada.")
            return
        p = Path(raw)
        if not p.is_dir():
            messagebox.showerror("Imágenes", f"La carpeta no existe:\n{p}")
            return
        self._open_path(p)

    def on_load_excel_clicked(self):
        raw_path = self.excel_path_var.get().strip().strip('"')
        initialdir = None
        if raw_path:
            p = Path(raw_path)
            try:
                if p.is_file():
                    initialdir = str(p.parent)
                elif p.is_dir():
                    initialdir = str(p)
                else:
                    initialdir = str(p.parent)
            except Exception:
                initialdir = None

        filetypes = [
            ("Archivos de Excel y CSV", "*.xlsx *.xlsm *.csv"),
            ("Archivos de Excel", "*.xlsx *.xlsm"),
            ("Archivos CSV", "*.csv"),
            ("Todos los archivos", "*.*"),
        ]
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo de productos (Excel/CSV)",
            filetypes=filetypes,
            initialdir=initialdir if initialdir else None,
        )
        if not filename:
            return

        path_obj = Path(filename)
        self.excel_path_var.set(str(path_obj))

        # sincroniza fila encabezados desde UI antes de cargar
        try:
            self._sync_excel_header_row()
        except Exception:
            pass

        self.load_excel(str(path_obj), interactive=True)
        self.save_config()

    def on_load_excel_from_entry(self, event=None):
        path = self.excel_path_var.get().strip().strip('"')
        if not path:
            messagebox.showwarning("Excel", "Por favor pega la ruta completa del archivo de Excel/CSV.")
            return
        path_obj = Path(path)
        if not path_obj.exists():
            messagebox.showerror("Excel", f"La ruta no existe:\n{path}")
            return

        try:
            self._sync_excel_header_row()
        except Exception:
            pass

        self.load_excel(str(path_obj), interactive=True)
        self.save_config()

    def _norm_header(self, s: str) -> str:
        t = normalize_text("" if s is None else str(s))
        t = " ".join(t.split())
        return t.strip()

    def _find_required_columns(self, path: str):
        expected_code = self._norm_header("Id")
        expected_name = self._norm_header("Nombre producto")

        ext = Path(path).suffix.lower()
        sheet_name = getattr(self, "excel_sheet_name", "catalogo") if ext in {".xlsx", ".xlsm"} else None

        if ext in {".xlsx", ".xlsm"}:
            names = get_excel_sheet_names(path)
            if names and sheet_name not in names:
                messagebox.showerror(
                    "Excel",
                    "No se encontró la hoja requerida.\n\n"
                    f"Hoja esperada: {sheet_name}\n"
                    f"Hojas disponibles: {', '.join(names)}"
                )
                return (None, None)

        hdr_row = int(getattr(self, "excel_header_row_number", 1))
        if hdr_row < 1:
            hdr_row = 1

        headers, _rows = get_table_preview(
            path=path,
            sheet_name=sheet_name,
            has_headers=True,
            header_row_number=hdr_row,
            max_rows=5
        )
        if not headers:
            messagebox.showerror(
                "Excel/CSV",
                "No pude leer los encabezados del archivo.\n"
                f"Verifica que tenga encabezados en la fila {hdr_row}."
            )
            return (None, None)

        norm_headers = [self._norm_header(h) for h in headers]
        code_idx = None
        name_idx = None

        for i, h in enumerate(norm_headers):
            if h == expected_code and code_idx is None:
                code_idx = i
            if h == expected_name and name_idx is None:
                name_idx = i

        if code_idx is None or name_idx is None:
            shown = ", ".join([str(h) for h in headers])
            messagebox.showerror(
                "Excel/CSV",
                "No se encontraron las columnas requeridas por encabezado.\n\n"
                "Se esperan EXACTAMENTE estos encabezados (sin importar mayúsculas):\n"
                " - Id\n"
                " - Nombre producto\n\n"
                f"Encabezados detectados en la fila {hdr_row}:\n"
                f"{shown}"
            )
            return (None, None)

        if code_idx == name_idx:
            messagebox.showerror("Excel/CSV", "Error: la columna 'Id' y 'Nombre producto' no pueden ser la misma.")
            return (None, None)

        return (int(code_idx), int(name_idx))

    def load_excel(self, path: str, interactive: bool = True):
        # Hoja fija como venías usando
        self.excel_sheet_name = "catalogo"
        self.excel_has_headers = True

        # ========= NUEVO: usar fila de encabezados elegida en UI =========
        try:
            self._sync_excel_header_row()
        except Exception:
            pass
        # ================================================================

        code_idx, name_idx = self._find_required_columns(path)
        if code_idx is None or name_idx is None:
            return

        self.excel_code_col_index = code_idx
        self.excel_name_col_index = name_idx

        products = load_products(
            path,
            code_col_index=int(self.excel_code_col_index),
            name_col_index=int(self.excel_name_col_index),
            has_headers=True,
            header_row_number=int(getattr(self, "excel_header_row_number", 1)),
            sheet_name=self.excel_sheet_name,
        )
        if not products:
            return

        self.products = products
        self.code_to_name = build_code_to_name_index(products)
        self.configure_product_tree_columns()
        self.apply_search_filter()

        if interactive:
            messagebox.showinfo(
                "Excel/CSV",
                f"Se cargaron {len(self.products)} productos desde:\n{Path(path).name}\n\n"
                "Reglas usadas:\n"
                "- Hoja: catalogo\n"
                f"- Encabezados: fila {int(getattr(self, 'excel_header_row_number', 1))}\n"
                "- Columnas: Id / Nombre producto"
            )

    def on_update_routes_in_excel(self):
        raw = (self.excel_path_var.get() or "").strip().strip('"')
        if not raw:
            messagebox.showwarning("Ruta", "Primero carga el Excel.")
            return
        p = Path(raw)
        if not p.exists() or p.is_dir():
            messagebox.showerror("Ruta", "La ruta del Excel no es válida.")
            return
        ext = p.suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            messagebox.showerror("Ruta", "Esta opción solo funciona con archivos .xlsx/.xlsm.")
            return

        root = getattr(self, "image_folder", None)
        if not root or not hasattr(root, "is_dir") or not root.is_dir():
            messagebox.showwarning("Ruta", "Primero carga la carpeta principal de imágenes.")
            return

        try:
            if hasattr(self, "reload_image_files"):
                self.reload_image_files(keep_current=True)
        except Exception:
            pass

        code_to_image_path = getattr(self, "code_to_image_path", {}) or {}
        mapping = {}
        for code_key, img_path in code_to_image_path.items():
            try:
                mapping[str(code_key)] = str(Path(img_path).resolve())
            except Exception:
                mapping[str(code_key)] = str(img_path)

        ok, err, filas = update_ruta_column_in_xlsx(
            path=str(p),
            code_col_index=int(getattr(self, "excel_code_col_index", 0)),
            codekey_to_path=mapping,
            sheet_name=getattr(self, "excel_sheet_name", "catalogo"),
            header_row_number=int(getattr(self, "excel_header_row_number", 1)),
            ruta_header="Ruta",
        )
        if not ok:
            messagebox.showerror(
                "Actualizar ruta",
                f"No se pudo actualizar la columna 'Ruta':\n\n{err}\n\n"
                f"Asegúrate de que exista un encabezado llamado 'Ruta' en la fila {int(getattr(self, 'excel_header_row_number', 1))}."
            )
            return

        messagebox.showinfo(
            "Actualizar ruta",
            "Listo.\n\n"
            f"Filas procesadas: {filas}\n\n"
            "Si el archivo estaba abierto en Excel, ciérralo e inténtalo de nuevo."
        )

    def apply_search_filter(self):
        query_norm = normalize_text(self.search_var.get())
        terms = [t for t in query_norm.split() if t]

        if not self.products:
            filtered = []
        elif not terms:
            filtered = list(self.products)
        else:
            filtered = []
            for p in self.products:
                text = p.get("search_text", "")
                if all(term in text for term in terms):
                    filtered.append(p)

        self.filtered_products = filtered
        self.refresh_product_list()
        self._update_products_labels()

    def refresh_product_list(self):
        if not hasattr(self, "tree"):
            return

        self._tree_iid_to_product = {}
        for item in self.tree.get_children():
            self.tree.delete(item)

        code_to_folder = getattr(self, "code_to_folder", {}) or {}
        code_to_image_path = getattr(self, "code_to_image_path", {}) or {}
        cols = list(self.tree["columns"]) if self.tree["columns"] else ["code", "name", "folder"]

        for idx, p in enumerate(self.filtered_products):
            code = str(p.get("code", "") or "").strip()
            name = str(p.get("name", "") or "").strip()

            folder_label = ""
            if code:
                try:
                    key = sanitize_filename(code)
                    img_path = code_to_image_path.get(key)
                    if img_path is not None and hasattr(self, "_folder_label_for_file"):
                        folder_label = self._folder_label_for_file(img_path)
                    else:
                        folder_label = code_to_folder.get(key, "") or ""
                except Exception:
                    folder_label = ""

            values_map = {"code": code, "name": name, "folder": folder_label}
            row_values = tuple(values_map.get(c, "") for c in cols)

            row_number = p.get("row_number", None)
            iid = f"r{row_number}" if row_number is not None else f"i{idx}"
            self.tree.insert("", "end", iid=str(iid), values=row_values)
            self._tree_iid_to_product[str(iid)] = p

    def _get_focused_product_data(self):
        if not hasattr(self, "tree"):
            return None

        item_id = ""
        try:
            item_id = self.tree.focus() or ""
        except Exception:
            item_id = ""

        if not item_id:
            try:
                sel = self.tree.selection()
                if sel:
                    item_id = sel[0]
            except Exception:
                item_id = ""

        if not item_id:
            return None

        mp = getattr(self, "_tree_iid_to_product", {}) or {}
        return mp.get(item_id)

    def on_toggle_show_matches(self):
        try:
            enabled = bool(self.show_matches_var.get())
        except Exception:
            enabled = False

        if enabled:
            if not (getattr(self, "code_to_name", {}) or {}):
                messagebox.showwarning("Coincidencias", "Primero debes cargar el Excel/CSV para usar 'Ver coincidencias'.")
                try:
                    self.show_matches_var.set(False)
                except Exception:
                    pass
                return

            try:
                if hasattr(self, "show_no_code_images_only_var"):
                    self.show_no_code_images_only_var.set(False)
            except Exception:
                pass

            try:
                idxs = self._get_nav_indices()
            except Exception:
                idxs = []

            if not idxs:
                messagebox.showinfo("Coincidencias", "No hay coincidencias para navegar con el filtro actual.")
                try:
                    self.show_matches_var.set(False)
                except Exception:
                    pass
                return

            try:
                cur = int(getattr(self, "current_image_index", -1))
            except Exception:
                cur = -1

            if cur not in idxs:
                self.current_image_index = idxs[0]
            self.show_current_image()
        else:
            try:
                self.show_current_image()
            except Exception:
                pass

    def _build_stem_from_product(self, product: dict) -> str:
        if not product:
            return ""
        mode = str(
            getattr(self, "rename_mode_var", None).get()
            if hasattr(self, "rename_mode_var")
            else "both"
        ).lower()
        if mode not in ("code", "name", "both"):
            mode = "both"

        code = str(product.get("code", "") or "").strip()
        name = str(product.get("name", "") or "").strip()

        if mode == "code":
            return code
        if mode == "name":
            return name
        if code and name:
            return f"{code}{FILENAME_PART_SEPARATOR}{name}"
        return code or name

    def _find_product_by_code_key(self, code_key: str):
        if not code_key:
            return None
        for p in (getattr(self, "products", []) or []):
            c = str(p.get("code", "") or "").strip()
            if not c:
                continue
            if sanitize_filename(c) == code_key:
                return p
        return None

    def on_rename_current_from_product(self):
        if not (getattr(self, "code_to_name", {}) or {}):
            messagebox.showerror("Renombrar", "Primero debes cargar el Excel/CSV para poder renombrar usando la tabla.")
            return
        product = self._get_focused_product_data()
        if not product:
            messagebox.showwarning(
                "Renombrar actual",
                "Selecciona un producto en la tabla (la línea del Excel) y luego pulsa 'Renombrar actual (Excel)'."
            )
            return
        if not getattr(self, "image_files", None) or not (0 <= getattr(self, "current_image_index", -1) < len(self.image_files)):
            messagebox.showerror("Renombrar actual", "No hay una imagen actual seleccionada para renombrar.")
            return

        new_stem = self._build_stem_from_product(product)
        if not str(new_stem or "").strip():
            messagebox.showerror(
                "Renombrar actual",
                "No se pudo construir el nuevo nombre con la fila seleccionada.\n"
                "Verifica que el producto tenga Código y/o Nombre, y el modo de renombrado elegido."
            )
            return

        product_code = str(product.get("code", "") or "").strip()
        product_name = str(product.get("name", "") or "").strip()

        self.rename_current_image_with_name(
            name=new_stem,
            origin="RENAME_CURRENT_FROM_SELECTED_ROW",
            product_code=product_code,
            product_name=product_name
        )

    def on_rename_all_from_product(self):
        if not (getattr(self, "code_to_name", {}) or {}):
            messagebox.showerror("Renombrar", "Primero debes cargar el Excel/CSV para poder renombrar por código.")
            return
        if not getattr(self, "image_folder", None):
            messagebox.showwarning("Renombrar", "Primero debes cargar la carpeta principal de imágenes.")
            return

        recursive = bool(getattr(self, "expand_scope_var", None).get() if hasattr(self, "expand_scope_var") else False)
        alcance = "carpeta principal + subcarpetas" if recursive else "solo carpeta principal"

        ok = messagebox.askyesno(
            "Renombrar todas",
            "Esto renombrará imágenes según el Excel, usando el código detectado en cada archivo.\n\n"
            f"Alcance: {alcance}\n"
            f"Modo: {self.rename_mode_var.get()}\n\n"
            "¿Deseas continuar?"
        )
        if not ok:
            return
        self.rename_all_images_using_excel(recursive=recursive)

    # ---- edición TreeView (NO venía completa en tu pegado; la incluyo funcional) ----
    def on_tree_click_store_column(self, event):
        try:
            self._cancel_tree_edit()
        except Exception:
            pass
        try:
            col = self.tree.identify_column(event.x)
            self._last_tree_click_col = col
        except Exception:
            self._last_tree_click_col = None

    def on_tree_edit_cancel_request(self, event=None):
        try:
            self._cancel_tree_edit()
        except Exception:
            pass

    def on_tree_edit_request(self, event=None):
        if not hasattr(self, "tree"):
            return
        item_id = self.tree.focus()
        if not item_id:
            return

        col_token = getattr(self, "_last_tree_click_col", None)
        if col_token not in ("#1", "#2"):
            col_token = "#2"
        col_name = "code" if col_token == "#1" else "name"
        if col_name not in ("code", "name"):
            return

        bbox = self.tree.bbox(item_id, col_name)
        if not bbox:
            return

        x, y, w, h = bbox
        old_val = self.tree.set(item_id, col_name)
        self._tree_editing = {"item_id": item_id, "col": col_name, "old": old_val}

        entry = getattr(self, "_tree_edit_entry", None)
        try:
            if entry is None or not entry.winfo_exists():
                entry = None
        except Exception:
            entry = None

        if entry is None:
            self._tree_edit_entry = ttk.Entry(self.tree)
            entry = self._tree_edit_entry

        entry.place(x=x, y=y, width=w, height=h)
        entry.delete(0, "end")
        entry.insert(0, str(old_val))
        entry.select_range(0, "end")
        entry.focus_set()

        entry.bind("<Return>", self._commit_tree_edit, add="+")
        entry.bind("<Escape>", lambda e: self._cancel_tree_edit(), add="+")
        entry.bind("<FocusOut>", self._commit_tree_edit, add="+")

    def _cancel_tree_edit(self):
        try:
            entry = getattr(self, "_tree_edit_entry", None)
            if entry is not None and entry.winfo_exists():
                entry.place_forget()
        except Exception:
            pass
        self._tree_editing = None

    def _commit_tree_edit(self, event=None):
        editing = getattr(self, "_tree_editing", None)
        entry = getattr(self, "_tree_edit_entry", None)
        if not editing or entry is None:
            return

        try:
            new_val = str(entry.get()).strip()
        except Exception:
            new_val = ""

        item_id = editing.get("item_id")
        col = editing.get("col")
        old_val = str(editing.get("old", ""))

        # cerrar editor visual
        self._cancel_tree_edit()

        if str(new_val) == str(old_val):
            return

        mp = getattr(self, "_tree_iid_to_product", {}) or {}
        product = mp.get(str(item_id))
        if not product:
            return

        row_number = product.get("row_number")
        if not row_number:
            messagebox.showerror("Editar", "No se pudo identificar la fila (row_number) para guardar el cambio.")
            return

        raw_path = (self.excel_path_var.get() or "").strip().strip('"')
        if not raw_path:
            messagebox.showerror("Editar", "No hay archivo Excel/CSV cargado.")
            return
        p = Path(raw_path)
        if not p.exists() or p.is_dir():
            messagebox.showerror("Editar", "La ruta del Excel/CSV no es válida.")
            return

        # valores actuales
        cur_code = str(product.get("code", "") or "")
        cur_name = str(product.get("name", "") or "")

        new_code = cur_code
        new_name = cur_name
        if col == "code":
            new_code = new_val
        elif col == "name":
            new_name = new_val
        else:
            return

        ok, err = update_product_in_file(
            path=str(p),
            row_number=int(row_number),
            code_col_index=int(getattr(self, "excel_code_col_index", 0)),
            name_col_index=int(getattr(self, "excel_name_col_index", 1)),
            sheet_name=getattr(self, "excel_sheet_name", "catalogo"),
            has_headers=True,
            header_row_number=int(getattr(self, "excel_header_row_number", 1)),
            new_code=new_code,
            new_name=new_name,
        )

        if not ok:
            messagebox.showerror("Editar", f"No se pudo guardar el cambio:\n\n{err}")
            return

        # actualizar memoria
        product["code"] = new_code
        product["name"] = new_name
        product["search_text"] = normalize_text(f"{new_code} {new_name} {product.get('category','')} {product.get('brand','')} {product.get('unit_price','')} {product.get('stock','')}")

        # reconstruir índices + refrescar vista
        self.code_to_name = build_code_to_name_index(getattr(self, "products", []) or [])
        self.apply_search_filter()

    # ---- fin edición TreeView ----


class RenameUIMixin:

    # =========================
    # NUEVO: Asignar códigos disponibles a imágenes sin código
    # =========================
    def _excel_numeric_code_stats(self):
        """
        Retorna (ok, err, excel_ints_set, min_int, max_int, width)
        width = longitud máxima encontrada en códigos numéricos (para zfill).
        """
        products = list(getattr(self, "products", []) or [])
        nums = []
        widths = []
        for p in products:
            raw_code = str(p.get("code", "") or "").strip()
            if not raw_code:
                continue
            if re.search(r"[A-Za-z]", raw_code):
                continue
            cleaned = raw_code.replace(" ", "").replace("\u00a0", "")
            digits_only = "".join(ch for ch in cleaned if ch.isdigit())
            if cleaned.isdigit():
                try:
                    nums.append(int(cleaned))
                    widths.append(len(digits_only) if digits_only else len(cleaned))
                except Exception:
                    pass
                continue
            parsed = _parse_number_for_excel(cleaned)
            if isinstance(parsed, (int, float)) and float(parsed).is_integer():
                try:
                    nums.append(int(parsed))
                    if digits_only:
                        widths.append(len(digits_only))
                    else:
                        widths.append(len(str(int(parsed))))
                except Exception:
                    pass

        if not nums:
            return (False, "No encontré Id numéricos en el Excel para calcular huecos.", set(), 0, 0, 0)

        min_int = min(nums)
        max_int = max(nums)
        width = max(widths) if widths else len(str(max_int))
        return (True, "", set(nums), int(min_int), int(max_int), int(width))

    def _format_code_with_width(self, n: int, width: int) -> str:
        try:
            if width and width > 1:
                return str(int(n)).zfill(int(width))
        except Exception:
            pass
        return str(int(n))

    def _iter_available_codes_in_range(self, min_int: int, max_int: int, occupied_ints: set, need: int):
        """
        Genera códigos disponibles (enteros) en [min_int, max_int] excluyendo occupied_ints,
        sin iterar todo el rango completo; se detiene al alcanzar 'need'.
        """
        if need <= 0:
            return []

        # Ordena ocupados dentro del rango
        occ = sorted([x for x in occupied_ints if isinstance(x, int) and min_int <= x <= max_int])

        out = []
        cur = int(min_int)

        for x in occ:
            if len(out) >= need:
                break
            if x < cur:
                continue
            while cur < x and len(out) < need:
                out.append(cur)
                cur += 1
            cur = x + 1

        while cur <= max_int and len(out) < need:
            if cur not in occupied_ints:
                out.append(cur)
            cur += 1

        return out

    def _iter_available_codes_with_fallback(self, min_int: int, max_int: int, occupied_ints: set, need: int):
        """
        Genera códigos disponibles primero dentro de [min_int, max_int] y, si no hay
        suficientes, continúa desde (max_int + 1) hacia arriba evitando ocupados.
        """
        if need <= 0:
            return []

        out = self._iter_available_codes_in_range(min_int, max_int, occupied_ints, need)

        cur = int(max_int) + 1
        while len(out) < need:
            if cur not in occupied_ints:
                out.append(cur)
            cur += 1

        return out

    def _extract_name_from_stem_after_code(self, stem: str, code_key: str) -> str:
        """
        Intenta obtener nombre del producto desde el nombre del archivo.
        Ej: '12345_Shampoo_Azul' -> 'Shampoo Azul'
        """
        if not stem:
            return ""

        sep = FILENAME_PART_SEPARATOR or "_"
        raw = str(stem)

        # intentar con separador principal
        if sep and sep in raw:
            first, rest = raw.split(sep, 1)
            if sanitize_filename(first) == str(code_key):
                name = rest
            else:
                name = ""
        else:
            name = ""

        # fallback: si no salió, intenta con '_' (común)
        if not name and "_" in raw:
            first, rest = raw.split("_", 1)
            if sanitize_filename(first) == str(code_key):
                name = rest

        name = (name or "").replace("_", " ").replace("-", " ").strip()
        return name

    def _scan_used_code_ints_in_images(self, recursive: bool):
        used = set()
        root = getattr(self, "image_folder", None)
        if not root or not hasattr(root, "is_dir") or not root.is_dir():
            return used

        files = []
        try:
            if hasattr(self, "_iter_scope_image_files"):
                files = self._iter_scope_image_files(recursive=recursive)
        except Exception:
            files = []

        for p in files:
            try:
                ck = extract_code_key_from_stem(p.stem)
            except Exception:
                ck = ""
            if not ck or not looks_like_product_code(ck):
                continue
            if str(ck).isdigit():
                try:
                    used.add(int(str(ck)))
                except Exception:
                    pass
        return used

    def on_assign_codes_to_images_without_code(self):
        # Reglas: solo a imágenes cuyo nombre NO inicia con un código válido.
        if not (getattr(self, "products", []) or []):
            messagebox.showerror("Códigos", "Primero carga el Excel/CSV para conocer el rango de Id.")
            return
        if not getattr(self, "image_folder", None):
            messagebox.showwarning("Códigos", "Primero carga la carpeta principal de imágenes.")
            return

        recursive = bool(getattr(self, "expand_scope_var", None).get() if hasattr(self, "expand_scope_var") else False)
        alcance = "carpeta principal + subcarpetas" if recursive else "solo carpeta principal"

        ok_stats, err_stats, excel_ints, min_int, max_int, width = self._excel_numeric_code_stats()
        if not ok_stats:
            messagebox.showerror("Códigos", err_stats)
            return

        # Detectar imágenes sin código válido al inicio
        scope_files = []
        try:
            scope_files = self._iter_scope_image_files(recursive=recursive)
        except Exception:
            scope_files = []

        targets = []
        for path in scope_files:
            try:
                ck = extract_code_key_from_stem(path.stem)
            except Exception:
                ck = ""
            if not looks_like_product_code(ck):
                targets.append(path)

        if not targets:
            messagebox.showinfo("Códigos", "No encontré imágenes sin código válido para asignar.")
            return

        # Ocupados: lo que ya existe en Excel + lo que ya usan imágenes con código numérico
        used_images_ints = self._scan_used_code_ints_in_images(recursive=recursive)
        occupied = set(excel_ints) | set(used_images_ints)

        avail = self._iter_available_codes_with_fallback(min_int, max_int, occupied, need=len(targets))

        if len(avail) < len(targets):
            msg_extra = (
                f"Imágenes sin código: {len(targets)}\n"
                f"Códigos disponibles: {len(avail)}\n\n"
                "Se asignarán códigos solo a la cantidad disponible."
            )
        else:
            uses_after_max = any(c > max_int for c in avail)
            if uses_after_max:
                msg_extra = (
                    f"Imágenes sin código detectadas: {len(targets)}\n"
                    "No hay huecos en el rango del Excel; se continuará desde el mayor Id."
                )
            else:
                msg_extra = f"Imágenes sin código detectadas: {len(targets)}"

        go = messagebox.askyesno(
            "Asignar códigos",
            "Esto RENOMBRARÁ archivos de imagen agregando un Id al inicio.\n\n"
            f"Alcance: {alcance}\n"
            f"Rango Excel: {min_int} .. {max_int}\n\n"
            f"{msg_extra}\n\n"
            "¿Deseas continuar?"
        )
        if not go:
            return

        assigned = 0
        errors = 0

        # Renombrar, asignando en orden ascendente
        for i, img_path in enumerate(targets[:len(avail)]):
            try:
                code_int = avail[i]
                code_str = self._format_code_with_width(code_int, width)

                # Nuevo nombre: CODIGO + _ + NOMBRE_ACTUAL
                new_stem_raw = f"{code_str}{FILENAME_PART_SEPARATOR}{img_path.stem}"
                new_stem = self._clean_stem(new_stem_raw) if hasattr(self, "_clean_stem") else sanitize_filename(new_stem_raw)

                new_path = self._unique_path_same_folder(img_path, new_stem) if hasattr(self, "_unique_path_same_folder") else img_path.with_name(new_stem + img_path.suffix)

                if new_path == img_path:
                    continue

                img_path.rename(new_path)
                assigned += 1
            except Exception:
                errors += 1

        # refrescar
        try:
            self._rebuild_indexes_and_refresh(focus_path=None)
        except Exception:
            try:
                self.reload_image_files(keep_current=False)
                self.show_current_image()
            except Exception:
                pass

        try:
            self.save_config()
        except Exception:
            pass

        messagebox.showinfo(
            "Asignar códigos",
            "Proceso terminado.\n\n"
            f"Asignadas: {assigned}\n"
            f"Errores: {errors}\n\n"
            "Nota: esto NO agrega filas al Excel. Si quieres crear los productos nuevos,\n"
            "usa 'Agregar productos faltantes al Excel'."
        )

    # =========================
    # NUEVO: Agregar productos faltantes al Excel desde imágenes
    # =========================
    def on_add_missing_products_to_excel(self):
        raw_excel = (getattr(self, "excel_path_var", None).get() if hasattr(self, "excel_path_var") else "").strip().strip('"')
        if not raw_excel:
            messagebox.showerror("Excel", "Primero carga el Excel/CSV.")
            return
        p = Path(raw_excel)
        if not p.exists() or p.is_dir():
            messagebox.showerror("Excel", "La ruta del Excel/CSV no es válida.")
            return

        root = getattr(self, "image_folder", None)
        if not root or not hasattr(root, "is_dir") or not root.is_dir():
            messagebox.showwarning("Excel", "Primero carga la carpeta principal de imágenes.")
            return

        recursive = bool(getattr(self, "expand_scope_var", None).get() if hasattr(self, "expand_scope_var") else False)
        alcance = "carpeta principal + subcarpetas" if recursive else "solo carpeta principal"

        excel_codes = set((getattr(self, "code_to_name", {}) or {}).keys())

        # Escanear imágenes y detectar códigos válidos que NO estén en Excel
        scope_files = []
        try:
            scope_files = self._iter_scope_image_files(recursive=recursive)
        except Exception:
            scope_files = []

        missing = {}  # code_key -> name
        for img_path in scope_files:
            try:
                ck = extract_code_key_from_stem(img_path.stem)
            except Exception:
                ck = ""
            if not ck or not looks_like_product_code(ck):
                continue

            code_key = str(ck)
            if sanitize_filename(code_key) in excel_codes:
                continue

            name = self._extract_name_from_stem_after_code(img_path.stem, sanitize_filename(code_key))
            missing[sanitize_filename(code_key)] = name

        if not missing:
            messagebox.showinfo("Excel", "No encontré productos faltantes en el Excel (según códigos de imágenes).")
            return

        # Orden: numérico si es posible
        def _sort_key(code_key: str):
            s = str(code_key)
            return (0, int(s)) if s.isdigit() else (1, s)

        missing_codes_sorted = sorted(missing.keys(), key=_sort_key)

        # Preparar filas a agregar (Id / Nombre producto)
        rows_to_add = []
        for code_key in missing_codes_sorted:
            rows_to_add.append({
                "code": code_key,
                "name": str(missing.get(code_key, "") or "").strip()
            })

        go = messagebox.askyesno(
            "Agregar productos",
            "Esto ESCRIBIRÁ nuevas filas en tu Excel/CSV (Id / Nombre producto).\n"
            "Asegúrate de tener el archivo cerrado en Excel.\n\n"
            f"Alcance: {alcance}\n"
            f"Filas nuevas a agregar: {len(rows_to_add)}\n\n"
            "¿Deseas continuar?"
        )
        if not go:
            return

        ok, err, added = append_products_to_file(
            path=str(p),
            products_to_add=rows_to_add,
            code_col_index=int(getattr(self, "excel_code_col_index", 0)),
            name_col_index=int(getattr(self, "excel_name_col_index", 1)),
            sheet_name=getattr(self, "excel_sheet_name", "catalogo"),
            has_headers=True,
            header_row_number=int(getattr(self, "excel_header_row_number", 1)),
        )

        if not ok:
            messagebox.showerror("Agregar productos", f"No se pudo agregar:\n\n{err}")
            return

        # Recargar Excel y refrescar vista
        try:
            self.load_excel(str(p), interactive=False)
        except Exception:
            pass
        try:
            self.apply_search_filter()
        except Exception:
            pass
        try:
            self.refresh_product_list()
        except Exception:
            pass
        try:
            self.save_config()
        except Exception:
            pass

        messagebox.showinfo(
            "Agregar productos",
            "Listo.\n\n"
            f"Filas agregadas: {added}\n\n"
            "Si el archivo estaba abierto en Excel, ciérralo y vuelve a intentar."
        )

    
    def update_undo_button_state(self):
        if hasattr(self, "btn_undo"):
            if self.last_rename_action:
                self.btn_undo.state(["!disabled"])
            else:
                self.btn_undo.state(["disabled"])

    def _rebuild_indexes_and_refresh(self, focus_path: Path = None):
        try:
            if hasattr(self, "reload_image_files"):
                self.reload_image_files(keep_current=False)
        except Exception:
            pass

        if focus_path is not None:
            try:
                if focus_path in getattr(self, "image_files", []):
                    self.current_image_index = self.image_files.index(focus_path)
                else:
                    self.current_image_index = 0 if getattr(self, "image_files", []) else -1
            except Exception:
                self.current_image_index = 0 if getattr(self, "image_files", []) else -1

        try:
            self.show_current_image()
        except Exception:
            pass

        try:
            if hasattr(self, "refresh_product_list"):
                self.refresh_product_list()
        except Exception:
            pass

    def on_undo(self):
        action = self.last_rename_action
        if not action:
            messagebox.showinfo("Deshacer", "No hay cambios para deshacer.")
            return

        old_path = action["old_path"]
        new_path = action["new_path"]

        if not new_path.exists():
            messagebox.showerror("Deshacer", "No se encontró el archivo renombrado para deshacer el cambio.")
            self.last_rename_action = None
            self.update_undo_button_state()
            return

        try:
            new_path.rename(old_path)
        except Exception as e:
            messagebox.showerror("Deshacer", f"No se pudo deshacer el cambio:\n{e}")
            return

        self.last_rename_action = None
        self.update_undo_button_state()
        self._rebuild_indexes_and_refresh(focus_path=old_path)

        try:
            self.save_config()
        except Exception:
            pass

    def on_statusbar_rename_reset(self, event=None):
        if getattr(self, "_get_view_mode", lambda: "single")() != "single":
            try:
                self.status_rename_var.set("")
            except Exception:
                pass
            return

        if not getattr(self, "image_files", None) or not (0 <= getattr(self, "current_image_index", -1) < len(self.image_files)):
            try:
                self.status_rename_var.set("")
            except Exception:
                pass
            return

        try:
            p = self.image_files[self.current_image_index]
            self.status_rename_var.set(p.stem)
        except Exception:
            pass

    def on_statusbar_rename_apply(self, event=None):
        if getattr(self, "_get_view_mode", lambda: "single")() != "single":
            messagebox.showinfo("Renombrar", "En la vista de coincidencias no se puede renombrar. Abre una imagen individual.")
            return

        if not getattr(self, "image_files", None) or not (0 <= getattr(self, "current_image_index", -1) < len(self.image_files)):
            messagebox.showwarning("Renombrar", "No hay una imagen actual seleccionada.")
            return

        try:
            raw = str(self.status_rename_var.get()).strip()
        except Exception:
            raw = ""

        if not raw:
            messagebox.showwarning("Renombrar", "Escribe un nombre en la barra de estado.")
            return

        current_path = self.image_files[self.current_image_index]

        proposed = raw
        try:
            pp = Path(raw)
            if pp.suffix and (pp.suffix.lower() == current_path.suffix.lower() or pp.suffix.lower() in IMAGE_EXTENSIONS):
                proposed = pp.stem
        except Exception:
            proposed = raw

        self.rename_current_image_with_name(name=proposed, origin="STATUS_BAR")

    def _clean_stem(self, text: str) -> str:
        safe = sanitize_filename(text)
        safe = re.sub(r"_+", "_", safe).strip("_")
        return safe

    def _unique_path_same_folder(self, base_path: Path, new_stem: str) -> Path:
        ext = base_path.suffix
        candidate = base_path.with_name(f"{new_stem}{ext}")
        if candidate == base_path:
            return candidate
        if not candidate.exists():
            return candidate
        counter = 1
        while True:
            cand = base_path.with_name(f"{new_stem}_{counter}{ext}")
            if not cand.exists() and cand != base_path:
                return cand
            counter += 1

    def rename_current_image_with_name(self, name: str, origin: str = "MANUAL", product_code: str = "", product_name: str = ""):
        if not self.image_files or not (0 <= self.current_image_index < len(self.image_files)):
            messagebox.showwarning("Renombrar", "No hay imagen cargada para renombrar.")
            return

        current_path = self.image_files[self.current_image_index]
        if current_path.suffix.lower() not in IMAGE_EXTENSIONS:
            messagebox.showwarning("Renombrar", "El archivo actual no es una imagen soportada.")
            return

        old_path = current_path
        new_stem = self._clean_stem(name)
        if not new_stem:
            messagebox.showwarning("Renombrar", "El nombre propuesto no es válido.")
            return

        new_path = self._unique_path_same_folder(old_path, new_stem)
        if new_path == old_path:
            return

        try:
            old_path.rename(new_path)
        except Exception as e:
            messagebox.showerror("Renombrar", f"No se pudo renombrar el archivo:\n{e}")
            return

        try:
            self.image_files[self.current_image_index] = new_path
        except Exception:
            pass

        self.last_rename_action = {
            "old_path": old_path,
            "new_path": new_path,
            "product_code": product_code,
            "product_name": product_name,
        }
        self.update_undo_button_state()

        self._rebuild_indexes_and_refresh(focus_path=new_path)

        try:
            self.save_config()
        except Exception:
            pass

    def _iter_scope_image_files(self, recursive: bool):
        root = getattr(self, "image_folder", None)
        if not root or not hasattr(root, "is_dir") or not root.is_dir():
            return []
        files = []
        try:
            if recursive:
                for p in root.rglob("*"):
                    if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS:
                        files.append(p)
            else:
                for p in root.iterdir():
                    if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS:
                        files.append(p)
        except Exception:
            files = []
        return sorted(files, key=lambda x: x.name.lower())

    def rename_all_images_using_excel(self, recursive: bool = False):
        if not (getattr(self, "code_to_name", {}) or {}):
            messagebox.showerror("Renombrar", "No hay Excel cargado (no existe el índice code_to_name).")
            return

        root = getattr(self, "image_folder", None)
        if not root or not hasattr(root, "is_dir") or not root.is_dir():
            messagebox.showwarning("Renombrar", "Primero debes cargar la carpeta principal de imágenes.")
            return

        scope_files = self._iter_scope_image_files(recursive=recursive)
        if not scope_files:
            messagebox.showinfo("Renombrar", "No se encontraron imágenes en el alcance seleccionado.")
            return

        current_path = None
        if getattr(self, "image_files", None) and 0 <= getattr(self, "current_image_index", -1) < len(self.image_files):
            try:
                current_path = self.image_files[self.current_image_index]
            except Exception:
                current_path = None

        renamed = 0
        skipped_no_code = 0
        skipped_not_in_excel = 0
        skipped_no_name_built = 0
        errors = 0

        code_to_name = getattr(self, "code_to_name", {}) or {}
        for path in scope_files:
            try:
                code_key = extract_code_key_from_stem(path.stem)
            except Exception:
                code_key = ""

            if not code_key:
                skipped_no_code += 1
                continue

            if code_key not in code_to_name:
                skipped_not_in_excel += 1
                continue

            product = None
            try:
                if hasattr(self, "_find_product_by_code_key"):
                    product = self._find_product_by_code_key(code_key)
            except Exception:
                product = None

            if not product:
                product = {"code": code_key, "name": code_to_name.get(code_key, "")}

            try:
                new_stem_raw = self._build_stem_from_product(product) if hasattr(self, "_build_stem_from_product") else ""
                new_stem = self._clean_stem(new_stem_raw)
                if not new_stem:
                    skipped_no_name_built += 1
                    continue

                new_path = self._unique_path_same_folder(path, new_stem)
                if new_path == path:
                    continue

                path.rename(new_path)
                if current_path == path:
                    current_path = new_path
                renamed += 1
            except Exception:
                errors += 1

        try:
            self.reload_image_files(keep_current=False)
            if current_path and current_path in getattr(self, "image_files", []):
                self.current_image_index = self.image_files.index(current_path)
        except Exception:
            pass

        try:
            self.show_current_image()
        except Exception:
            pass
        try:
            self.refresh_product_list()
        except Exception:
            pass

        self.last_rename_action = None
        self.update_undo_button_state()

        try:
            self.save_config()
        except Exception:
            pass

        alcance = "carpeta principal + subcarpetas" if recursive else "solo carpeta principal"
        msg = (
            f"Alcance: {alcance}\n\n"
            f"Renombradas: {renamed}\n"
            f"Sin código detectable: {skipped_no_code}\n"
            f"Código no existe en Excel: {skipped_not_in_excel}\n"
            f"Sin nombre válido a construir: {skipped_no_name_built}\n"
            f"Errores: {errors}"
        )
        messagebox.showinfo("Renombrar todas", msg)


# =========================
# APP (antes app.py + main.py)
# =========================

class ImageRenamerApp(
    tk.Tk,
    WindowMixin,
    ProductsUIMixin,
    FilesUIMixin,
    RenameUIMixin,
):
    def __init__(self):
        tk.Tk.__init__(self)

        self.title("Renombrador de Imágenes por Código")
        self.geometry("1100x700")
        self.minsize(900, 600)

        self.initial_window_geometry = None
        self._config_last_image_path = None
        self._config_tree_columns = None

        self.products = []
        self.filtered_products = []
        self.code_to_name = {}

        self.excel_sheet_name = "catalogo"
        self.excel_has_headers = True
        self.excel_header_row_number = 2
        self.excel_code_col_index = 0
        self.excel_name_col_index = 1
        self.excel_category_col_index = -1
        self.excel_brand_col_index = -1
        self.excel_unit_price_col_index = -1
        self.excel_stock_col_index = -1

        self.image_folder = None
        self.all_image_files = []
        self.image_files = []
        self.current_image_index = -1
        self.current_image = None
        self.current_photo = None

        self.code_to_folder = {}
        self.code_to_image_path = {}

        self.excel_path_var = tk.StringVar()
        self.image_folder_path_var = tk.StringVar()
        self.search_var = tk.StringVar()

        # ====== NUEVO: variable UI para fila de encabezados ======
        self.excel_header_row_var = tk.IntVar(value=int(self.excel_header_row_number))
        # =========================================================

        self.show_no_code_images_only_var = tk.BooleanVar(value=False)
        self.show_matches_var = tk.BooleanVar(value=False)
        self.rename_mode_var = tk.StringVar(value="both")
        self.expand_scope_var = tk.BooleanVar(value=False)

        self.status_filename_var = tk.StringVar(value="")
        self.status_rename_var = tk.StringVar(value="")

        self.last_rename_action = None

        self._search_after_id = None
        self._search_delay_ms = 260

        self._create_style()
        self._create_menu()
        self._create_widgets()

        self.search_var.trace_add("write", self._on_search_var_write)

        self.load_config_and_autoload()

        self.protocol("WM_DELETE_WINDOW", self.on_app_close)
        self.after(0, self.apply_initial_window_state)
        self.after(100, self.set_paned_half)

    def _on_search_var_write(self, *args):
        after_id = getattr(self, "_search_after_id", None)
        if after_id is not None:
            try:
                self.after_cancel(after_id)
            except Exception:
                pass
        delay = getattr(self, "_search_delay_ms", 260)
        self._search_after_id = self.after(delay, self.apply_search_filter)

    def load_config_and_autoload(self):
        if not CONFIG_FILE.exists():
            return
        try:
            with CONFIG_FILE.open("r", encoding="utf-8") as f:
                data = json.load(f)

            excel_path = data.get("excel_path", "")
            image_folder = data.get("image_folder", "")

            self.initial_window_geometry = data.get("window_geometry")
            tree_cols = data.get("tree_columns", {})
            if isinstance(tree_cols, dict):
                self._config_tree_columns = tree_cols

            self.search_var.set(data.get("search_text", ""))

            mode = str(data.get("rename_mode", "both")).strip().lower()
            if mode not in ("code", "name", "both"):
                mode = "both"
            self.rename_mode_var.set(mode)

            self.expand_scope_var.set(bool(data.get("expand_scope", False)))
            self.show_no_code_images_only_var.set(bool(data.get("show_no_code_images_only", False)))
            self.show_matches_var.set(bool(data.get("show_matches", False)))

            # Mantener hoja fija como antes
            self.excel_sheet_name = "catalogo"
            self.excel_has_headers = True

            # ====== NUEVO: cargar fila encabezados desde config ======
            hdr = data.get("excel_header_row", 2)
            try:
                hdr = int(hdr)
            except Exception:
                hdr = 2
            if hdr < 1:
                hdr = 1
            self.excel_header_row_number = int(hdr)
            try:
                self.excel_header_row_var.set(int(hdr))
            except Exception:
                pass
            # ========================================================

            last_image = data.get("current_image", "")
            if last_image:
                try:
                    self._config_last_image_path = Path(last_image)
                except Exception:
                    self._config_last_image_path = None

            if excel_path:
                self.excel_path_var.set(excel_path)
                if Path(excel_path).exists():
                    self.load_excel(excel_path, interactive=False)

            if image_folder:
                self.image_folder_path_var.set(image_folder)
                folder_path = Path(image_folder)
                if folder_path.is_dir():
                    self.set_image_folder(folder_path)

            try:
                if bool(self.show_matches_var.get()):
                    self.on_toggle_show_matches()
            except Exception:
                pass

        except Exception as e:
            print("No se pudo leer config:", e)

    def save_config(self):
        data = {
            "excel_path": self.excel_path_var.get().strip(),
            "image_folder": self.image_folder_path_var.get().strip(),
        }
        try:
            self.update_idletasks()
            data["window_geometry"] = self.geometry()

            if hasattr(self, "tree"):
                cols = {}
                try:
                    for col in self.tree["columns"]:
                        try:
                            cols[col] = self.tree.column(col, "width")
                        except Exception:
                            pass
                except Exception:
                    pass
                data["tree_columns"] = cols

            data["search_text"] = self.search_var.get()

            mode = str(self.rename_mode_var.get()).strip().lower()
            if mode not in ("code", "name", "both"):
                mode = "both"
            data["rename_mode"] = mode

            data["expand_scope"] = bool(self.expand_scope_var.get())
            data["show_no_code_images_only"] = bool(self.show_no_code_images_only_var.get())
            data["show_matches"] = bool(self.show_matches_var.get())

            # ====== NUEVO: guardar fila de encabezados ======
            try:
                data["excel_header_row"] = int(self.excel_header_row_var.get())
            except Exception:
                data["excel_header_row"] = int(getattr(self, "excel_header_row_number", 2))
            # ==============================================

            if self.image_files and 0 <= self.current_image_index < len(self.image_files):
                data["current_image"] = str(self.image_files[self.current_image_index])
            else:
                data["current_image"] = ""

        except Exception as e:
            print("No se pudo recolectar estado:", e)

        try:
            with CONFIG_FILE.open("w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print("No se pudo guardar config:", e)

    def apply_initial_window_state(self):
        try:
            if self.initial_window_geometry:
                self.geometry(self.initial_window_geometry)
            self.update_idletasks()

            try:
                self.state("zoomed")
            except tk.TclError:
                try:
                    self.attributes("-zoomed", True)
                except tk.TclError:
                    w = self.winfo_screenwidth()
                    h = self.winfo_screenheight()
                    self.geometry(f"{w}x{h}+0+0")
        except Exception as e:
            print("No se pudo maximizar la ventana:", e)

    def set_paned_half(self):
        try:
            if not hasattr(self, "paned"):
                return
            self.update_idletasks()
            total_width = self.paned.winfo_width()
            if total_width > 0:
                self.paned.sashpos(0, total_width // 2)
        except Exception as e:
            print("No se pudo ajustar el panel:", e)

    def on_app_close(self):
        try:
            if hasattr(self, "_cancel_tree_edit"):
                self._cancel_tree_edit()
        except Exception:
            pass
        self.save_config()
        self.destroy()


if __name__ == "__main__":
    app = ImageRenamerApp()
    app.mainloop()
